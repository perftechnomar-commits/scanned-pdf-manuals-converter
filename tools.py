from __future__ import annotations

import base64
import io
import json
import os
import re
import tempfile
import time
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence
from difflib import SequenceMatcher
from datetime import datetime, timezone

import pandas as pd
import requests
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from pypdf.generic import RectangleObject


TOOLS_VERSION = "4.18.12"

MACHINERY_SHEET = "1.Machineries|Sub|Units"
SPARE_PARTS_SHEET = "2.Spare Parts"

MACHINERY_COLUMNS = [
    "CODE",
    "NAME",
    "MAKER",
    "MODEL",
    "TYPE",
    "INSTR.BOOK",
    "SPECIFICATIONS",
    "MCH_TP(M/S/U)",
]

BENEFIT_SPARE_COLUMNS = [
    "MACHINERY",
    "PART NO",
    "DESCRIPTION",
    "CODE",
    "ITEM NO",
    "UNIT",
    "QNT",
]

REVIEW_COLUMNS = [
    "INCLUDE",
    "READY",
    "MACHINERY",
    "PART NO",
    "DESCRIPTION",
    "CODE",
    "ITEM NO",
    "UNIT",
    "QNT",
    "SOURCE PAGE",
    "SECTION START PAGE",
    "TABLE TITLE",
    "SECTION CODE",
    "SECTION MAKER",
    "SECTION MODEL",
    "CONFIDENCE",
    "DETECTED MACHINERY",
    "ASSIGNMENT SOURCE",
    "WARNING",
]

SUBMACHINERY_REVIEW_COLUMNS = [
    "INCLUDE",
    "CODE",
    "NAME",
    "MAKER",
    "MODEL",
    "TYPE",
    "INSTR.BOOK",
    "SPECIFICATIONS",
    "MCH_TP(M/S/U)",
    "FIRST PAGE",
    "LAST PAGE",
    "PARTS FOUND",
    "CONFIDENCE",
    "VARIANTS",
    "DETECTION KEYS",
    "ORIGIN",
]

MACHINERY_TYPES = ["Main Machinery", "SubMachinery"]
UNIT_OPTIONS = ["", "PCS", "SET"]

MAX_MACHINERY_ROWS = 605  # B5:B609 is the template's named machinery range.
MAX_SPARE_ROWS = 1438  # Rows 4:1441 in the spare-parts import sheet.

ProgressCallback = Callable[[int, int, str], None]

PAGE_CLASSIFICATION_COLUMNS = [
    "SOURCE PAGE",
    "CLASSIFICATION",
    "PROCESS",
    "SCORE",
    "REASON",
    "CHARACTERS",
]

PAGE_FILTER_MODES = [
    "Conservative (recommended)",
    "Strict",
    "Off",
]


# ---------------------------------------------------------------------------
# General cleanup and validation helpers
# ---------------------------------------------------------------------------


def clean_text(value: Any) -> str:
    """Return a compact, single-line string suitable for review and Excel."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value)
    text = text.replace("\x00", " ")
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"[`*_]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_text(value).upper())


def excel_safe_text(value: Any) -> str | None:
    """Protect text cells from accidental formula execution while preserving display."""
    text = clean_text(value)
    if not text:
        return None
    if text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def quantity_to_number(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        number = float(value)
    else:
        text = clean_text(value).replace(",", ".")
        match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        number = float(match.group(0))

    if number.is_integer():
        return int(number)
    return number


def normalize_unit(value: Any, default_unit: str = "PCS") -> str:
    text = clean_text(value).upper()
    if "SET" in text:
        return "SET"
    if any(token in text for token in ("PCS", "PC", "PIECE", "EA", "EACH", "NO.")):
        return "PCS"
    return default_unit if default_unit in UNIT_OPTIONS else ""


def clamp_confidence(value: Any, fallback: float = 0.70) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = fallback
    return max(0.0, min(1.0, number))


def empty_review_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=REVIEW_COLUMNS)


def empty_additional_machinery_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=MACHINERY_COLUMNS)


def empty_submachinery_review_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=SUBMACHINERY_REVIEW_COLUMNS)


# ---------------------------------------------------------------------------
# PDF selection and OCR
# ---------------------------------------------------------------------------


def parse_page_spec(spec: str, total_pages: int) -> list[int]:
    """
    Parse an inclusive, one-based page expression such as ``1-20,25,30-35``.

    Returns zero-based page indexes. Blank or ``all`` selects every page.
    """
    if total_pages < 1:
        return []

    text = clean_text(spec).lower()
    if not text or text == "all":
        return list(range(total_pages))

    selected: list[int] = []
    seen: set[int] = set()

    for token in text.split(","):
        token = token.strip()
        if not token:
            continue

        if "-" in token:
            left, right = token.split("-", 1)
            start = int(left) if left.strip() else 1
            end = int(right) if right.strip() else total_pages
            if start > end:
                start, end = end, start
            numbers = range(start, end + 1)
        else:
            numbers = [int(token)]

        for page_number in numbers:
            if not 1 <= page_number <= total_pages:
                raise ValueError(
                    f"Page {page_number} is outside this PDF's range of 1-{total_pages}."
                )
            index = page_number - 1
            if index not in seen:
                selected.append(index)
                seen.add(index)

    if not selected:
        raise ValueError("The page selection did not contain any valid pages.")
    return selected


def pdf_page_count(pdf_bytes: bytes) -> int:
    return len(PdfReader(io.BytesIO(pdf_bytes)).pages)


def chunks(values: Sequence[int], chunk_size: int) -> Iterable[list[int]]:
    chunk_size = max(1, int(chunk_size))
    for start in range(0, len(values), chunk_size):
        yield list(values[start : start + chunk_size])


def _temporary_file(data: bytes, suffix: str) -> str:
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        handle.write(data)
        handle.flush()
        return handle.name
    finally:
        handle.close()


def _delete_file(path: str | None) -> None:
    if not path:
        return
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


def _normalize_api_key(api_key: str) -> str:
    """Normalize a Mistral key without ever logging or returning it in an error."""
    key = str(api_key or "").strip().strip('"').strip("'").strip()
    if key.lower().startswith("bearer "):
        key = key[7:].strip()
    if not key:
        raise RuntimeError("A Mistral API key is required.")
    return key


def _safe_api_error_text(value: Any, api_key: str = "") -> str:
    """Return a concise error message with credentials and very long bodies removed."""
    text = str(value or "").replace("\x00", " ")
    if api_key:
        text = text.replace(api_key, "***REDACTED***")
    # Some upstream libraries include the key after labels such as api_key= or Bearer.
    text = re.sub(
        r"(?i)(api[_ -]?key\s*[:=]\s*|authorization\s*[:=]\s*bearer\s+)([^\s,;\]\[{}]+)",
        r"\1***REDACTED***",
        text,
    )
    text = re.sub(r"\s+", " ", text).strip()
    return text[:1200]


def _mistral_ocr_request(
    api_key: str,
    document: dict[str, str],
    model: str = "mistral-ocr-latest",
    timeout_seconds: int = 600,
    max_retries: int = 5,
) -> dict[str, Any]:
    """Call Mistral's official OCR REST endpoint directly.

    This intentionally bypasses ``py-mistral-helper`` and its transitive SDK
    dependencies. PDF and image bytes are sent as supported base64 data URLs,
    while external sources are passed through as ordinary URLs.
    """
    key = _normalize_api_key(api_key)
    endpoint = os.getenv("MISTRAL_OCR_URL", "https://api.mistral.ai/v1/ocr")
    payload = {
        "model": model,
        "document": document,
        "include_image_base64": False,
    }
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    retryable_statuses = {404, 408, 409, 425, 429, 500, 502, 503, 504}
    last_error: Exception | None = None

    for attempt in range(1, max(1, int(max_retries)) + 1):
        try:
            response = requests.post(
                endpoint,
                headers=headers,
                json=payload,
                timeout=max(30, int(timeout_seconds)),
            )

            if 200 <= response.status_code < 300:
                try:
                    body = response.json()
                except ValueError as exc:
                    raise RuntimeError("Mistral OCR returned non-JSON output.") from exc
                if not isinstance(body, dict):
                    raise RuntimeError("Mistral OCR returned an unexpected response type.")
                return body

            try:
                body_value: Any = response.json()
            except ValueError:
                body_value = response.text
            detail = _safe_api_error_text(body_value, key)

            if response.status_code in retryable_statuses and attempt < max_retries:
                retry_after = response.headers.get("Retry-After", "")
                try:
                    delay = max(1.0, float(retry_after))
                except (TypeError, ValueError):
                    delay = min(20.0, float(2 ** (attempt - 1)))
                time.sleep(delay)
                continue

            if response.status_code in {401, 403}:
                raise RuntimeError(
                    f"Mistral OCR authentication failed (HTTP {response.status_code}). "
                    "Verify the key stored in Streamlit Secrets and its workspace access. "
                    f"Service response: {detail or 'No additional details.'}"
                )
            if response.status_code == 402:
                raise RuntimeError(
                    "Mistral OCR rejected the request because billing or workspace "
                    f"payment is not enabled (HTTP 402). Service response: {detail}"
                )
            if response.status_code == 429:
                raise RuntimeError(
                    "Mistral OCR rate or usage limit reached (HTTP 429). Wait and retry, "
                    f"or check the workspace Limits and Usage pages. Service response: {detail}"
                )
            raise RuntimeError(
                f"Mistral OCR request failed (HTTP {response.status_code}). "
                f"Service response: {detail or 'No additional details.'}"
            )
        except requests.Timeout as exc:
            last_error = exc
            if attempt < max_retries:
                time.sleep(min(20.0, float(2 ** (attempt - 1))))
                continue
        except requests.RequestException as exc:
            last_error = exc
            if attempt < max_retries:
                time.sleep(min(20.0, float(2 ** (attempt - 1))))
                continue
        except RuntimeError:
            raise

    raise RuntimeError(
        "Mistral OCR could not be reached after repeated attempts: "
        + _safe_api_error_text(last_error, key)
    )


def _pdf_data_url(pdf_bytes: bytes) -> str:
    encoded = base64.b64encode(pdf_bytes).decode("ascii")
    return f"data:application/pdf;base64,{encoded}"


def _image_data_url(image_bytes: bytes, suffix: str) -> str:
    extension = str(suffix or "").lower().lstrip(".")
    mime_type = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "webp": "image/webp",
        "gif": "image/gif",
    }.get(extension, "image/png")
    encoded = base64.b64encode(image_bytes).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def _response_pages(response: dict[str, Any]) -> list[dict[str, Any]]:
    pages = response.get("pages", []) if isinstance(response, dict) else []
    if pages is None:
        return []
    if not isinstance(pages, list):
        raise RuntimeError("Mistral OCR returned an invalid pages collection.")
    return [page for page in pages if isinstance(page, dict)]


def ocr_pdf_bytes(
    api_key: str,
    pdf_bytes: bytes,
    page_indexes: Sequence[int] | None = None,
    pages_per_request: int = 25,
    progress: ProgressCallback | None = None,
) -> list[tuple[int, str]]:
    """OCR a PDF in page chunks and preserve original one-based page numbers."""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    indexes = list(page_indexes) if page_indexes is not None else list(range(len(reader.pages)))
    if not indexes:
        return []

    page_chunks = list(chunks(indexes, pages_per_request))
    extracted: list[tuple[int, str]] = []

    for chunk_number, page_chunk in enumerate(page_chunks, start=1):
        if progress:
            progress(
                chunk_number - 1,
                len(page_chunks),
                f"OCR request {chunk_number}/{len(page_chunks)}",
            )

        writer = PdfWriter()
        for index in page_chunk:
            writer.add_page(reader.pages[index])

        buffer = io.BytesIO()
        writer.write(buffer)
        response = _mistral_ocr_request(
            api_key=api_key,
            document={
                "type": "document_url",
                "document_url": _pdf_data_url(buffer.getvalue()),
            },
        )

        response_pages = _response_pages(response)
        if not response_pages:
            raise RuntimeError(
                f"OCR request {chunk_number}/{len(page_chunks)} completed but returned no pages."
            )

        for local_index, page in enumerate(response_pages):
            if local_index >= len(page_chunk):
                break
            original_page = page_chunk[local_index] + 1
            extracted.append((original_page, clean_markdown(page.get("markdown", ""))))

        if progress:
            progress(
                chunk_number,
                len(page_chunks),
                f"OCR request {chunk_number}/{len(page_chunks)} complete",
            )

    extracted.sort(key=lambda item: item[0])
    return extracted


def ocr_document_url(api_key: str, document_url: str) -> list[tuple[int, str]]:
    response = _mistral_ocr_request(
        api_key=api_key,
        document={
            "type": "document_url",
            "document_url": clean_text(document_url),
        },
    )
    return [
        (index + 1, clean_markdown(page.get("markdown", "")))
        for index, page in enumerate(_response_pages(response))
    ]


def ocr_image_bytes(api_key: str, image_bytes: bytes, suffix: str) -> list[tuple[int, str]]:
    response = _mistral_ocr_request(
        api_key=api_key,
        document={
            "type": "image_url",
            "image_url": _image_data_url(image_bytes, suffix),
        },
    )
    return [
        (index + 1, clean_markdown(page.get("markdown", "")))
        for index, page in enumerate(_response_pages(response))
    ]


def ocr_image_url(api_key: str, image_url: str) -> list[tuple[int, str]]:
    response = _mistral_ocr_request(
        api_key=api_key,
        document={
            "type": "image_url",
            "image_url": clean_text(image_url),
        },
    )
    return [
        (index + 1, clean_markdown(page.get("markdown", "")))
        for index, page in enumerate(_response_pages(response))
    ]

def clean_markdown(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("\x00", "").strip()


def _engineering_drawing_rotation_candidate(markdown: Any, source_text: Any = "") -> bool:
    """Return True when OCR likely saw only part of a rotated engineering drawing.

    This is manufacturer-agnostic. It looks for evidence of an engineering title block,
    drawing/section context, or a non-orderable construction table while a genuine
    orderable-parts header is still absent. Only such pages are eligible for the
    additional orientation OCR pass.
    """
    combined = clean_text(f"{markdown}\n{source_text}").lower()
    compact = re.sub(r"[^a-z0-9]+", "", combined)
    if not combined:
        return False
    if _is_engineering_article_table(markdown):
        return False

    drawing_context = bool(
        re.search(r"\b(?:drawing|drawings|dimension|technical data|assembly|sheet)\b", combined)
        or re.search(r"\b\d+\.\d+(?:\.\d+)+\b", combined)
    )
    construction_table = bool(
        "composition" in combined
        and "designation" in combined
        and re.search(r"\bpos\.?\b", combined)
    )
    legend_hint = bool(
        re.search(r"\b(?:legend|legenda|key)\s*:?\b", combined, flags=re.I)
    )
    title_block_hint = any(
        token in compact
        for token in (
            "documentno", "documentnumber", "drawingno", "drawingnumber",
            "dwgno", "responsibledepartment", "noofsheets", "revisionno",
            "sheetno", "creator", "approved", "title",
        )
    )
    document_number = bool(re.search(r"(?<!\d)\d{6,10}(?!\d)", combined))
    # A searchable PDF often exposes only the chapter heading and drawing number;
    # the embedded drawing, legend and article table remain rotated image/vector
    # content. Treat that combination as sufficient rescue evidence so drawings
    # without an already recognized table are not permanently assembly-only.
    return bool(
        document_number
        and drawing_context
        and (construction_table or title_block_hint or legend_hint or not _is_engineering_article_table(markdown))
    )


def _orientation_table_evidence(markdown: Any) -> int:
    """Score an OCR orientation by orderable-table and title-block evidence."""
    text = clean_text(markdown).lower()
    compact = re.sub(r"[^a-z0-9]+", "", text)
    score = 0
    signals = {
        "articleno": 8,
        "articlenumber": 8,
        "namedesignation": 7,
        "partname": 6,
        "materialblank": 3,
        "documentno": 4,
        "drawingno": 4,
        "purchasedarticle": 4,
        "quantity": 2,
        "qty": 2,
    }
    for token, weight in signals.items():
        if token in compact:
            score += weight
    score += min(12, len(re.findall(r"(?m)^\s*\d{6,10}(?:\s+[A-Z0-9]{1,4})?\b", str(markdown or ""))))
    if _is_engineering_article_table(markdown):
        score += 20
    legend_entries = _engineering_legend_entries(markdown)
    if legend_entries:
        score += 12 + min(12, len(legend_entries) * 2)
    variant_rows = _headerless_engineering_variant_rows(0, markdown, {})
    if variant_rows:
        score += 14 + min(14, len(variant_rows))
    return score


def _cropped_rotated_pdf_page_bytes(
    pdf_bytes: bytes,
    page_number: int,
    rotation: int,
    fractions: tuple[float, float, float, float],
) -> bytes:
    """Return one cropped PDF page for focused OCR without raster dependencies.

    ``fractions`` is (left, bottom, right, top) in the original page coordinate
    system. The crop deliberately overlaps neighbouring regions in callers so a
    table that sits on a split boundary is still visible in at least one request.
    This is generic engineering-drawing recovery and is not tied to any maker.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    page = reader.pages[int(page_number) - 1]
    media = page.mediabox
    left, bottom, right, top = fractions
    width = float(media.right) - float(media.left)
    height = float(media.top) - float(media.bottom)
    x0 = float(media.left) + max(0.0, min(1.0, left)) * width
    y0 = float(media.bottom) + max(0.0, min(1.0, bottom)) * height
    x1 = float(media.left) + max(0.0, min(1.0, right)) * width
    y1 = float(media.bottom) + max(0.0, min(1.0, top)) * height
    if x1 <= x0 or y1 <= y0:
        raise ValueError("Invalid engineering-drawing OCR crop.")
    box = RectangleObject([x0, y0, x1, y1])
    page.cropbox = box
    page.trimbox = box
    page.mediabox = box
    if rotation:
        page.rotate(int(rotation) % 360)
    writer = PdfWriter()
    writer.add_page(page)
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def _focused_engineering_region_ocr(
    api_key: str,
    pdf_bytes: bytes,
    page_number: int,
    rotation: int,
    base_markdown: str,
    max_regions: int = 4,
    force: bool = False,
) -> tuple[str, list[str]]:
    """Recover tiny orderable tables that full-page OCR can detect but not read.

    A common failure on dense technical drawings is that full-page OCR recognizes
    the semantic table headers but the actual article identifiers are too small to
    survive. In that case we OCR overlapping half-page regions at the already
    selected orientation. The resulting text is appended only when it increases
    source-backed identifier evidence or otherwise strengthens the table evidence.
    """
    combined = clean_markdown(base_markdown)
    messages: list[str] = []
    article_table = _is_engineering_article_table(combined)
    legend_page = _is_engineering_legend_page(combined)
    variant_rows = _headerless_engineering_variant_rows(0, combined, {})
    # ``force`` is used only after native PDF evidence has already proved that the
    # page is an equipment drawing. It breaks the previous circular dependency in
    # which a tiny table had to be recognized by full-page OCR before it was allowed
    # to receive the higher-resolution OCR needed to recognize that same table.
    if not force and (not combined or not (article_table or legend_page or variant_rows)):
        return combined, messages

    metadata = _engineering_title_block_metadata(combined, require_article_table=False)
    section_code = clean_text(metadata.get("section_code", ""))
    identifier_count = (
        len(_engineering_article_identifier_evidence(combined, section_code))
        if article_table
        else (
            len(_engineering_legend_entries(combined))
            if legend_page
            else len(variant_rows)
        )
    )
    # Headerless variant matrices commonly contain 10-20 rows. Three visible rows
    # prove the pattern but do not prove coverage, so keep focused OCR active until
    # a more representative set has been read. Ordinary article/legend pages retain
    # the lower threshold used by the prior build.
    sufficient_identifier_count = 20 if (article_table or variant_rows) else 2
    if identifier_count >= sufficient_identifier_count:
        return combined, messages

    # Overlapping halves preserve the full width/height of wide tables better than
    # four isolated quadrants. Rotating occurs after cropping, using the same
    # orientation that won the full-page rescue.
    regions = [
        ("LEFT HALF", (0.00, 0.00, 0.58, 1.00)),
        ("RIGHT HALF", (0.42, 0.00, 1.00, 1.00)),
        ("BOTTOM HALF", (0.00, 0.00, 1.00, 0.58)),
        ("TOP HALF", (0.00, 0.42, 1.00, 1.00)),
    ][: max(1, int(max_regions))]

    best_score = _orientation_table_evidence(combined)
    used_regions = 0
    for label, fractions in regions:
        try:
            crop_bytes = _cropped_rotated_pdf_page_bytes(
                pdf_bytes=pdf_bytes,
                page_number=page_number,
                rotation=rotation,
                fractions=fractions,
            )
            response = _mistral_ocr_request(
                api_key=api_key,
                document={
                    "type": "document_url",
                    "document_url": _pdf_data_url(crop_bytes),
                },
            )
            response_pages = _response_pages(response)
            region_markdown = (
                clean_markdown(response_pages[0].get("markdown", ""))
                if response_pages else ""
            )
            if not region_markdown:
                continue
            candidate = (
                combined
                + f"\n\n===== FOCUSED REGION OCR {label} =====\n"
                + region_markdown
            ).strip()
            candidate_article_table = _is_engineering_article_table(candidate)
            candidate_legend_page = _is_engineering_legend_page(candidate)
            candidate_variant_rows = _headerless_engineering_variant_rows(
                0, candidate, {}
            )
            candidate_metadata = _engineering_title_block_metadata(
                candidate, require_article_table=False
            )
            candidate_code = clean_text(
                candidate_metadata.get("section_code", section_code)
            )
            candidate_count = (
                len(_engineering_article_identifier_evidence(candidate, candidate_code))
                if candidate_article_table
                else (
                    len(_engineering_legend_entries(candidate))
                    if candidate_legend_page
                    else len(candidate_variant_rows)
                )
            )
            candidate_score = _orientation_table_evidence(candidate)
            candidate_has_details = bool(
                candidate_article_table
                or candidate_legend_page
                or candidate_variant_rows
            )
            if candidate_has_details and (
                candidate_count > identifier_count or candidate_score > best_score + 2
            ):
                combined = candidate
                identifier_count = candidate_count
                best_score = candidate_score
                section_code = candidate_code
                used_regions += 1
            sufficient_identifier_count = (
                20
                if (
                    _is_engineering_article_table(combined)
                    or _is_headerless_engineering_variant_table(combined)
                )
                else 3
            )
            if identifier_count >= sufficient_identifier_count:
                break
        except Exception as exc:
            messages.append(
                f"PDF page {page_number}: focused {label.lower()} OCR was unavailable; "
                f"other recovery paths continued. Details: {_safe_api_error_text(exc, api_key)}"
            )

    if used_regions:
        messages.insert(
            0,
            f"PDF page {page_number}: focused region OCR added {used_regions} higher-resolution "
            f"drawing region(s) and recovered {identifier_count} source drawing-detail candidate(s).",
        )
    elif identifier_count < 2:
        messages.insert(
            0,
            f"PDF page {page_number}: the full-page OCR recognized a drawing-detail source, but "
            "focused region OCR still could not read enough coded entries.",
        )
    return combined, messages


def recover_rotated_engineering_drawing_pages(
    api_key: str,
    pdf_bytes: bytes,
    extracted_pages: Sequence[tuple[int, str]],
    candidate_page_numbers: Sequence[int] | None = None,
    progress: ProgressCallback | None = None,
    max_pages: int = 12,
) -> tuple[list[tuple[int, str]], list[str], list[int]]:
    """Retry likely rotated engineering drawings using alternate page orientations.

    The recovery is generic: a candidate page is tried at 90, 270, then 180 degrees.
    The best result is accepted only when it exposes a coherent orderable-parts table
    with stronger structural evidence than the original OCR. Original page numbering
    is retained and the rescue text is appended for downstream reconciliation.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    original_lookup = {int(page): clean_markdown(markdown) for page, markdown in extracted_pages}
    allowed = (
        {int(page) for page in candidate_page_numbers}
        if candidate_page_numbers is not None
        else set(original_lookup)
    )
    candidates: list[int] = []
    for page_number in sorted(allowed):
        if page_number not in original_lookup or not 1 <= page_number <= len(reader.pages):
            continue
        source_text = ""
        try:
            source_text = reader.pages[page_number - 1].extract_text() or ""
        except Exception:
            source_text = ""
        if _engineering_drawing_rotation_candidate(original_lookup[page_number], source_text):
            candidates.append(page_number)
        if len(candidates) >= max(1, int(max_pages)):
            break

    if not candidates:
        return list(extracted_pages), [], []

    messages: list[str] = []
    rescued_pages: list[int] = []
    updated = dict(original_lookup)
    total = len(candidates)

    for index, page_number in enumerate(candidates, start=1):
        if progress:
            progress(index - 1, total, f"Drawing orientation OCR rescue {index}/{total} (PDF page {page_number})")
        original_score = _orientation_table_evidence(original_lookup[page_number])
        best_markdown = ""
        best_rotation = 0
        best_score = original_score
        try:
            for rotation in (90, 270, 180):
                page_reader = PdfReader(io.BytesIO(pdf_bytes))
                page = page_reader.pages[page_number - 1]
                page.rotate(rotation)
                writer = PdfWriter()
                writer.add_page(page)
                buffer = io.BytesIO()
                writer.write(buffer)
                response = _mistral_ocr_request(
                    api_key=api_key,
                    document={"type": "document_url", "document_url": _pdf_data_url(buffer.getvalue())},
                )
                response_pages = _response_pages(response)
                rescue_markdown = clean_markdown(response_pages[0].get("markdown", "")) if response_pages else ""
                score = _orientation_table_evidence(rescue_markdown)
                if score > best_score:
                    best_markdown, best_rotation, best_score = rescue_markdown, rotation, score
                if rescue_markdown and (
                    _is_engineering_article_table(rescue_markdown)
                    or _is_engineering_legend_page(rescue_markdown)
                    or _is_headerless_engineering_variant_table(rescue_markdown)
                ):
                    best_markdown, best_rotation, best_score = rescue_markdown, rotation, score
                    break

            # A first OCR pass may already see the Article-No. header but only a
            # fraction of a dense table. Keep the original orientation eligible for
            # focused region OCR even when no rotated pass scores higher.
            if not best_markdown and (
                _is_engineering_article_table(original_lookup[page_number])
                or _is_engineering_legend_page(original_lookup[page_number])
                or _is_headerless_engineering_variant_table(
                    original_lookup[page_number]
                )
            ):
                best_markdown = original_lookup[page_number]
                best_rotation = 0
                best_score = original_score

            # Native/source-text gating already proved this is a likely equipment
            # drawing. If every full-page orientation missed its tiny orderable
            # matrix, keep the original pass as the base and let focused crops find
            # the Article-No. / Name-Designation rows directly.
            if not best_markdown:
                best_markdown = original_lookup[page_number]
                best_rotation = 0
                best_score = original_score

            best_has_details = bool(
                _is_engineering_article_table(best_markdown)
                or _is_engineering_legend_page(best_markdown)
                or _is_headerless_engineering_variant_table(best_markdown)
            )
            focused_messages: list[str] = []
            focused_markdown = best_markdown
            focused_markdown, focused_messages = _focused_engineering_region_ocr(
                api_key=api_key,
                pdf_bytes=pdf_bytes,
                page_number=page_number,
                rotation=best_rotation,
                base_markdown=best_markdown,
                force=not best_has_details,
            )
            focused_has_details = bool(
                _is_engineering_article_table(focused_markdown)
                or _is_engineering_legend_page(focused_markdown)
                or _is_headerless_engineering_variant_table(focused_markdown)
            )

            if focused_has_details and (
                best_score > original_score
                or best_rotation == 0
                or focused_markdown != best_markdown
            ):
                best_markdown = focused_markdown or best_markdown
                if best_rotation == 0:
                    updated[page_number] = best_markdown
                    messages.append(
                        f"PDF page {page_number}: focused OCR expanded the orderable drawing table in its original orientation."
                    )
                else:
                    updated[page_number] = (
                        original_lookup[page_number]
                        + f"\n\n===== ORIENTATION OCR RESCUE {best_rotation} DEG =====\n"
                        + best_markdown
                    ).strip()
                    messages.append(
                        f"PDF page {page_number}: orientation OCR recovered a coded drawing legend or orderable spare table at {best_rotation} degrees."
                    )
                rescued_pages.append(page_number)
                messages.extend(focused_messages)
            else:
                messages.append(
                    f"PDF page {page_number}: alternate-orientation OCR was attempted but did not reveal stronger coded drawing details, so the original OCR was kept."
                )
        except Exception as exc:
            messages.append(
                f"PDF page {page_number}: alternate-orientation OCR rescue was unavailable; the original OCR was kept. Details: {_safe_api_error_text(exc, api_key)}"
            )
        if progress:
            progress(index, total, f"Drawing orientation OCR rescue {index}/{total} complete")

    rebuilt = [(int(page), updated.get(int(page), clean_markdown(markdown))) for page, markdown in extracted_pages]
    return rebuilt, messages, rescued_pages


# Backward-compatible alias for deployments that still import the older helper name.
recover_rotated_alfa_laval_drawing_pages = recover_rotated_engineering_drawing_pages


# ---------------------------------------------------------------------------
# Local page classification / pre-filtering
# ---------------------------------------------------------------------------


_POSITIVE_PAGE_PHRASES: tuple[tuple[str, int], ...] = (
    ("spare parts", 5),
    ("list of parts", 5),
    ("parts list", 5),
    ("part no", 4),
    ("part number", 4),
    ("article no", 5),
    ("article number", 5),
    ("name/designation", 4),
    ("part name", 3),
    ("ref. no", 3),
    ("ref no", 3),
    ("reference no", 3),
    ("replacement parts", 4),
    ("recommended spare parts", 6),
    ("spare part number", 6),
    ("sparepart number", 6),
    ("illustrated parts", 4),
    ("item no", 3),
    ("item number", 3),
    ("position no", 3),
    ("description", 2),
    ("designation", 2),
    ("denomination", 2),
    ("quantity", 2),
    (" qty", 2),
    (" qnt", 2),
)

_NEGATIVE_PAGE_PHRASES: tuple[tuple[str, int], ...] = (
    ("table of contents", 7),
    ("revision history", 6),
    ("list of revisions", 6),
    ("document revisions", 6),
    ("foreword", 4),
    ("preface", 4),
    ("general description", 3),
    ("operating instructions", 3),
    ("safety instructions", 3),
    ("maintenance instructions", 2),
)

_PART_IDENTIFIER_PATTERN = re.compile(
    r"(?<![A-Z0-9])[A-Z0-9][A-Z0-9./_-]{3,}(?![A-Z0-9])",
    flags=re.IGNORECASE,
)


def _markdown_table_signal(markdown: str) -> tuple[int, int]:
    """Return (table rows, relevant-header hits) for OCR markdown."""
    lines = markdown.splitlines()
    table_rows = sum(1 for line in lines if line.count("|") >= 2)
    header_hits = 0
    for line in lines:
        if "|" not in line:
            continue
        lowered = line.lower()
        if any(
            phrase in lowered
            for phrase in (
                "part no",
                "part number",
                "article no",
                "article number",
                "name/designation",
                "part name",
                "ref. no",
                "ref no",
                "reference no",
                "item no",
                "item number",
                "description",
                "designation",
                "quantity",
                "qty",
                "qnt",
            )
        ):
            header_hits += 1
    return table_rows, header_hits


def classify_ocr_pages(
    extracted_pages: Sequence[tuple[int, str]],
    mode: str = "Conservative (recommended)",
) -> tuple[list[tuple[int, str]], pd.DataFrame]:
    """
    Classify OCR pages locally before paid structured extraction.

    Conservative mode skips only pages that are very clearly front matter or prose.
    Strict mode processes only strong spare-parts candidates. Off processes everything.
    No extra API calls are made.
    """
    if mode not in PAGE_FILTER_MODES:
        mode = "Conservative (recommended)"

    selected: list[tuple[int, str]] = []
    records: list[dict[str, Any]] = []

    for page_number, markdown in extracted_pages:
        text = clean_markdown(markdown)
        lowered = text.lower()
        characters = len(text)
        table_rows, header_hits = _markdown_table_signal(text)
        identifier_hits = len(_PART_IDENTIFIER_PATTERN.findall(text))

        explicit_spare_source = _looks_like_explicit_spare_number_page(text)
        positive_score = sum(weight for phrase, weight in _POSITIVE_PAGE_PHRASES if phrase in lowered)
        negative_score = sum(weight for phrase, weight in _NEGATIVE_PAGE_PHRASES if phrase in lowered)
        score = positive_score + min(table_rows, 8) + (header_hits * 3) + min(identifier_hits // 4, 5) - negative_score

        strong_candidate = (
            explicit_spare_source
            or positive_score >= 5
            or header_hits >= 1
            or (table_rows >= 4 and identifier_hits >= 3)
        )
        obvious_non_parts = (
            characters < 40
            or (negative_score >= 5 and not strong_candidate)
            or (table_rows == 0 and positive_score == 0 and identifier_hits < 2 and characters < 1800)
        )

        if mode == "Off":
            classification = "All pages"
            process_page = True
            reason = "Filtering disabled"
        elif strong_candidate:
            classification = (
                "Explicit spare-number source"
                if explicit_spare_source
                else "Spare-parts candidate"
            )
            process_page = True
            reason = (
                f"parts signals={positive_score}; table rows={table_rows}; "
                f"header hits={header_hits}; identifiers={identifier_hits}; "
                f"explicit spare label={'yes' if explicit_spare_source else 'no'}"
            )
        elif obvious_non_parts:
            classification = "Skipped obvious non-parts page"
            process_page = False
            reason = (
                f"negative signals={negative_score}; table rows={table_rows}; "
                f"parts signals={positive_score}; identifiers={identifier_hits}"
            )
        else:
            classification = "Ambiguous"
            process_page = mode == "Conservative (recommended)"
            reason = (
                f"score={score}; table rows={table_rows}; "
                f"parts signals={positive_score}; identifiers={identifier_hits}"
            )

        if process_page:
            selected.append((page_number, text))

        records.append(
            {
                "SOURCE PAGE": int(page_number),
                "CLASSIFICATION": classification,
                "PROCESS": bool(process_page),
                "SCORE": int(score),
                "REASON": reason,
                "CHARACTERS": int(characters),
            }
        )

    return selected, pd.DataFrame(records, columns=PAGE_CLASSIFICATION_COLUMNS)


# ---------------------------------------------------------------------------
# Structured spare-parts extraction
# ---------------------------------------------------------------------------


MULTILINGUAL_ORDER_CATALOG_PROFILE = "multilingual_order_catalog"
ENGINEERING_ARTICLE_DRAWING_PROFILE = "engineering_article_drawing"
ALFA_LAVAL_ARTICLE_DRAWING_PROFILE = ENGINEERING_ARTICLE_DRAWING_PROFILE  # compatibility alias


def _is_multilingual_order_catalog(markdown: Any) -> bool:
    """Recognize older multilingual catalogues whose real identifier is Order-No.

    These catalogues commonly place a drawing position at the far left, followed
    by burner/equipment applicability, size, Order-No., weight, and three parallel
    German/English/French designation columns.  The layout is materially different
    from the Ident-No. tables used by newer manuals, so it needs its own mapping.
    """
    text = clean_text(markdown).lower()
    order_signal = any(
        token in text
        for token in (
            "order-no",
            "order no",
            "orderno",
            "bestell-nr",
            "bestell nr",
            "no de commande",
        )
    )
    multilingual_signal = (
        "designation" in text
        and any(token in text for token in ("bezeichnung", "benennung"))
    )
    layout_signal = any(
        token in text
        for token in (
            "burner serie",
            "burner series",
            "brenner-typenreihe",
            "type bruleur",
            "type brûleur",
            "pict.",
            "photo",
            "appr. kg",
        )
    )
    return bool(order_signal and multilingual_signal and layout_signal)


def _is_engineering_article_table(markdown: Any) -> bool:
    """Recognize an orderable engineering-drawing table by semantic headers.

    The detector is manufacturer-neutral. It requires an Article Number style
    identifier plus a name/designation field and at least one additional drawing-
    table/title-block signal so ordinary prose containing those words is ignored.
    """
    text = clean_text(markdown).lower()
    compact = re.sub(r"[^a-z0-9]+", "", text)
    article_signal = any(token in compact for token in ("articleno", "articlenumber", "articlenr"))
    name_signal = "namedesignation" in compact or ("name" in text and "designation" in text)
    drawing_signal = any(
        token in compact
        for token in (
            "materialblank", "artrev", "documentno", "documentnumber",
            "drawingno", "drawingnumber", "dwgno", "purchasedarticle",
            "sheetno", "revisionno",
        )
    )
    return bool(article_signal and name_signal and drawing_signal)


_ENGINEERING_LEGEND_HEADER_RE = re.compile(
    r"^\s*(?:legend|legenda|key)\s*:?\s*$",
    flags=re.IGNORECASE,
)
_ENGINEERING_LEGEND_CODE_RE = re.compile(
    r"^(?:\d{1,4}|[A-Z]{1,5}\d{1,4}(?:[-/.][A-Z0-9]{1,6})*)$",
    flags=re.IGNORECASE,
)
_ENGINEERING_LEGEND_STOP_RE = re.compile(
    r"^(?:title|document\s*(?:no\.?|number)|drawing\s*(?:no\.?|number)|"
    r"sheet\s*(?:no\.?|number)|revision|responsible\s+department|creator|"
    r"approved|first\s+angle|book\s+no\.?|flanges?\b)",
    flags=re.IGNORECASE,
)
_ENGINEERING_LEGEND_SPEC_RE = re.compile(
    r"\s+(?=(?:DN\s*[/A-Z0-9]|PN\s*\d|ISO[- ]?G|INLET\b|OUTLET\b|"
    r"CONNECTING\b|CONE\b|AL\s*[-:]|\d{2,4}\s*[-–]\s*\d{2,4}\s*V\b|"
    r"M\d+(?:\s*[X×]\s*\d+)?\b|[Ø⌀]\s*\d))",
    flags=re.IGNORECASE,
)


def _valid_engineering_legend_code(value: Any) -> str:
    text = clean_text(value).upper().strip(" .,:;|#")
    if not text or not _ENGINEERING_LEGEND_CODE_RE.fullmatch(text):
        return ""
    if _is_date_like_section_code(text):
        return ""
    return text


def _clean_engineering_legend_description(value: Any) -> str:
    text = clean_text(value).strip(" |:;-–—")
    if not text:
        return ""
    # Keep the coded component/service name and discard the dimensional or
    # installation specification that follows it. Markdown tables already keep
    # these in separate cells; flattened OCR normally retains one of these markers.
    text = _ENGINEERING_LEGEND_SPEC_RE.split(text, maxsplit=1)[0]
    text = re.sub(r"\s+", " ", text).strip(" |:;-–—")
    if not text or len(text) > 80 or not re.search(r"[A-Za-z]", text):
        return ""
    if _ENGINEERING_LEGEND_STOP_RE.match(text):
        return ""
    return text.upper()


def _engineering_legend_entries(markdown: Any) -> list[tuple[str, str]]:
    """Return source-coded entries from a labelled engineering drawing legend.

    The parser is deliberately restricted to a visible Legend/Legenda/Key block.
    It supports Markdown table cells and flattened aligned OCR, stops at the title
    block, and never turns nearby dimensions or specifications into identifiers.
    """
    raw_lines = str(markdown or "").splitlines()
    active = False
    remaining = 0
    entries: list[tuple[str, str]] = []
    seen: set[str] = set()

    for raw_line in raw_lines:
        cleaned_line = clean_text(raw_line).strip()
        header_candidate = re.sub(r"^[#>*+\-\s]+", "", cleaned_line)
        header_candidate = header_candidate.strip(" |:;-–—")
        header_cells = [
            clean_text(cell).strip(" |:;-–—")
            for cell in re.split(r"\s*\|\s*", cleaned_line)
            if clean_text(cell).strip(" |:;-–—")
        ]
        if _ENGINEERING_LEGEND_HEADER_RE.fullmatch(header_candidate) or any(
            _ENGINEERING_LEGEND_HEADER_RE.fullmatch(cell)
            for cell in header_cells
        ):
            active = True
            remaining = 28
            continue
        if not active:
            continue
        if remaining <= 0:
            active = False
            continue
        remaining -= 1
        if not cleaned_line:
            continue

        plain = re.sub(r"^[#>*+\-\s]+", "", cleaned_line).strip()
        plain = plain.strip(" |")
        if _ENGINEERING_LEGEND_STOP_RE.match(plain):
            active = False
            continue

        cells = [
            clean_text(cell).strip(" |:;-–—")
            for cell in re.split(r"\s*\|\s*|\t+|\s{2,}", plain)
            if clean_text(cell).strip(" |:;-–—")
        ]
        code = ""
        description = ""
        if len(cells) >= 2:
            code = _valid_engineering_legend_code(cells[0])
            description = _clean_engineering_legend_description(cells[1])
        if not code:
            match = re.match(
                r"^\s*(?P<code>\d{1,4}|[A-Z]{1,5}\d{1,4}(?:[-/.][A-Z0-9]{1,6})*)"
                r"\s+(?P<description>.+?)\s*$",
                plain,
                flags=re.IGNORECASE,
            )
            if match:
                code = _valid_engineering_legend_code(match.group("code"))
                description = _clean_engineering_legend_description(
                    match.group("description")
                )
        key = normalize_key(code)
        if code and description and key not in seen:
            seen.add(key)
            entries.append((code, description))

    # One isolated short token is too easy to confuse with a callout. A labelled
    # legend containing at least two coherent entries is strong page-level evidence.
    return entries if len(entries) >= 2 else []


def _is_engineering_legend_page(markdown: Any) -> bool:
    return bool(_engineering_legend_entries(markdown))


# Backward-compatible internal alias. New logic is intentionally manufacturer-neutral.
def _is_alfa_laval_article_drawing(markdown: Any) -> bool:
    return _is_engineering_article_table(markdown)


def _document_extraction_profile(
    extracted_pages: Sequence[tuple[int, str]],
) -> str:
    # Orderable engineering-drawing tables can appear hundreds of pages into a
    # compiled system manual, so scan page-by-page instead of sampling only the
    # beginning of the document. This is local string inspection and adds no API cost.
    for _, markdown in extracted_pages:
        if (
            _is_engineering_article_table(markdown)
            or _is_headerless_engineering_variant_table(markdown)
        ):
            return ENGINEERING_ARTICLE_DRAWING_PROFILE
    sample = "\n".join(str(markdown or "") for _, markdown in extracted_pages[:20])
    if _is_multilingual_order_catalog(sample):
        return MULTILINGUAL_ORDER_CATALOG_PROFILE
    return ""


def _profile_prompt(profile: str) -> str:
    if profile == ENGINEERING_ARTICLE_DRAWING_PROFILE:
        return """
AUTOMATIC LAYOUT PROFILE - ENGINEERING DRAWING TITLE BLOCK + ARTICLE TABLE:
- A page is a spare-parts source when it contains Article No. / Art. Rev. /
  Name/Designation / Material/Blank / Note.
- A drawing title block can be authoritative for the sub-machinery: Document No./Drawing No. is
  section_code and Title/Assembly is section_name_english when both are clearly paired. Example: Document No. 9007280
  and Title Cable means section_code="9007280", section_name_english="CABLE".
- Article No. is the genuine spare identifier and populates ident_no, therefore both
  PART NO and CODE. Preserve printed spaces and leading zeroes.
- Name/Designation is the spare description. Art. Rev. is revision metadata, not
  ITEM NO. Material/Blank (including PURCHASED ARTICLE) is classification, not the
  spare description. Note is specification text, not quantity.
- If there is no genuine drawing-position column, item_no must be blank. If there is
  no genuine quantity column, quantity must be null. Note values such as 10 m, 25 m,
  50 m, 75 m, 100 m, or Length acc. to order are not QNT.
- Reject a separate Pos. / Designation / Composition construction table as a spare
  table. CONDUCTOR, FILLER, WRAPPING, EMC SCREEN and SHEATH are not spare rows unless
  that logical row also has a genuine Article No.
- A labelled drawing Legend/Legenda/Key can contain coded component entries even
  when no Article-No. table is present. For a drawing parent that otherwise has no
  detailed spare rows, preserve each printed legend code as ident_no and the short
  English legend name as description_english. Ignore the dimensional/specification
  continuation after that name. If the same page also has a genuine Article-No. /
  Name-Designation table, extract the Article-No. table and do not duplicate its
  legend callouts.
- Some drawings contain a headerless two-column orderable-variant table. When at
  least three rows repeat one long article-number family followed by a variant
  suffix (for example `9007170 72/92`) and an English name (`Valve DN500 / A500`),
  preserve the complete printed first cell as ident_no and the second cell as
  description_english. Do not shorten the identifier to its shared family stem.
- Never use chapter headings such as Dimension drawings including technical data,
  dates, sheet numbers, or revision numbers as the sub-machinery code/name.
""".strip()
    if profile != MULTILINGUAL_ORDER_CATALOG_PROFILE:
        return ""
    return """
AUTOMATIC LAYOUT PROFILE - MULTILINGUAL ORDER-NO. CATALOGUE:
- The column headed Bestell-Nr. / Order-No. / No de commande is the genuine
  spare-part identifier. Copy it to ident_no. It will populate both PART NO and CODE.
- Bild / Pict. / Photo is the drawing position and must populate item_no. It is
  never the spare-part code.
- Use only the middle English Designation column. Ignore the German Bezeichnung
  and the final French Designation column.
- Burner serie / Type bruleur and Size/Grandeur describe applicability. They are
  not part numbers, quantities, or sub-machinery codes.
- ca. kg / appr. kg / env. kg is WEIGHT. Never map it to quantity. Return quantity=null.
- A single printed item position can contain several burner-series/size variants
  and several Order-No. values. Return one row for EVERY Order-No.; repeat the same
  item_no and English description on each variant row.
- German assembly words such as Brennermotor, Luftregelung, Stellantrieb, and
  Brennergehause are hierarchy/title text, not a manufacturer. Leave section_maker
  empty unless the page explicitly labels a maker/manufacturer or shows a clear brand.
- Section headings are simple numbered English headings such as "3. Servo drive
  for oil burners". Use the integer as section_code and the exact English heading
  as section_name_english, without appending the code in parentheses. Carry it
  through consecutive continuation pages until the next numbered section heading.
- A repeated item position, Order-No., or individual spare description is never a
  hierarchy heading. Do not promote table-body rows to sub-machineries.
""".strip()


EXTRACTION_SYSTEM_PROMPT = """
You are a precise technical-document extraction engine for marine and industrial
spare-parts manuals. Convert OCR markdown into structured rows for the Benefit
machinery and spare-parts import template.

Return one JSON object with exactly this top-level structure:
{
  "spare_parts": [
    {
      "section_code": "",
      "section_name_english": "",
      "section_maker": "",
      "section_model": "",
      "table_title": "",
      "section_start_page": 1,
      "source_part_no": "",
      "ident_no": "",
      "item_no": "",
      "description_english": "",
      "unit": "",
      "quantity": null,
      "source_page": 1,
      "confidence": 0.0
    }
  ]
}

Benefit mapping rules:
1. Extract every genuine spare-part row. Never summarize, merge, or invent rows.
2. section_code is the drawing, table, plate, assembly, or sub-machinery code that
   governs the row. Preserve all punctuation and leading zeros. If a section has
   two printed codes, join them with " / ".
3. section_name_english is the English assembly/sub-machinery title only, in
   UPPERCASE. If the source already prints several languages, choose the English
   wording. Translate only when no English wording is printed.
4. section_maker and section_model come from the relevant PDF page/section. A
   section-specific maker/model overrides the manual-level maker/model. Do not
   invent either value. Only use an explicit maker/manufacturer label or a clear
   brand/logo. German assembly/title words such as Brennermotor, Luftregelung,
   Stellantrieb, Brennergehause, and Geblase are not makers.
5. The first table column is often a drawing position even when its multilingual
   header contains "Part-No.". When it contains sequential callouts such as 1,
   2, 3, 1-12, P101, etc., return it as item_no.
5a. When one printed drawing position/item number is visually merged across two
   or more spare-part records, repeat that same item_no on EVERY returned record.
   Never leave a continuation record blank merely because the source cell is
   merged or printed only once.
6. ident_no is the identifier under headers such as Ident-Nr., Ident-No., Code,
   Material Code, Spare Part Code, or equivalent. This value will populate BOTH
   Benefit CODE and Benefit PART NO.
6a. Under a multilingual Bestell-Nr. / Order-No. / No de commande header,
    Order-No. is ident_no. Never confuse it with the Bild/Pict./Photo position,
    burner-series applicability, size, unit marker, or weight.
6b. OCR may flatten several Order-No. values into one cell. Return one distinct
    spare_parts record for every printed Order-No.; never combine them into one code.
6c. When a parts list uses a single Ref. No. / Reference No. column with composite
    values such as 03036-55 or 03036-58A, the FULL printed reference is ident_no
    (therefore PART NO and CODE), while only its final callout suffix is item_no
    (55 or 58A). Never copy the full composite Ref. No. into ITEM NO.
6d. Successive pages titled Parts List for Figure X-1, X-2, X-3, etc. belong to
    one equipment section when the manual/equipment heading is unchanged. Do not
    create one sub-machinery per sheet. Keep the same section code/name across all
    list sheets, even when illustrated drawing pages occur between them.
6e. Genuine spare rows can also appear outside formal tables. When prose/list text
    explicitly labels a value as "Spare part number", "Sparepart number",
    "Replacement part No.", or equivalent, create one row for that labelled part.
    Use the nearest short item description (for example UV lamp set, temperature
    transmitter, lamp power supply) as description_english. A nearby component
    heading may be used as section_name_english, but NEVER invent section_code.
    Repeated mentions of the same labelled spare number are evidence for one spare,
    not separate orderable rows unless they clearly belong to different parent
    components.
6f. On an equipment drawing, a block explicitly headed Legend, Legenda, or Key may
    define coded component/service entries. Use the exact printed legend code as
    ident_no and only the short English legend name as description_english; exclude
    dimensions, pressure ratings, connection details, voltage/frequency, inlet/outlet
    text, and installation notes. Apply legend recovery only when that drawing parent
    has no stronger detailed spare rows. When the page also contains an Article No. /
    Name-Designation table, the Article-No. rows take priority and legend callouts are
    supporting evidence rather than additional duplicate spares.
6g. A source-confirmed equipment drawing may also contain a headerless two-column
    orderable-variant list. When at least three rows repeat one long article family
    plus a printed suffix (for example `9007170 72/92`) beside an English name
    (`Valve DN500 / A500`), return every full first-cell value as ident_no and the
    second cell as description_english. Preserve spaces, slashes and leading zeroes.
    The common family number alone is not the spare code.
7. source_part_no is any separate manufacturer Part No. printed by the source.
   Capture it only for audit context; it is not used automatically in Benefit.
8. description_english is the individual spare-part name only, in ENGLISH and
   UPPERCASE. When German/English/French are printed together, return ONLY the EXACT
   printed English phrase. Never concatenate the language variants. OCR may flatten
   the three printed lines into one string; still identify and isolate the English
   span. Do not paraphrase, modernize, shorten, expand, or replace it with a synonym.
   Preserve source dimensions, standards, stage numbers, punctuation, and symbols.
   When no English phrase is printed, translate the complete term into precise
   technical English and lower confidence because translation was required.
8a. Apply the same language rule to section_name_english: isolate the printed English
   title when present; otherwise translate the complete section title into English.
8b. Flattened OCR examples that MUST be cleaned:
   KURBELGEHÄUSE CRANK CASE CARTER -> CRANK CASE
   WELLENDICHTRING RADIAL PACKING RING BAGUE D'ETANCHEITE -> RADIAL PACKING RING
   LEISTUNGSSCHILD MANUFACTURER'S NAME PLATE PLAQUE SIGNALETIQUE -> MANUFACTURER'S NAME PLATE
   MANOMETER 1. STUFE PRESSURE GAUGE 1ST STAGE MANOMETRE 1ER ETAGE -> PRESSURE GAUGE 1ST STAGE
9. quantity must be numeric or null. unit is PCS, SET, or an empty string.
10. source_page is the PAGE marker supplied in the user message.
11. section_start_page is the first PDF page where the section begins, including
    its sectional/exploded drawing page when that page precedes the parts table.
12. Continuation pages keep exactly the same section_code, section_name_english,
    section_maker, section_model, and section_start_page only when the current page
    has no new drawing/table code in its own header. A code printed in a spare-part
    row, description, cross-reference, or table body is never a section_code.
12a. section_name_english contains only the clean English assembly title. Do not
     append section_code in parentheses and do not append ordering notes or the
     description of the first spare part below the heading.
13. Do not convert ordinary contents/index entries, drawing callouts, page numbers,
    headers, or descriptive prose into spare-part rows. The exception is a high-signal
    non-tabular list where the source explicitly labels the identifier as a spare or
    replacement part number as described in rule 6e.
14. confidence is 0 to 1 and reflects OCR quality, row alignment, English-language
    selection, and section matching.
15. A weight column headed ca. kg, appr. kg, env. kg, weight, or equivalent is not
    quantity. Return quantity=null when the source has no genuine quantity column.

Critical example:
Source columns:
  Teil-Nr./Part-No. | Ident-Nr./Ident-No. | Benennung/Designation | Qty
Source row:
  2 | 10.10.10.40 | Gleitlager / slide bearing / palier | 1
Return:
  item_no="2", ident_no="10.10.10.40",
  description_english="SLIDE BEARING", quantity=1.

Merged-position continuation example:
  2 | 53.20.1-.2 | 2/2-way solenoid valve 1st and 2nd stage | 2
    | 53.20.3    | 2/2-way solenoid valve 3rd stage         | 1
Return TWO rows and repeat item_no="2" on BOTH rows.
Do NOT return the section drawing code as the spare-part identifier.
""".strip()


def _page_batches(
    extracted_pages: Sequence[tuple[int, str]],
    pages_per_batch: int,
    max_chars: int = 30000,
) -> list[list[tuple[int, str]]]:
    batches: list[list[tuple[int, str]]] = []
    current: list[tuple[int, str]] = []
    current_chars = 0

    for item in extracted_pages:
        page_chars = len(item[1])
        would_overflow = current and (
            len(current) >= max(1, pages_per_batch)
            or current_chars + page_chars > max_chars
        )
        if would_overflow:
            batches.append(current)
            current = []
            current_chars = 0
        current.append(item)
        current_chars += page_chars

    if current:
        batches.append(current)
    return batches


def _parse_json_object(content: Any) -> dict[str, Any]:
    """Parse a JSON object while tolerating fences or harmless surrounding text."""
    if isinstance(content, list):
        content = "".join(
            str(chunk.get("text", ""))
            for chunk in content
            if isinstance(chunk, dict)
        )
    text = str(content or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text)

    candidates: list[str] = [text]
    first_brace = text.find("{")
    last_brace = text.rfind("}")
    if first_brace >= 0 and last_brace > first_brace:
        candidates.append(text[first_brace : last_brace + 1])

    decoder = json.JSONDecoder()
    for candidate in candidates:
        if not candidate:
            continue
        for normalized in (
            candidate,
            re.sub(r",\s*([}\]])", r"\1", candidate),
        ):
            try:
                parsed = json.loads(normalized, strict=False)
            except json.JSONDecodeError:
                try:
                    parsed, _ = decoder.raw_decode(normalized)
                except json.JSONDecodeError:
                    continue
            if isinstance(parsed, list):
                return {"spare_parts": parsed}
            if isinstance(parsed, dict):
                return parsed

    raise json.JSONDecodeError("Could not parse a complete JSON object", text, 0)


def _mistral_json_request(
    api_key: str,
    model: str,
    messages: list[dict[str, str]],
    timeout_seconds: int = 300,
    max_retries: int = 2,
) -> dict[str, Any]:
    endpoint = os.getenv(
        "MISTRAL_CHAT_COMPLETIONS_URL",
        "https://api.mistral.ai/v1/chat/completions",
    )
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "max_tokens": 8192,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    last_error: Exception | None = None
    for attempt in range(max_retries):
        try:
            response = requests.post(
                endpoint,
                headers=headers,
                json=payload,
                timeout=timeout_seconds,
            )
            if response.status_code in {429, 500, 502, 503, 504} and attempt < max_retries - 1:
                time.sleep(2 ** (attempt + 1))
                continue
            response.raise_for_status()
            body = response.json()
            return _parse_json_object(body["choices"][0]["message"]["content"])
        except (
            requests.RequestException,
            KeyError,
            IndexError,
            TypeError,
            json.JSONDecodeError,
            ValueError,
        ) as exc:
            last_error = exc
            if attempt < max_retries - 1:
                time.sleep(1 + attempt)

    raise RuntimeError(f"Mistral structured extraction failed: {last_error}")


DOCUMENT_PROFILE_SYSTEM_PROMPT = """
You are a technical-document layout analyst. Examine representative OCR pages from
one marine or industrial spare-parts manual and describe the manual's extraction
pattern. Return exactly one JSON object with this structure:
{
  "document_family": "",
  "languages": [],
  "english_selection_rule": "",
  "part_identifier_header": "",
  "part_identifier_rule": "",
  "item_number_header": "",
  "item_number_rule": "",
  "quantity_header": "",
  "quantity_rule": "",
  "hierarchy": {
    "major_code_pattern": "",
    "subsection_code_pattern": "",
    "heading_detection_rule": "",
    "continuation_rule": "",
    "examples": [{"code": "", "name_english": "", "page": 1}]
  },
  "table_rules": [],
  "exclude_as_codes": [],
  "uncertainties": [],
  "confidence": 0.0
}

Rules:
1. Base every conclusion on text visibly present in the supplied OCR pages. Do not
   invent headings, translations, columns, codes, or page ranges.
2. Distinguish hierarchy codes from spare-part identifiers, drawing positions,
   burner/application variants, maker names, and model names such as SQM10 or ASZ12.
3. Preserve printed punctuation and trailing zeroes in hierarchy examples.
4. Explain how to select an already printed English phrase. Translation is allowed
   only when no English wording is printed.
5. Identify vertically merged cells, continuation pages, parallel language columns,
   repeated headers, weight columns, and multi-value Order-No./Ident-No. cells when
   the evidence supports them.
6. Keep rules concise and operational. Put any doubtful conclusion in uncertainties
   and lower confidence rather than guessing.
7. confidence is from 0 to 1 and describes confidence in the overall profile.
""".strip()


DOCUMENT_PROFILE_JSON_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "document_family": {"type": "string"},
        "languages": {"type": "array", "items": {"type": "string"}},
        "english_selection_rule": {"type": "string"},
        "part_identifier_header": {"type": "string"},
        "part_identifier_rule": {"type": "string"},
        "item_number_header": {"type": "string"},
        "item_number_rule": {"type": "string"},
        "quantity_header": {"type": "string"},
        "quantity_rule": {"type": "string"},
        "hierarchy": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "major_code_pattern": {"type": "string"},
                "subsection_code_pattern": {"type": "string"},
                "heading_detection_rule": {"type": "string"},
                "continuation_rule": {"type": "string"},
                "examples": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "code": {"type": "string"},
                            "name_english": {"type": "string"},
                            "page": {"type": "integer"},
                        },
                        "required": ["code", "name_english", "page"],
                    },
                },
            },
            "required": [
                "major_code_pattern", "subsection_code_pattern",
                "heading_detection_rule", "continuation_rule", "examples",
            ],
        },
        "table_rules": {"type": "array", "items": {"type": "string"}},
        "exclude_as_codes": {"type": "array", "items": {"type": "string"}},
        "uncertainties": {"type": "array", "items": {"type": "string"}},
        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
    },
    "required": [
        "document_family", "languages", "english_selection_rule",
        "part_identifier_header", "part_identifier_rule", "item_number_header",
        "item_number_rule", "quantity_header", "quantity_rule", "hierarchy",
        "table_rules", "exclude_as_codes", "uncertainties", "confidence",
    ],
}


def _openai_response_text(body: Any) -> str:
    """Return the assistant text from a raw Responses API result."""
    if not isinstance(body, dict):
        return ""
    direct = body.get("output_text")
    if isinstance(direct, str) and direct.strip():
        return direct
    fragments: list[str] = []
    for item in body.get("output", []) or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []) or []:
            if not isinstance(content, dict):
                continue
            if content.get("type") == "output_text" and content.get("text"):
                fragments.append(str(content["text"]))
    return "".join(fragments)


def _openai_json_request(
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    timeout_seconds: int = 180,
) -> dict[str, Any]:
    """Make one bounded OpenAI Responses API call with strict JSON output.

    This function deliberately performs no automatic retry. An optional verifier
    must not consume a restricted key repeatedly or delay the primary Mistral path.
    """
    key = str(api_key or "").strip().strip('"').strip("'")
    if not key:
        raise RuntimeError("OpenAI verification key is not configured.")
    endpoint = os.getenv("OPENAI_RESPONSES_URL", "https://api.openai.com/v1/responses")
    payload = {
        "model": clean_text(model) or "gpt-5.6-sol",
        "input": [
            {
                "role": "system",
                "content": [{"type": "input_text", "text": system_prompt}],
            },
            {
                "role": "user",
                "content": [{"type": "input_text", "text": user_prompt}],
            },
        ],
        "text": {
            "verbosity": "low",
            "format": {
                "type": "json_schema",
                "name": "document_profile",
                "description": "Evidence-based spare-parts manual layout profile.",
                "strict": True,
                "schema": DOCUMENT_PROFILE_JSON_SCHEMA,
            },
        },
        "max_output_tokens": 6000,
    }
    try:
        response = requests.post(
            endpoint,
            headers={
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            json=payload,
            timeout=max(30, int(timeout_seconds)),
        )
    except (requests.Timeout, requests.RequestException, OSError) as exc:
        raise RuntimeError(
            "OpenAI verification could not be reached: "
            + _safe_api_error_text(exc, key)
        ) from exc

    if not 200 <= int(response.status_code) < 300:
        try:
            error_value: Any = response.json()
        except (ValueError, TypeError):
            error_value = getattr(response, "text", "")
        detail = _safe_api_error_text(error_value, key)
        status = int(response.status_code)
        reason = {
            401: "authentication was rejected",
            403: "the project or key lacks permission for this model",
            404: "the configured model or endpoint is unavailable",
            429: "the project quota or rate limit was reached",
        }.get(status, "the optional service returned an error")
        raise RuntimeError(
            f"OpenAI verification skipped because {reason} (HTTP {status}). "
            f"{detail}".strip()
        )

    try:
        body = response.json()
        content = _openai_response_text(body)
        if not content:
            raise ValueError("The response contained no output text.")
        return _parse_json_object(content)
    except (ValueError, TypeError, json.JSONDecodeError) as exc:
        raise RuntimeError(
            "OpenAI verification returned an unusable structured response."
        ) from exc


def _representative_profile_pages(
    extracted_pages: Sequence[tuple[int, str]],
    max_pages: int = 24,
    max_chars_per_page: int = 7000,
) -> list[tuple[int, str]]:
    """Select an evenly distributed, bounded sample for one document-level pass."""
    pages = [(int(page), str(markdown or "")) for page, markdown in extracted_pages]
    if len(pages) <= max(1, int(max_pages)):
        selected = pages
    else:
        last = len(pages) - 1
        indexes = {
            round(position * last / (max(2, int(max_pages)) - 1))
            for position in range(max(2, int(max_pages)))
        }
        selected = [pages[index] for index in sorted(indexes)]
    character_limit = max(1000, int(max_chars_per_page))
    return [
        (page, markdown[:character_limit])
        for page, markdown in selected
        if clean_text(markdown)
    ]


def _profile_text_list(value: Any, maximum: int = 20) -> list[str]:
    if not isinstance(value, list):
        return []
    result: list[str] = []
    for item in value:
        text = clean_text(item)
        if text and text not in result:
            result.append(text)
        if len(result) >= maximum:
            break
    return result


def _normalize_document_profile(
    value: Any,
    model: str,
    analyzed_pages: Sequence[int],
) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    hierarchy_source = value.get("hierarchy", {})
    hierarchy_source = hierarchy_source if isinstance(hierarchy_source, dict) else {}
    examples: list[dict[str, Any]] = []
    for item in hierarchy_source.get("examples", []):
        if not isinstance(item, dict):
            continue
        page_value = quantity_to_number(item.get("page"))
        examples.append(
            {
                "code": clean_text(item.get("code", "")),
                "name_english": clean_text(item.get("name_english", "")),
                "page": int(page_value) if page_value is not None else None,
            }
        )
        if len(examples) >= 30:
            break
    profile = {
        "document_family": clean_text(value.get("document_family", "")),
        "languages": _profile_text_list(value.get("languages", []), maximum=12),
        "english_selection_rule": clean_text(value.get("english_selection_rule", "")),
        "part_identifier_header": clean_text(value.get("part_identifier_header", "")),
        "part_identifier_rule": clean_text(value.get("part_identifier_rule", "")),
        "item_number_header": clean_text(value.get("item_number_header", "")),
        "item_number_rule": clean_text(value.get("item_number_rule", "")),
        "quantity_header": clean_text(value.get("quantity_header", "")),
        "quantity_rule": clean_text(value.get("quantity_rule", "")),
        "hierarchy": {
            "major_code_pattern": clean_text(hierarchy_source.get("major_code_pattern", "")),
            "subsection_code_pattern": clean_text(hierarchy_source.get("subsection_code_pattern", "")),
            "heading_detection_rule": clean_text(hierarchy_source.get("heading_detection_rule", "")),
            "continuation_rule": clean_text(hierarchy_source.get("continuation_rule", "")),
            "examples": examples,
        },
        "table_rules": _profile_text_list(value.get("table_rules", [])),
        "exclude_as_codes": _profile_text_list(value.get("exclude_as_codes", [])),
        "uncertainties": _profile_text_list(value.get("uncertainties", [])),
        "confidence": clamp_confidence(value.get("confidence"), fallback=0.0),
        "analysis_model": clean_text(model),
        "analyzed_pages": [int(page) for page in analyzed_pages],
    }
    meaningful = any(
        profile.get(key)
        for key in (
            "document_family", "languages", "part_identifier_header",
            "item_number_header", "table_rules",
        )
    ) or bool(examples)
    return profile if meaningful else {}


def analyze_document_profile_with_ai(
    api_key: str,
    model: str,
    extracted_pages: Sequence[tuple[int, str]],
    additional_instructions: str = "",
) -> tuple[dict[str, Any], list[str]]:
    """Create one evidence-based document profile before row extraction."""
    sample = _representative_profile_pages(extracted_pages)
    if not sample:
        raise ValueError("No OCR text was available for adaptive document analysis.")
    page_text = "\n\n".join(
        f"===== PDF PAGE {page} =====\n{markdown}"
        for page, markdown in sample
    )
    instruction_text = clean_text(additional_instructions)
    user_prompt = (
        "Analyze the following representative pages and return only the required "
        "document-profile JSON.\n\n"
    )
    if instruction_text:
        user_prompt += (
            "User-supplied manual notes (treat as guidance and verify against the OCR):\n"
            f"{instruction_text}\n\n"
        )
    user_prompt += page_text
    result = _mistral_json_request(
        api_key=api_key,
        model=clean_text(model) or "mistral-large-2512",
        messages=[
            {"role": "system", "content": DOCUMENT_PROFILE_SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        timeout_seconds=300,
        max_retries=3,
    )
    profile = _normalize_document_profile(
        result,
        clean_text(model) or "mistral-large-2512",
        [page for page, _ in sample],
    )
    if not profile:
        raise ValueError("Large-model analysis returned an empty document profile.")
    confidence = float(profile.get("confidence", 0.0))
    return profile, [
        "Adaptive document analysis completed with "
        f"{profile['analysis_model']} across {len(sample)} representative page(s) "
        f"(profile confidence {confidence:.0%})."
    ]


def _merge_document_profiles(
    primary: dict[str, Any] | None,
    verifier: dict[str, Any] | None,
) -> dict[str, Any]:
    """Combine Mistral's profile with an optional independent verifier."""
    base = dict(primary) if isinstance(primary, dict) else {}
    check = dict(verifier) if isinstance(verifier, dict) else {}
    if not check:
        return base
    if not base:
        return check

    base_confidence = clamp_confidence(base.get("confidence"), fallback=0.0)
    check_confidence = clamp_confidence(check.get("confidence"), fallback=0.0)
    verifier_is_strong = check_confidence >= max(0.60, base_confidence - 0.05)

    scalar_fields = (
        "document_family", "english_selection_rule", "part_identifier_header",
        "part_identifier_rule", "item_number_header", "item_number_rule",
        "quantity_header", "quantity_rule",
    )
    for field in scalar_fields:
        if clean_text(check.get(field, "")) and (
            verifier_is_strong or not clean_text(base.get(field, ""))
        ):
            base[field] = check[field]

    for field in ("languages", "table_rules", "exclude_as_codes", "uncertainties"):
        combined: list[Any] = []
        seen: set[str] = set()
        for item in list(base.get(field, []) or []) + list(check.get(field, []) or []):
            key = normalize_key(item)
            if key and key not in seen:
                combined.append(item)
                seen.add(key)
        base[field] = combined

    base_hierarchy = dict(base.get("hierarchy", {}) or {})
    check_hierarchy = dict(check.get("hierarchy", {}) or {})
    for field in (
        "major_code_pattern", "subsection_code_pattern",
        "heading_detection_rule", "continuation_rule",
    ):
        if clean_text(check_hierarchy.get(field, "")) and (
            verifier_is_strong or not clean_text(base_hierarchy.get(field, ""))
        ):
            base_hierarchy[field] = check_hierarchy[field]

    examples: list[dict[str, Any]] = []
    seen_examples: set[tuple[str, int | None]] = set()
    for example in list(base_hierarchy.get("examples", []) or []) + list(
        check_hierarchy.get("examples", []) or []
    ):
        if not isinstance(example, dict):
            continue
        page_value = quantity_to_number(example.get("page"))
        key = (
            normalize_key(example.get("code", "")),
            int(page_value) if page_value is not None else None,
        )
        if key[0] and key not in seen_examples:
            examples.append(example)
            seen_examples.add(key)
    base_hierarchy["examples"] = examples[:60]
    base["hierarchy"] = base_hierarchy
    base["confidence"] = max(base_confidence, check_confidence)
    base["analysis_model"] = " + ".join(
        dict.fromkeys(
            clean_text(value)
            for value in (base.get("analysis_model", ""), check.get("analysis_model", ""))
            if clean_text(value)
        )
    )
    base["analyzed_pages"] = sorted(
        {
            int(quantity_to_number(page))
            for page in list(base.get("analyzed_pages", []) or [])
            + list(check.get("analyzed_pages", []) or [])
            if quantity_to_number(page) is not None
        }
    )
    base["openai_verified"] = True
    return base


def verify_document_profile_with_openai(
    api_key: str,
    model: str,
    extracted_pages: Sequence[tuple[int, str]],
    existing_profile: dict[str, Any] | None = None,
    additional_instructions: str = "",
) -> tuple[dict[str, Any], list[str]]:
    """Optionally cross-check a document profile without ever blocking Mistral.

    Missing credentials, restricted models, exhausted quota, timeouts, malformed
    responses, and every other OpenAI-side failure return the original profile.
    """
    original = dict(existing_profile) if isinstance(existing_profile, dict) else {}
    if not str(api_key or "").strip():
        return original, [
            "Optional OpenAI verification was skipped because OPENAI_API_KEY is not configured."
        ]
    selected_model = clean_text(model) or "gpt-5.6-sol"
    try:
        sample = _representative_profile_pages(
            extracted_pages,
            max_pages=60,
            max_chars_per_page=5500,
        )
        if not sample:
            return original, [
                "Optional OpenAI verification was skipped because no OCR text was available."
            ]
        page_text = "\n\n".join(
            f"===== PDF PAGE {page} =====\n{markdown}"
            for page, markdown in sample
        )
        instruction_text = clean_text(additional_instructions)
        user_prompt = (
            "Independently verify the document layout and hierarchy in these OCR pages. "
            "A numeric value is a sub-machinery code only when the page visibly presents "
            "it as a section heading; ordinary item positions must not become sections. "
            "Return only the required JSON.\n\n"
        )
        if instruction_text:
            user_prompt += (
                "User notes (verify them against the OCR rather than assuming they are true):\n"
                f"{instruction_text}\n\n"
            )
        user_prompt += page_text
        result = _openai_json_request(
            api_key=api_key,
            model=selected_model,
            system_prompt=DOCUMENT_PROFILE_SYSTEM_PROMPT,
            user_prompt=user_prompt,
        )
        verified = _normalize_document_profile(
            result,
            selected_model,
            [page for page, _ in sample],
        )
        if not verified:
            raise ValueError("OpenAI returned an empty document profile.")
        merged = _merge_document_profiles(original, verified)
        confidence = float(verified.get("confidence", 0.0))
        return merged, [
            f"Optional OpenAI verification completed with {selected_model} across "
            f"{len(sample)} OCR page(s) (verification confidence {confidence:.0%})."
        ]
    except Exception as exc:
        # This is intentionally fail-open. Mistral OCR/extraction and the local
        # deterministic parser remain fully usable when OpenAI is unavailable.
        return original, [
            "Optional OpenAI verification was unavailable and was bypassed; "
            "Mistral processing continued normally. Details: "
            + _safe_api_error_text(exc, str(api_key or ""))
        ]


def _adaptive_profile_prompt(profile: dict[str, Any] | None) -> str:
    if not isinstance(profile, dict) or not profile:
        return ""
    prompt_profile = {
        key: profile.get(key)
        for key in (
            "document_family", "languages", "english_selection_rule",
            "part_identifier_header", "part_identifier_rule",
            "item_number_header", "item_number_rule", "quantity_header",
            "quantity_rule", "hierarchy", "table_rules", "exclude_as_codes",
            "uncertainties", "confidence",
        )
    }
    return (
        "ADAPTIVE DOCUMENT PROFILE (generated once from representative pages):\n"
        + json.dumps(prompt_profile, ensure_ascii=False, indent=2)
        + "\nUse this profile as document-wide guidance, but never override contradictory "
        "evidence on the current page or invent a value that is not printed."
    )


def _build_extraction_prompt(
    batch: Sequence[tuple[int, str]],
    additional_instructions: str,
    catalog_hint: str = "",
    profile_hint: str = "",
    adaptive_profile_hint: str = "",
) -> str:
    page_text = "\n\n".join(
        f"===== PAGE {page_number} =====\n{markdown}"
        for page_number, markdown in batch
    )
    prompt = (
        "Extract all genuine spare-parts rows from the following OCR markdown. "
        "Return only the required JSON object. If the pages contain no spare-parts "
        "rows, return {\"spare_parts\": []}.\n\n"
    )
    if profile_hint:
        prompt += profile_hint + "\n\n"
    if adaptive_profile_hint:
        prompt += adaptive_profile_hint + "\n\n"
    if catalog_hint:
        prompt += (
            "Authoritative section catalogue detected from this same PDF. Use these "
            "codes and English names whenever the current page belongs to one of them:\n"
            f"{catalog_hint}\n\n"
        )
    if clean_text(additional_instructions):
        prompt += (
            "Manual-specific instructions:\n"
            f"{clean_text(additional_instructions)}\n\n"
        )
    return prompt + page_text


def _normalize_batch_source_pages(
    batch_rows: Sequence[dict[str, Any]],
    batch: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    valid_pages = {int(page) for page, _ in batch}
    only_page = next(iter(valid_pages)) if len(valid_pages) == 1 else None
    normalized: list[dict[str, Any]] = []
    for item in batch_rows:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        parsed_page = quantity_to_number(row.get("source_page"))
        page_number = int(parsed_page) if parsed_page is not None else None
        if page_number not in valid_pages and only_page is not None:
            page_number = only_page
        if page_number is not None:
            row["source_page"] = page_number

        parsed_start = quantity_to_number(row.get("section_start_page"))
        section_start = int(parsed_start) if parsed_start is not None else None
        if section_start is not None and page_number is not None:
            section_start = min(section_start, page_number)
        row["section_start_page"] = section_start or page_number

        section_name = clean_text(
            row.get("section_name_english", row.get("detected_machinery", ""))
        ).upper()
        description = clean_text(
            row.get("description_english", row.get("description", ""))
        ).upper()
        ident_no = clean_text(row.get("ident_no", row.get("code", "")))
        legacy_part_no = clean_text(row.get("part_no", ""))
        section_code = clean_text(row.get("section_code", ""))
        if not ident_no and legacy_part_no and normalize_key(legacy_part_no) != normalize_key(section_code):
            ident_no = legacy_part_no

        row["section_code"] = section_code
        row["section_name_english"] = section_name
        row["detected_machinery"] = section_name
        row["section_maker"] = clean_text(row.get("section_maker", "")).upper()
        row["section_model"] = clean_text(row.get("section_model", "")).upper()
        row["description_english"] = description
        row["description"] = description
        row["ident_no"] = ident_no
        row["code"] = ident_no
        row["part_no"] = ident_no
        row["item_no"] = clean_text(row.get("item_no", ""))

        table_title = clean_text(row.get("table_title"))
        if not section_name and table_title and not _is_generic_machinery_name(table_title):
            row["section_name_english"] = table_title.upper()
            row["detected_machinery"] = table_title.upper()
        normalized.append(row)
    return normalized


def _propagate_detected_machinery_context(
    rows: Sequence[dict[str, Any]],
    max_page_gap: int = 2,
) -> list[dict[str, Any]]:
    """Carry a clear section heading onto nearby continuation rows conservatively."""
    result = [dict(row) for row in rows if isinstance(row, dict)]
    indexed = sorted(
        enumerate(result),
        key=lambda item: (
            quantity_to_number(item[1].get("source_page")) or 10**9,
            item[0],
        ),
    )
    last_detected = ""
    last_title = ""
    last_section_page: int | None = None
    last_source_page: int | None = None

    for _, row in indexed:
        source_value = quantity_to_number(row.get("source_page"))
        source_page = int(source_value) if source_value is not None else None
        detected = clean_text(row.get("detected_machinery"))
        title = clean_text(row.get("table_title"))
        start_value = quantity_to_number(row.get("section_start_page"))
        section_page = int(start_value) if start_value is not None else None

        if detected and not _is_generic_machinery_name(detected):
            last_detected = detected
            last_title = title or detected
            last_section_page = section_page or source_page
            last_source_page = source_page
            continue

        nearby = (
            source_page is not None
            and last_source_page is not None
            and 0 <= source_page - last_source_page <= max_page_gap
        )
        if nearby and last_detected:
            row["detected_machinery"] = last_detected
            if not title:
                row["table_title"] = last_title
            if section_page is None:
                row["section_start_page"] = last_section_page
            row["machinery_inherited"] = True
            last_source_page = source_page

    return result


def extract_spare_parts_with_ai(
    api_key: str,
    model: str,
    extracted_pages: Sequence[tuple[int, str]],
    pages_per_batch: int = 3,
    max_chars_per_batch: int = 12000,
    additional_instructions: str = "",
    document_profile: dict[str, Any] | None = None,
    coverage_model: str = "",
    progress: ProgressCallback | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Extract rows with automatic divide-and-retry recovery.

    If a multi-page response is malformed, the batch is split in half recursively.
    If a single page still fails, the local markdown-table parser is used for that
    page so one bad response does not discard the remainder of a large manual.
    """
    batches = _page_batches(
        extracted_pages,
        pages_per_batch=max(1, int(pages_per_batch)),
        max_chars=max(2000, int(max_chars_per_batch)),
    )
    rows: list[dict[str, Any]] = []
    messages: list[str] = []
    extraction_profile = _document_extraction_profile(extracted_pages)
    profile_hint = _profile_prompt(extraction_profile)
    adaptive_profile_hint = _adaptive_profile_prompt(document_profile)
    catalog_hint = _catalog_prompt_hint(
        build_section_catalog(extracted_pages).get("sections", [])
    )
    if extraction_profile == ENGINEERING_ARTICLE_DRAWING_PROFILE:
        messages.append(
            "Automatically detected an engineering drawing title-block Article No. catalogue. "
            "Document No./Title define the sub-machinery and only Article No. rows are treated as spares."
        )
    if extraction_profile == MULTILINGUAL_ORDER_CATALOG_PROFILE:
        messages.append(
            "Automatically detected a multilingual Order-No. catalogue. "
            "Order-No. is being mapped to PART NO/CODE, Pict./Photo to ITEM NO, "
            "and weight is being excluded from QNT."
        )

    def process_batch(
        batch: list[tuple[int, str]],
        label: str,
        depth: int = 0,
    ) -> None:
        try:
            result = _mistral_json_request(
                api_key=api_key,
                model=model,
                messages=[
                    {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": _build_extraction_prompt(
                            batch,
                            additional_instructions,
                            catalog_hint,
                            profile_hint,
                            adaptive_profile_hint,
                        ),
                    },
                ],
            )
            batch_rows = result.get("spare_parts", [])
            if not isinstance(batch_rows, list):
                raise ValueError("JSON did not contain a spare_parts list")
            normalized_batch_rows = _normalize_batch_source_pages(batch_rows, batch)

            # Some models return a valid but empty JSON object for dense historical
            # catalogues.  That is not an API failure, so the old recovery path never
            # ran. Retry once with an explicit coverage instruction before allowing
            # the deterministic Markdown parser to supplement the page later.
            coverage_profile = ""
            expected_direct_rows: list[dict[str, Any]] = []
            identifier_evidence_count = 0
            coverage_table_present = False
            if extraction_profile == MULTILINGUAL_ORDER_CATALOG_PROFILE:
                coverage_profile = "Order-No. catalogue"
                expected_direct_rows = _direct_table_rows(batch)
                identifier_evidence_count = _catalog_order_number_evidence_count(batch)
                coverage_table_present = any(
                    _is_multilingual_order_catalog(markdown)
                    for _, markdown in batch
                )
            elif extraction_profile == ENGINEERING_ARTICLE_DRAWING_PROFILE:
                coverage_profile = "Article-No. engineering table"
                expected_direct_rows = _direct_table_rows(batch)
                identifier_evidence_count = _engineering_article_identifier_evidence_count(batch)
                coverage_table_present = any(
                    _is_engineering_article_table(markdown)
                    for _, markdown in batch
                )

            expected_count = max(len(expected_direct_rows), identifier_evidence_count)
            coverage_is_sparse = bool(
                coverage_table_present
                and (
                    len(normalized_batch_rows) < 1
                    or (
                        expected_count >= 3
                        and len(normalized_batch_rows) < max(1, int(expected_count * 0.85))
                    )
                )
            )
            if coverage_is_sparse:
                expected_wording = f"approximately {expected_count}" if expected_count else "multiple"
                if extraction_profile == ENGINEERING_ARTICLE_DRAWING_PROFILE:
                    recovery_detail = (
                        "Article-No. records. Reconstruct the orderable table even when OCR "
                        "has emitted it in column-major reading order. Article No. is the "
                        "identifier; Name/Designation is the description; Art. Rev., "
                        "Material/Blank and Note are metadata and must not become ITEM NO or QNT."
                    )
                else:
                    recovery_detail = (
                        "Order-No. records. Read every table row and every Order-No. variant. "
                        "Do not summarize the page."
                    )
                recovery_instructions = "\n\n".join(
                    value
                    for value in (
                        clean_text(additional_instructions),
                        (
                            "COVERAGE RECOVERY: The prior pass returned too few rows even "
                            f"though source evidence indicates {expected_wording} {recovery_detail}"
                        ),
                    )
                    if value
                )
                try:
                    retry = _mistral_json_request(
                        api_key=api_key,
                        model=model,
                        messages=[
                            {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                            {
                                "role": "user",
                                "content": _build_extraction_prompt(
                                    batch,
                                    recovery_instructions,
                                    catalog_hint,
                                    profile_hint,
                                    adaptive_profile_hint,
                                ),
                            },
                        ],
                    )
                    retry_rows = retry.get("spare_parts", [])
                    if isinstance(retry_rows, list):
                        normalized_retry_rows = _normalize_batch_source_pages(
                            retry_rows, batch
                        )
                        if len(normalized_retry_rows) > len(normalized_batch_rows):
                            normalized_batch_rows = normalized_retry_rows
                            messages.append(
                                f"Recovered {len(normalized_batch_rows)} row(s) from "
                                f"catalogue page(s) {batch[0][0]}-{batch[-1][0]} after "
                                "automatic coverage retry."
                            )
                except Exception as recovery_error:
                    messages.append(
                        f"Coverage retry was unavailable for page(s) "
                        f"{batch[0][0]}-{batch[-1][0]}; existing extracted rows were "
                        f"retained. Details: {recovery_error}"
                    )

            still_sparse = bool(
                coverage_table_present
                and (
                    len(normalized_batch_rows) < 1
                    or (
                        expected_count >= 3
                        and len(normalized_batch_rows) < max(1, int(expected_count * 0.85))
                    )
                )
            )
            recovery_model = clean_text(coverage_model)
            if (
                still_sparse
                and recovery_model
                and normalize_key(recovery_model) != normalize_key(model)
            ):
                large_recovery_instructions = "\n\n".join(
                    value
                    for value in (
                        clean_text(additional_instructions),
                        (
                            "HIGH-ACCURACY COVERAGE RECOVERY: Earlier extraction passes "
                            "missed source-backed records on this page. Reconstruct every "
                            "orderable table record from the visible semantic columns, preserve "
                            "the printed hierarchy/title-block context, and never promote metadata "
                            "or construction/composition entries into spare parts."
                        ),
                    )
                    if value
                )
                try:
                    large_retry = _mistral_json_request(
                        api_key=api_key,
                        model=recovery_model,
                        messages=[
                            {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                            {
                                "role": "user",
                                "content": _build_extraction_prompt(
                                    batch,
                                    large_recovery_instructions,
                                    catalog_hint,
                                    profile_hint,
                                    adaptive_profile_hint,
                                ),
                            },
                        ],
                    )
                    large_rows = large_retry.get("spare_parts", [])
                    if isinstance(large_rows, list):
                        normalized_large_rows = _normalize_batch_source_pages(
                            large_rows, batch
                        )
                        if len(normalized_large_rows) > len(normalized_batch_rows):
                            normalized_batch_rows = normalized_large_rows
                            messages.append(
                                f"Recovered {len(normalized_large_rows)} row(s) from "
                                f"catalogue page(s) {batch[0][0]}-{batch[-1][0]} with "
                                f"the high-accuracy model {recovery_model}."
                            )
                except Exception as recovery_error:
                    messages.append(
                        f"High-accuracy coverage recovery was unavailable for page(s) "
                        f"{batch[0][0]}-{batch[-1][0]}; existing extracted rows were "
                        f"retained. Details: {recovery_error}"
                    )

            rows.extend(normalized_batch_rows)
            return
        except Exception as exc:
            if len(batch) > 1:
                midpoint = max(1, len(batch) // 2)
                left = batch[:midpoint]
                right = batch[midpoint:]
                page_range = f"{batch[0][0]}-{batch[-1][0]}"
                messages.append(
                    f"Recovered {label} (pages {page_range}) by automatically splitting "
                    "the malformed/oversized response into smaller requests."
                )
                process_batch(left, f"{label}.1", depth + 1)
                process_batch(right, f"{label}.2", depth + 1)
                return

            page_number = batch[0][0]
            fallback_rows = extract_spare_parts_from_markdown_tables(batch)
            rows.extend(fallback_rows)
            if fallback_rows:
                messages.append(
                    f"Page {page_number}: AI JSON remained invalid; the local table "
                    f"parser recovered {len(fallback_rows)} row(s)."
                )
            else:
                messages.append(
                    f"Page {page_number}: structured extraction failed and no local "
                    f"table rows were recoverable. Details: {exc}"
                )

    for batch_index, batch in enumerate(batches, start=1):
        if progress:
            progress(
                batch_index - 1,
                len(batches),
                f"Structuring candidate batch {batch_index}/{len(batches)}",
            )
        process_batch(list(batch), f"batch {batch_index}")
        if progress:
            progress(
                batch_index,
                len(batches),
                f"Structuring candidate batch {batch_index}/{len(batches)} complete",
            )

    # Carry clear headings onto nearby continuation pages, then avoid duplicate
    # recovery messages when recursive splitting happened repeatedly.
    rows = _propagate_detected_machinery_context(rows)
    deduplicated_messages = list(dict.fromkeys(messages))
    return rows, deduplicated_messages




ENGLISH_NORMALIZATION_SYSTEM_PROMPT = """
You are a strict technical-language normalizer for marine and industrial spare-parts
manuals. Return one JSON object with exactly this structure:
{
  "items": [
    {"id": "D0", "english": ""}
  ]
}

For every input item:
1. Return only the English technical term in UPPERCASE.
2. The source may contain German, English, and French concatenated because OCR removed
   line breaks. Isolate the English wording; do not return all languages together.
3. When an English phrase is already printed, reproduce that phrase faithfully rather
   than replacing it with a synonym or paraphrase.
4. When no English wording is present, translate the complete term into precise
   technical English.
5. Remove German and French wording from the result. Keep proper names, maker/model
   names, dimensions, standards, symbols, stage numbers, and forms such as 2/2-WAY.
6. Never change identifiers, codes, item numbers, quantities, or dimensions.
7. Return exactly one result for every supplied id and no additional commentary.

Examples:
KURBELGEHÄUSE CRANK CASE CARTER -> CRANK CASE
WELLENDICHTRING RADIAL PACKING RING BAGUE D'ETANCHEITE -> RADIAL PACKING RING
LEISTUNGSSCHILD MANUFACTURER'S NAME PLATE PLAQUE SIGNALETIQUE -> MANUFACTURER'S NAME PLATE
MANOMETER 1. STUFE PRESSURE GAUGE 1ST STAGE MANOMETRE 1ER ETAGE -> PRESSURE GAUGE 1ST STAGE
ÖLSTANDSAUGE MIT DICHTRING -> OIL LEVEL SIGHT GLASS WITH SEALING RING
""".strip()


def _english_normalization_batches(
    items: Sequence[dict[str, Any]],
    max_items: int = 40,
    max_chars: int = 22000,
) -> list[list[dict[str, Any]]]:
    batches: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_chars = 0
    for item in items:
        item_chars = len(json.dumps(item, ensure_ascii=False))
        if current and (
            len(current) >= max(1, int(max_items))
            or current_chars + item_chars > max(2000, int(max_chars))
        ):
            batches.append(current)
            current = []
            current_chars = 0
        current.append(item)
        current_chars += item_chars
    if current:
        batches.append(current)
    return batches


def enforce_english_only_with_ai(
    api_key: str,
    model: str,
    rows: Sequence[dict[str, Any]],
    progress: ProgressCallback | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Isolate printed English or translate unresolved text before review/export.

    Descriptions are sent only when they are foreign, flattened multilingual text,
    or not confidently English. Every unique sub-machinery title is normalized once
    and then propagated by section code so one section cannot receive several names.
    """
    normalized_rows = [dict(row) for row in rows if isinstance(row, dict)]
    if not normalized_rows:
        return normalized_rows, []

    key = _normalize_api_key(api_key)
    request_items: list[dict[str, Any]] = []
    description_targets: dict[str, int] = {}
    section_targets: dict[str, str] = {}
    section_item_by_key: dict[str, str] = {}
    section_rows: dict[str, list[int]] = {}

    for row_index, row in enumerate(normalized_rows):
        current_description = clean_text(
            row.get("description_english", row.get("description", ""))
        )
        source_description = _clean_markdown_cell(
            row.get("source_description_raw", current_description)
        )
        if (
            clean_text(row.get("description_source", "")).lower() != "printed english"
            and (
                _text_needs_english_normalization(source_description)
                or _text_needs_english_normalization(current_description)
            )
        ):
            item_id = f"D{row_index}"
            request_items.append(
                {
                    "id": item_id,
                    "kind": "spare_part_description",
                    "source_text": source_description,
                    "current_text": current_description,
                    "section_code": clean_text(row.get("section_code", "")),
                    "item_no": clean_text(row.get("item_no", "")),
                    "ident_no": clean_text(row.get("ident_no", row.get("code", ""))),
                }
            )
            description_targets[item_id] = row_index

        section_code = clean_text(row.get("section_code", ""))
        current_section = clean_text(
            row.get("section_name_english", row.get("detected_machinery", ""))
        )
        source_section = _clean_markdown_cell(
            row.get("source_section_name_raw", row.get("table_title", current_section))
        )
        section_key = normalize_key(section_code) or f"NAME{normalize_key(current_section)}"
        section_rows.setdefault(section_key, []).append(row_index)
        section_needs_cleanup = (
            _text_needs_english_normalization(source_section)
            or _text_needs_english_normalization(current_section)
        )
        if (
            section_needs_cleanup
            and section_key not in section_item_by_key
            and (source_section or current_section)
        ):
            item_id = f"S{len(section_item_by_key)}"
            request_items.append(
                {
                    "id": item_id,
                    "kind": "sub_machinery_title",
                    "source_text": source_section or current_section,
                    "current_text": current_section,
                    "section_code": section_code,
                }
            )
            section_item_by_key[section_key] = item_id
            section_targets[item_id] = section_key

    if not request_items:
        return normalized_rows, [
            "All descriptions and sub-machinery titles were already confidently English."
        ]

    batches = _english_normalization_batches(request_items)
    returned: dict[str, str] = {}
    failed_ids: set[str] = set()

    def process_batch(batch: list[dict[str, Any]], label: str) -> None:
        try:
            result = _mistral_json_request(
                api_key=key,
                model=model or "mistral-small-latest",
                messages=[
                    {"role": "system", "content": ENGLISH_NORMALIZATION_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": (
                            "Normalize the following technical terms. Return only the required JSON object.\n\n"
                            + json.dumps({"items": batch}, ensure_ascii=False)
                        ),
                    },
                ],
                timeout_seconds=300,
                max_retries=3,
            )
            result_items = result.get("items", [])
            if not isinstance(result_items, list):
                raise ValueError("English normalization JSON did not contain an items list")
            batch_ids = {str(item.get("id", "")) for item in batch}
            seen_ids: set[str] = set()
            for result_item in result_items:
                if not isinstance(result_item, dict):
                    continue
                item_id = clean_text(result_item.get("id", ""))
                english = clean_text(result_item.get("english", "")).upper()
                if item_id in batch_ids and english:
                    returned[item_id] = english
                    seen_ids.add(item_id)
            missing = batch_ids - seen_ids
            if missing:
                raise ValueError(
                    "English normalization omitted item(s): " + ", ".join(sorted(missing))
                )
        except Exception:
            if len(batch) > 1:
                midpoint = max(1, len(batch) // 2)
                process_batch(batch[:midpoint], f"{label}.1")
                process_batch(batch[midpoint:], f"{label}.2")
            else:
                failed_ids.add(str(batch[0].get("id", "")))

    for batch_index, batch in enumerate(batches, start=1):
        if progress:
            progress(
                batch_index - 1,
                len(batches),
                f"English-only cleanup {batch_index}/{len(batches)}",
            )
        process_batch(batch, f"English batch {batch_index}")
        if progress:
            progress(
                batch_index,
                len(batches),
                f"English-only cleanup {batch_index}/{len(batches)} complete",
            )

    cleaned_descriptions = 0
    cleaned_sections = 0
    unresolved = 0

    for item_id, row_index in description_targets.items():
        english = returned.get(item_id, "")
        if not english:
            normalized_rows[row_index]["language_review"] = True
            unresolved += 1
            continue
        normalized_rows[row_index]["description_english"] = english
        normalized_rows[row_index]["description"] = english
        still_suspect = (
            _looks_likely_non_english(english)
            or len(_language_segments(english)) > 1
        )
        normalized_rows[row_index]["language_review"] = bool(still_suspect)
        normalized_rows[row_index]["english_cleanup_applied"] = True
        normalized_rows[row_index]["description_source"] = "AI English isolation/translation"
        if still_suspect:
            normalized_rows[row_index]["confidence"] = min(
                clamp_confidence(normalized_rows[row_index].get("confidence", 0.60)),
                0.55,
            )
            unresolved += 1
        else:
            # Translation improves language certainty but must not hide unrelated
            # section or OCR uncertainty.
            normalized_rows[row_index]["confidence"] = max(
                clamp_confidence(normalized_rows[row_index].get("confidence", 0.70)),
                0.72,
            )
            cleaned_descriptions += 1

    for item_id, section_key in section_targets.items():
        english = returned.get(item_id, "")
        if not english:
            unresolved += 1
            continue
        still_suspect = (
            _looks_likely_non_english(english)
            or len(_language_segments(english)) > 1
        )
        if still_suspect:
            unresolved += 1
            continue
        for row_index in section_rows.get(section_key, []):
            normalized_rows[row_index]["section_name_english"] = english
            normalized_rows[row_index]["detected_machinery"] = english
            normalized_rows[row_index]["section_english_cleanup_applied"] = True
        cleaned_sections += 1

    messages = [
        f"English-only cleanup normalized {cleaned_descriptions} spare-part description(s).",
        f"English-only cleanup normalized {cleaned_sections} unique sub-machinery title(s).",
    ]
    if unresolved or failed_ids:
        messages.append(
            f"{max(unresolved, len(failed_ids))} English cleanup item(s) remain flagged for review."
        )
    return normalized_rows, messages


# ---------------------------------------------------------------------------
# Deterministic Benefit table parsing and section catalogue
# ---------------------------------------------------------------------------


def _clean_markdown_cell(value: Any) -> str:
    """Clean a Markdown table cell while preserving meaningful line breaks.

    Mistral commonly separates German / English / French variants with ``<br>``.
    Keeping those boundaries lets the deterministic layer select the exact printed
    English phrase instead of trusting an AI paraphrase.
    """
    raw = "" if value is None else str(value)
    raw = raw.replace("\x00", " ").replace("\\|", "|")
    raw = re.sub(r"<br\s*/?>", "\n", raw, flags=re.IGNORECASE)
    raw = re.sub(r"[`*_]+", "", raw)
    lines = [re.sub(r"\s+", " ", line).strip() for line in raw.splitlines()]
    return "\n".join(line for line in lines if line)


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip().strip("|")
    return [
        _clean_markdown_cell(cell)
        for cell in re.split(r"(?<!\\)\|", stripped)
    ]

def _is_markdown_separator(line: str) -> bool:
    cells = _split_markdown_row(line)
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells)


def _html_table_blocks(markdown: str) -> list[list[list[str]]]:
    """Parse simple HTML tables sometimes emitted by OCR instead of Markdown."""
    raw = str(markdown or "")
    blocks: list[list[list[str]]] = []
    for table_html in re.findall(r"<table\b[^>]*>(.*?)</table>", raw, flags=re.I | re.S):
        rows: list[list[str]] = []
        for row_html in re.findall(r"<tr\b[^>]*>(.*?)</tr>", table_html, flags=re.I | re.S):
            cells = re.findall(r"<(?:th|td)\b[^>]*>(.*?)</(?:th|td)>", row_html, flags=re.I | re.S)
            if cells:
                cleaned = [
                    _clean_markdown_cell(re.sub(r"<[^>]+>", " ", cell))
                    for cell in cells
                ]
                if any(cleaned):
                    rows.append(cleaned)
        if len(rows) >= 2:
            blocks.append(rows)
    return blocks


def _markdown_table_blocks(markdown: str) -> list[list[list[str]]]:
    """Return tabular blocks from strict Markdown, loose pipe tables, or HTML.

    OCR engines are inconsistent about emitting the Markdown separator row. The
    parser therefore accepts a loose pipe table when the first row looks like a
    semantic technical-table header and at least one following row has a similar
    number of cells. This improves coverage without making any manufacturer-specific
    assumptions.
    """
    lines = str(markdown or "").splitlines()
    blocks: list[list[list[str]]] = []
    consumed: set[int] = set()
    index = 0
    while index + 1 < len(lines):
        if "|" in lines[index] and _is_markdown_separator(lines[index + 1]):
            block = [_split_markdown_row(lines[index])]
            consumed.update({index, index + 1})
            index += 2
            while index < len(lines) and "|" in lines[index]:
                if not _is_markdown_separator(lines[index]):
                    block.append(_split_markdown_row(lines[index]))
                consumed.add(index)
                index += 1
            if len(block) >= 2:
                blocks.append(block)
        else:
            index += 1

    # Loose pipe tables without the usual |---|---| separator.
    index = 0
    while index < len(lines):
        if index in consumed or lines[index].count("|") < 2:
            index += 1
            continue
        start = index
        raw_rows: list[list[str]] = []
        while index < len(lines) and lines[index].count("|") >= 2:
            if index not in consumed and not _is_markdown_separator(lines[index]):
                raw_rows.append(_split_markdown_row(lines[index]))
            index += 1
        if len(raw_rows) < 2:
            continue
        first_joined = " ".join(clean_text(value).lower() for value in raw_rows[0])
        header_signal = any(
            token in first_joined
            for token in (
                "article no", "article number", "ident", "order-no", "order no",
                "part no", "part-no", "ref. no", "ref no", "item no", "position",
                "description", "designation", "part name", "quantity", "qty",
                "material/blank", "composition", "document no", "document number",
                "drawing no", "drawing number", "dwg no", "title", "sheet no",
                "revision no", "responsible department", "creator", "approved",
            )
        )
        if header_signal:
            blocks.append(raw_rows)
            consumed.update(range(start, index))

    blocks.extend(_html_table_blocks(markdown))
    return blocks


def _semantic_table_role(headers: Sequence[str]) -> str:
    """Classify a table before extracting rows from it."""
    keys = [re.sub(r"[^a-z0-9]+", "", clean_text(value).lower()) for value in headers]
    mapped = [_canonical_source_header(value) for value in headers]
    has_identifier = "ident_no" in mapped or "source_part_no" in mapped
    has_description = "description_raw" in mapped
    has_quantity = "quantity" in mapped
    composition = any("composition" in key for key in keys)
    material_blank = any(key in {"materialblank", "material", "blank"} for key in keys)
    revision = any(key in {"revision", "revisionno", "rev", "artrev"} for key in keys)

    if composition and not has_identifier:
        return "MATERIAL_COMPOSITION"
    if has_identifier and has_description:
        return "SPARE_PARTS"
    if revision and not has_identifier:
        return "REVISION"
    if material_blank and not has_identifier:
        return "TECHNICAL_DATA"
    if has_quantity and has_description and not has_identifier:
        return "SPECIFICATION"
    return "UNKNOWN"


def _canonical_source_header(header: str) -> str | None:
    key = re.sub(r"[^a-z0-9]+", "", clean_text(header).lower())
    if not key:
        return None
    if key in {"articleno", "articlenumber", "articlenr"}:
        return "ident_no"
    if any(
        token in key
        for token in ("orderno", "orderingno", "bestellnr", "nodecommande")
    ):
        return "ident_no"
    if "ident" in key or key in {"code", "partcode", "sparepartcode", "materialcode", "articlecode"}:
        return "ident_no"
    if any(token in key for token in ("description", "designation", "benennung", "partname", "denomination")):
        return "description_raw"
    if any(token in key for token in ("quantity", "qty", "qnt", "menge", "numberoff", "nooff")):
        return "quantity"
    if key in {
        "itemno", "itemnumber", "itemnr",
        "position", "positionno", "positionnumber", "posno", "posnr",
        "refno", "referenceno", "referencenumber", "indexno", "indexnumber",
    }:
        return "item_no"
    if any(token in key for token in ("bildpictphoto", "pictphoto", "picturephoto")):
        return "item_no"
    if any(token in key for token in ("partno", "partnumber", "teilnr", "catalogno", "orderingno", "stockno", "pno")):
        return "source_part_no"
    if key in {"unit", "uom", "unitofmeasure"}:
        return "unit"
    return None


def _find_data_header_index(rows: Sequence[Sequence[str]]) -> int | None:
    for row_index, row in enumerate(rows[:5]):
        joined = " ".join(clean_text(value).lower() for value in row)
        joined_compact = re.sub(r"\s+", "", joined)
        has_description = any(
            token in joined
            for token in ("description", "designation", "benennung", "part name")
        ) or "partname" in joined_compact
        has_standard_identifier = any(
            token in joined
            for token in (
                "ident", "code", "part-no", "part no", "item", "position",
                "article no", "article-no", "article number", "order-no", "order no", "bestell-nr", "bestell nr",
                "no de commande", "pict.", "photo",
            )
        )
        has_reference_no = bool(
            re.search(r"\b(?:ref(?:erence)?\.?\s*no\.?)\b", joined, flags=re.I)
        )
        # A generic tool list may also have a Ref. No. column. Treat Ref. No. as
        # a spare-parts identifier only when the same header explicitly says Part Name.
        reference_parts_header = has_reference_no and (
            "part name" in joined or "partname" in joined_compact
        )
        if has_description and (has_standard_identifier or reference_parts_header):
            return row_index
    return None


_REFERENCE_IDENTIFIER_RE = re.compile(
    r"^\s*(?P<prefix>[A-Z0-9][A-Z0-9._/]*?)\s*[-/]\s*(?P<item>\d+[A-Z]?)\s*$",
    flags=re.IGNORECASE,
)


def _is_reference_no_header(value: Any) -> bool:
    key = re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())
    return key in {"refno", "referenceno", "referencenumber"}


def _reference_identifier_parts(value: Any) -> tuple[str, str, str] | None:
    """Split a composite Ref. No. into full identifier, stable prefix and item suffix."""
    text = clean_text(value).upper()
    match = _REFERENCE_IDENTIFIER_RE.fullmatch(text)
    if not match:
        return None
    prefix = clean_text(match.group("prefix")).upper()
    item = clean_text(match.group("item")).upper()
    if len(normalize_key(prefix)) < 3 or not re.search(r"\d", prefix):
        return None
    return text, prefix, item


def _reference_identifier_columns(
    block: Sequence[Sequence[str]],
    header_index: int,
    headers: Sequence[str],
) -> set[int]:
    """Find Ref. No. columns that are clearly composite catalogue identifiers.

    This is deliberately conservative: at least three printed values must share
    one stable prefix (for example 03036-1, 03036-2, 03036-58A). A plain drawing
    reference column containing 1, 2, 3 therefore remains an ITEM NO column.
    """
    result: set[int] = set()
    for column_index, header in enumerate(headers):
        if not _is_reference_no_header(header):
            continue
        parsed: list[tuple[str, str, str]] = []
        for row in block[header_index + 1 : header_index + 24]:
            if column_index >= len(row):
                continue
            item = _reference_identifier_parts(row[column_index])
            if item is not None:
                parsed.append(item)
        if len(parsed) < 3:
            continue
        prefix_counts: dict[str, int] = {}
        for _, prefix, _ in parsed:
            key = normalize_key(prefix)
            prefix_counts[key] = prefix_counts.get(key, 0) + 1
        dominant = max(prefix_counts.values(), default=0)
        if dominant >= 3 and dominant / len(parsed) >= 0.75:
            result.add(column_index)
    return result


def _logical_table_column_groups(
    mappings: Sequence[str | None],
) -> list[tuple[int, int]]:
    """Split side-by-side repeated parts lists into independent logical records."""
    identifier_starts = [
        index for index, canonical in enumerate(mappings) if canonical == "ident_no"
    ]
    if len(identifier_starts) < 2:
        return [(0, len(mappings))]

    groups: list[tuple[int, int]] = []
    for position, start in enumerate(identifier_starts):
        end = identifier_starts[position + 1] if position + 1 < len(identifier_starts) else len(mappings)
        group = mappings[start:end]
        if "description_raw" not in group:
            return [(0, len(mappings))]
        groups.append((start, end))
    return groups


_HYPHEN_SECTION_CODE_RE = re.compile(
    r"\b(?:[A-Z]{1,6}-)?\d+[A-Z0-9.]*-\d+[A-Z0-9.]*-\d+[A-Z0-9.]*\b",
    flags=re.IGNORECASE,
)
_PAREN_SECTION_CODE_RE = re.compile(r"\(([A-Z]{1,6}\d{2,5})\)", flags=re.IGNORECASE)


def _section_code_tokens(value: Any, permissive: bool = False) -> list[str]:
    text = clean_text(value).upper()
    found: list[str] = []
    for match in _HYPHEN_SECTION_CODE_RE.findall(text):
        token = clean_text(match).upper()
        if token and token not in found:
            found.append(token)
    for match in _PAREN_SECTION_CODE_RE.findall(text):
        token = clean_text(match).upper()
        if token and token not in found:
            found.append(token)
    if permissive and not found:
        for token in re.findall(r"\b[A-Z]{1,5}\d{2,5}\b", text):
            # Avoid treating common model forms such as L350 as section codes unless
            # they are explicitly parenthesised or the whole cell is code-like.
            if token.startswith("L") and token[1:].isdigit() and len(token) <= 5:
                continue
            if token not in found:
                found.append(token)
    return found


def _join_section_codes(tokens: Sequence[str]) -> str:
    cleaned = [clean_text(token).upper() for token in tokens if clean_text(token)]
    return " / ".join(dict.fromkeys(cleaned))


def _language_segments(value: Any) -> list[str]:
    """Return visually separated language variants without splitting values like 2/2-way."""
    raw = _clean_markdown_cell(value)
    if not raw:
        return []
    line_parts = [line.strip() for line in raw.splitlines() if line.strip()]
    if len(line_parts) >= 2:
        return line_parts
    # Only split slashes surrounded by whitespace. Technical values such as
    # ``2/2-way`` and dimensions such as ``10/12 mm`` must remain intact.
    slash_parts = [
        clean_text(part)
        for part in re.split(r"\s+/\s+", raw)
        if clean_text(part)
    ]
    return slash_parts if len(slash_parts) >= 2 else [clean_text(raw)]


def _looks_likely_non_english(value: Any) -> bool:
    text = clean_text(value).lower()
    markers = (
        # German
        "schraube", "mutter", "dichtung", "gleitlager", "stufe", "zylinder",
        "halter", "kupplung", "scheibe", "leitung", "gehäuse", "gehaeuse",
        "kurbelgehäuse", "kurbelgehaeuse", "kurbelwelle", "wellendichtring",
        "leistungs", "drehrichtung", "einschraub", "ölstand", "oelstand",
        "dichtring", "passfeder", "spannhülse", "spannhuelse", "deckel",
        "lager", "welle", "stutzen", "druckluft", "sicherungsring",
        "sechskant", "manometer", "benennung", "abdeckung", "drossel",
        # French
        "pièce", "piece de", "soupape", "tuyau", "arbre", "carter", "écrou",
        "ecrou", "palier", "étage", "etage", "raccord", "support pour",
        "vis hexagonale", "joint de", "bague", "couvercle", "plaque",
        "flèche", "fleche", "rondelle", "clavette", "manchon", "vilebrequin",
        "étanch", "etanch", "graisseur", "ressort", "tôle", "tole",
        "désignation", "designation française", "quantité", "quantite",
        "brûleur", "bruleur", "moteur", " pour ", "réglage", "reglage",
        "brenner", "gebläse", "geblase", "stellantrieb", "luftregelung",
        "pumpen", "druckwächter", "druckwachter", "flammkopf", "düsen",
        "dusen", "zündtrafo", "zundtrafo", "ölvorwärmer", "olvorwarmer",
    )
    return any(marker in text for marker in markers)


def _looks_likely_english(value: Any) -> bool:
    text = clean_text(value).lower()
    if not text or _looks_likely_non_english(text):
        return False
    markers = (
        " for ", " with ", " and ", " of ", "stage", "support", "motor",
        "bolt", "nut", "ring", "valve", "pipe", "gauge", "piece", "hose",
        "bearing", "coupling", "mounting", "separator", "filter", "plate",
        "disc", "cylinder", "head", "pressure", "suction", "flywheel",
        "socket", "clip", "flange", "fitting", "thermometer", "gasket",
        "joint", "seal", "spring", "shaft", "cover", "piston", "liner",
        "screw", "washer", "nozzle", "bracket", "cartridge", "crank",
        "case", "radial", "packing", "inspection", "hole", "manufacturer",
        "name", "arrow", "direction", "oil", "level", "sight", "glass",
        "straight", "male", "union", "gear", "wheel", "complete", "dowel",
        "key", "cap", "tube", "sieve", "seat", "guide", "rod", "lever",
        "handle", "body", "housing", "insert", "bush", "bushing", "pin",
        "plug", "stud", "panel", "rubber", "metal", "foot", "inside",
        "outside", "adjustable", "protecting", "line", "air", "compressed",
        "elbow", "reduction", "reducing", "threaded", "filter", "cartridge",
        "burner", "cable", "casing", "individual", "regulation", "servo", "drive",
        "blower", "pump", "combustion", "ignition", "transformer", "preheater",
        "monitor", "terminal", "electrical", "equipment", "train",
    )
    padded = f" {text} "
    return any(marker in padded for marker in markers)


def _text_needs_english_normalization(value: Any) -> bool:
    """Return True when text is foreign, multilingual, or not confidently English."""
    raw = _clean_markdown_cell(value)
    text = clean_text(raw)
    if not text:
        return True
    if len(_language_segments(raw)) >= 2:
        return True
    if _looks_likely_non_english(text):
        return True
    # Accented Latin characters are a strong signal in German/French/English manuals.
    if re.search(r"[À-ÿ]", text):
        return True
    return not _looks_likely_english(text)

def _english_variant(value: Any) -> str:
    segments = _language_segments(value)
    if not segments:
        return ""
    if len(segments) >= 3:
        return clean_text(segments[1]).upper()
    if len(segments) == 2:
        first, second = segments
        if _looks_likely_non_english(first) and _looks_likely_english(second):
            return clean_text(second).upper()
    return clean_text(segments[0]).upper()


_CATALOG_HEADING_CODE_RE = re.compile(
    r"^\s*(\d{1,2}(?:\.\d{1,2})?)\.?\s*$",
    flags=re.IGNORECASE,
)
_SIMPLE_SECTION_HEADING_RE = re.compile(
    r"^\s*(\d{1,2}(?:\.\d{1,2})?)\.?\s+(.+?[A-Za-zÀ-ÿ].*)$",
    flags=re.IGNORECASE,
)


def _catalog_section_code(value: Any) -> str:
    """Keep a printed catalogue heading code as text, including trailing zeroes.

    In this particular multilingual catalogue layout, subsection *headings* use
    decade values (``3.30``, ``6.40`` and ``18.50``). OCR occasionally drops the
    final zero and returns ``3.3``.  This helper is used only after a value has
    already passed the heading test; spare item positions are never reformatted.
    """
    code = clean_text(value).strip().rstrip(".")
    match = re.fullmatch(r"(\d{1,2})\.(\d)", code)
    if match:
        return f"{match.group(1)}.{match.group(2)}0"
    return code


def _catalog_heading_title(value: Any) -> str:
    """Return only the leading heading from a vertically merged title cell.

    These catalogues frequently merge a section title with the first spare-part
    description.  A wrapped title is joined only when the grammar clearly
    continues (``Burner casing and`` + ``individual parts``); later notes such as
    ``When ordering...`` or child descriptions such as ``Blower wheel`` are not.
    """
    lines = [
        re.sub(_SIMPLE_SECTION_HEADING_RE, r"\2", line).strip(" -:;|")
        for line in _catalog_cell_lines(value)
    ]
    lines = [line for line in lines if line and re.search(r"[A-Za-zÀ-ÿ]", line)]
    if not lines:
        return ""

    title_parts = [lines[0]]
    for line in lines[1:]:
        current = title_parts[-1]
        repeated = normalize_key(line).startswith(normalize_key(title_parts[0]))
        continues = bool(
            re.search(r"(?:\b(?:and|or|for|of|with|without|to)|[-/])$", current, re.I)
            or re.match(r"^(?:and|or|for|of|with|without|to)\b", line, re.I)
        )
        if repeated or not continues:
            break
        title_parts.append(line)
    return clean_text(" ".join(title_parts)).strip(" -:;|")


def _clean_catalog_section_name(value: Any, section_code: Any = "") -> str:
    """Remove OCR spillover and code labels from an assembly heading."""
    text = clean_text(value).strip(" -:;|").upper()
    code = clean_text(section_code).strip().rstrip(".")
    if not text:
        return ""
    if code:
        text = re.sub(
            rf"\s*\(\s*{re.escape(code)}\s*\)\s*$", "", text, flags=re.I
        ).strip()
        text = re.sub(
            rf"^\s*{re.escape(code)}\.?\s+", "", text, flags=re.I
        ).strip()

    # Ordering notes and the first child description are commonly flattened onto
    # the heading. They are not part of the sub-machinery name.
    text = re.split(r"\bWHEN\s+ORDERING\b", text, maxsplit=1, flags=re.I)[0].strip()
    words = text.split()
    for length in range(min(5, len(words) // 2), 0, -1):
        if words[:length] == words[length : length * 2]:
            text = " ".join(words[:length])
            break
    return clean_text(text).strip(" -:;|").upper()


def _clean_catalog_maker(value: Any) -> str:
    """Reject German assembly headings that OCR/AI can mislabel as makers."""
    text = clean_text(value).upper()
    compact = normalize_key(text)
    false_makers = {
        "BRENNERMOTOR",
        "BRENNERGEHAUSE",
        "BRENNERGEHAEUSE",
        "LUFTREGELUNG",
        "STELLANTRIEB",
        "GEBLASE",
        "GEBLAESE",
    }
    if compact in false_makers:
        return ""
    return text


def _catalog_row_without_heading(
    row: Sequence[Any], mappings: Sequence[str | None], section: dict[str, str]
) -> list[Any]:
    """Remove a merged heading prefix while retaining the first spare row."""
    values = list(row)
    code_key = normalize_key(section.get("section_code", ""))
    for index, canonical in enumerate(mappings):
        if index >= len(values):
            continue
        if canonical in {"item_no", "source_part_no"}:
            lines = _catalog_cell_lines(values[index])
            if lines:
                match = _CATALOG_HEADING_CODE_RE.fullmatch(lines[0])
                if match and normalize_key(match.group(1)) == code_key:
                    values[index] = "\n".join(lines[1:])
        elif canonical == "description_raw":
            lines = _catalog_cell_lines(values[index])
            title = _catalog_heading_title(values[index])
            if not lines or not title:
                continue
            consumed = 0
            rebuilt = ""
            for consumed_count in range(1, min(3, len(lines)) + 1):
                rebuilt = clean_text(" ".join(lines[:consumed_count]))
                if normalize_key(rebuilt) == normalize_key(title):
                    consumed = consumed_count
                    break
            if consumed:
                values[index] = "\n".join(lines[consumed:])
    return values


def _catalog_heading_code(value: Any) -> tuple[str, bool]:
    """Return a printed hierarchy code and whether the cell is clearly a heading."""
    lines = _catalog_cell_lines(value)
    if not lines:
        return "", False
    embedded = _SIMPLE_SECTION_HEADING_RE.match(lines[0])
    if embedded:
        return embedded.group(1), True

    codes: list[str] = []
    for line in lines:
        match = _CATALOG_HEADING_CODE_RE.fullmatch(line)
        if match:
            codes.append(match.group(1))
    if not codes:
        return "", False
    # ``1.`` is an unambiguous major heading. Repeated positions such as
    # ``12.1 / 12.1 / 12.1`` are applicability variants of one spare, not a
    # hierarchy. Decimal subsection headings in this catalogue use decade
    # boundaries (3.30, 6.40, 18.50) followed by distinct child positions.
    major_heading = bool(re.fullmatch(r"\s*\d{1,2}\.\s*", lines[0]))
    raw_code = codes[0]
    first_parts = raw_code.split(".")
    distinct_codes = list(dict.fromkeys(codes))
    decimal_heading = False
    if len(first_parts) == 2 and first_parts[1].isdigit():
        major, position = int(first_parts[0]), int(first_parts[1])
        # OCR can read a section heading such as 6.40 as 6.4. Treat the
        # one-digit form as its decade heading only when subsequent printed
        # drawing positions prove that it starts a subsection.
        heading_position = position * 10 if len(first_parts[1]) == 1 else position
        decimal_heading = bool(
            heading_position >= 10
            and heading_position % 10 == 0
            and any(
                len(parts := candidate.split(".")) == 2
                and parts[0].isdigit()
                and parts[1].isdigit()
                and int(parts[0]) == major
                and int(parts[1]) > heading_position
                for candidate in distinct_codes[1:]
            )
        )
    return (
        _catalog_section_code(raw_code) if decimal_heading else raw_code,
        bool(major_heading or decimal_heading),
    )


def _classic_catalog_section_from_row(
    row: Sequence[Any], mappings: Sequence[str | None] | None = None
) -> dict[str, str]:
    """Extract one major/subsection heading from a multilingual catalogue row."""
    values = list(row)
    code = ""
    strong_heading = False
    english = ""

    if mappings:
        item_indexes = [
            index
            for index, canonical in enumerate(mappings)
            if canonical in {"item_no", "source_part_no"}
        ]
        for index in item_indexes:
            if index >= len(values):
                continue
            candidate_code, candidate_is_heading = _catalog_heading_code(values[index])
            if candidate_code:
                code = candidate_code
                strong_heading = candidate_is_heading
                break

        ident_indexes = [
            index for index, canonical in enumerate(mappings) if canonical == "ident_no"
        ]
        has_order_number_text = any(
            index < len(values) and bool(clean_text(values[index]))
            for index in ident_indexes
        )
        # A decimal drawing position with an Order-No. is a spare row, not a
        # subsection. Merged heading/child rows remain valid when their item cell
        # proves that a heading precedes one or more child positions.
        if not code or (has_order_number_text and not strong_heading):
            return {}
        if strong_heading or (
            not has_order_number_text
            and re.fullmatch(r"\d{1,2}\.\d", clean_text(code).rstrip("."))
        ):
            code = _catalog_section_code(code)

        english_index = _classic_english_description_index(values, mappings)
        if english_index is not None and english_index < len(values):
            english = _catalog_heading_title(values[english_index])

    raw_candidates: list[str] = []
    for cell in values:
        segments = _language_segments(cell)
        for segment in segments:
            match = _SIMPLE_SECTION_HEADING_RE.match(segment)
            if match:
                if not code:
                    code = match.group(1)
                title = clean_text(match.group(2))
                if title:
                    raw_candidates.append(title)
            elif not mappings and re.fullmatch(r"\s*\d{1,2}\.\s*", segment):
                code = re.sub(r"\D", "", segment)
            elif code and re.search(r"[A-Za-zÀ-ÿ]", segment):
                raw_candidates.append(clean_text(segment))

    if not code:
        joined = " ".join(clean_text(cell) for cell in values if clean_text(cell))
        match = _SIMPLE_SECTION_HEADING_RE.match(joined)
        if match:
            code = match.group(1)
            raw_candidates.append(clean_text(match.group(2)))

    candidates: list[str] = []
    for value in raw_candidates:
        cleaned = clean_text(value).strip(" -:;|")
        lowered = cleaned.lower()
        if not cleaned or any(
            token in lowered
            for token in (
                "order-no", "bestell", "no de commande", "designation",
                "bezeichnung", "burner serie", "type bruleur", "type brûleur",
                "appr. kg", "env. kg",
            )
        ):
            continue
        if cleaned not in candidates:
            candidates.append(cleaned)

    if not english:
        english_candidates = [
            value
            for value in candidates
            if _looks_likely_english(value) and not _looks_likely_non_english(value)
        ]
        if english_candidates:
            english = english_candidates[0]
        elif len(candidates) >= 3:
            english = candidates[1]
        elif candidates:
            english = _english_variant("\n".join(candidates)) or candidates[0]

    if not code or not english:
        return {}
    return {
        "section_code": code,
        "section_name_raw": "\n".join(candidates) or english,
        "section_name_english": _clean_catalog_section_name(english, code),
    }


def _classic_catalog_sections(markdown: str) -> list[dict[str, str]]:
    """Return every numbered section heading, including two sections on one page."""
    if not _is_multilingual_order_catalog(markdown):
        return []
    result: list[dict[str, str]] = []
    seen: set[str] = set()

    # OCR sometimes places the real section heading immediately above the
    # Markdown table instead of inside its first row. Read those standalone
    # headings even when the same page also contains another valid table heading.
    for line in str(markdown or "").splitlines()[:100]:
        if "|" in line or line.lstrip().startswith("!["):
            continue
        cleaned = _clean_markdown_cell(line.lstrip("# "))
        if not cleaned:
            continue
        section = _classic_catalog_section_from_row([cleaned])
        code = section.get("section_code", "")
        if code and code not in seen:
            seen.add(code)
            result.append(section)

    for block in _markdown_table_blocks(markdown):
        header_index = _find_data_header_index(block)
        if header_index is None:
            continue
        headers = block[header_index]
        mappings = [_canonical_source_header(header) for header in headers]
        for row in block[header_index + 1 :]:
            section = _classic_catalog_section_from_row(row, mappings)
            code = section.get("section_code", "")
            if code and code not in seen:
                seen.add(code)
                result.append(section)
    return result


def _classic_catalog_section_metadata(markdown: str) -> dict[str, str]:
    """Return the first numbered section heading for page-level continuation."""
    sections = _classic_catalog_sections(markdown)
    return sections[0] if sections else {}


def _best_effort_english_description(value: Any) -> tuple[str, bool]:
    """Return uppercase English text and whether language review is still needed."""
    segments = _language_segments(value)
    if not segments:
        return "", True
    if len(segments) >= 3:
        return clean_text(segments[1]).upper(), False
    if len(segments) == 2:
        first, second = segments
        if _looks_likely_non_english(first) and _looks_likely_english(second):
            return clean_text(second).upper(), False
    text = clean_text(segments[0])
    # A single AI phrase is accepted only when it looks plausibly English. This
    # catches cases such as MANOMETER / STUFE being returned as the English line.
    needs_review = _looks_likely_non_english(text) or not _looks_likely_english(text)
    return text.upper(), needs_review


def _source_english_description(value: Any) -> tuple[str, bool]:
    """Return an exact English phrase printed in the source when it can be isolated."""
    segments = _language_segments(value)
    if not segments:
        return "", False
    if len(segments) >= 3:
        return clean_text(segments[1]).upper(), True
    if len(segments) == 2:
        first, second = segments
        if _looks_likely_non_english(first) and _looks_likely_english(second):
            return clean_text(second).upper(), True
    if len(segments) == 1 and _looks_likely_english(segments[0]):
        return clean_text(segments[0]).upper(), True
    return "", False


def _select_source_faithful_description(
    source_value: Any,
    ai_value: Any,
) -> tuple[str, bool, bool, str]:
    """Choose a description while keeping exact source English ahead of AI wording.

    Returns ``(description, needs_review, ai_disagreed, source_kind)``.
    """
    source_description, source_is_exact = _source_english_description(source_value)
    ai_description, ai_needs_review = _best_effort_english_description(ai_value)

    if source_is_exact and source_description:
        disagreed = bool(
            ai_description
            and SequenceMatcher(
                None,
                normalize_key(source_description),
                normalize_key(ai_description),
            ).ratio() < 0.92
        )
        # The printed English phrase is authoritative even when the AI produced a
        # fluent synonym or paraphrase.
        return source_description, False, disagreed, "printed English"

    if ai_description:
        return ai_description, bool(ai_needs_review), False, "AI English/translation"

    fallback = clean_text(source_value).upper()
    return fallback, True, False, "unresolved source text"

def _page_header_text(markdown: str, max_lines: int = 28) -> str:
    """Return only the page header / pre-table area used for section detection."""
    candidates: list[str] = []
    for block in _markdown_table_blocks(markdown):
        header_index = _find_data_header_index(block)
        if header_index is not None and header_index > 0:
            for row in block[:header_index]:
                candidates.extend(cell for cell in row if clean_text(cell))
            break

    if not candidates:
        for line in str(markdown or "").splitlines():
            cleaned = _clean_markdown_cell(line.lstrip("# "))
            if not cleaned or line.lstrip().startswith("!["):
                continue
            lowered = clean_text(cleaned).lower()
            if (
                any(token in lowered for token in ("description", "designation", "benennung"))
                and any(
                    token in lowered
                    for token in (
                        "ident", "part-no", "part no", "item", "order-no",
                        "order no", "bestell", "no de commande", "pict", "photo",
                    )
                )
            ):
                break
            candidates.append(cleaned)
            if len(candidates) >= max_lines:
                break
    return "\n".join(dict.fromkeys(clean_text(value) for value in candidates if clean_text(value)))


def _is_date_like_section_code(value: Any) -> bool:
    """Return True for values that are clearly calendar dates, not hierarchy codes.

    Automatic hierarchy detection must never promote drawing dates such as
    ``2018-05-31`` to sub-machinery codes. Manual user-entered codes are not touched.
    """
    text = clean_text(value).strip().upper()
    if not text:
        return False
    compact = re.sub(r"\s+", "", text)
    date_patterns = (
        r"(?:19|20)\d{2}[-/.](?:0?[1-9]|1[0-2])[-/.](?:0?[1-9]|[12]\d|3[01])",
        r"(?:0?[1-9]|[12]\d|3[01])[-/.](?:0?[1-9]|1[0-2])[-/.](?:19|20)\d{2}",
        r"(?:0?[1-9]|1[0-2])[-/.](?:0?[1-9]|[12]\d|3[01])[-/.]\d{2}",
        r"(?:0?[1-9]|[12]\d|3[01])[-/.](?:0?[1-9]|1[0-2])[-/.]\d{2}",
        r"(?:19|20)\d{6}",
    )
    return any(re.fullmatch(pattern, compact) for pattern in date_patterns)


def _valid_automatic_section_code(value: Any) -> str:
    """Normalize an automatically detected hierarchy code and reject metadata noise."""
    text = clean_text(value).strip(" :;|,").upper()
    if not text or _is_date_like_section_code(text):
        return ""
    normalized = normalize_key(text)
    # OCR frequently joins a drawing number to a printed pagination marker, for
    # example ``9025986_Page_5`` -> ``9025986PAPER5``. This identifies a page of a
    # multi-page document, not a new machinery code.
    if re.search(r"(?:PAGE|PAPER|SHEET)\d+(?:OF\d+)?$", normalized, flags=re.I):
        return ""
    if normalized in {
        "DATE", "REV", "REVISION", "REVISIONNO", "SHEET", "SHEETNO",
        "NOOFSHEETS", "PAGE", "PAGENO", "TITLE", "DOCUMENTNO",
        "DOCUMENTNUMBER", "DRAWINGNO", "DRAWINGNUMBER", "DWGNO",
    }:
        return ""
    return text


def _credible_section_maker(value: Any) -> str:
    """Return a plausible printed manufacturer or blank for shifted OCR prose."""
    text = clean_text(value).upper().strip(" -:;|,.")
    if not text or len(text) > 70 or len(text.split()) > 8:
        return ""
    if not re.search(r"[A-Z]", text):
        return ""
    normalized = normalize_key(text)
    if normalized in {
        "ALARMID", "ALARM", "ID", "DESCRIPTION", "DESIGNATION", "TITLE",
        "DRAWING", "DOCUMENT", "COMPONENT", "MODEL", "MANUFACTURER",
    }:
        return ""
    if any(
        phrase in text
        for phrase in (
            "STARTED BY", "SIGNAL FROM", "IS DIRECTED", "SEE DRAWING",
            "REPLACE ", "CHECK ", "BOOK NO", "PAGE ",
        )
    ):
        return ""
    return text


def _credible_section_model(value: Any) -> str:
    """Return a plausible equipment model or blank for narrative OCR fragments."""
    text = clean_text(value).upper().strip(" -:;|,.")
    if not text or len(text) > 90 or len(text.split()) > 12:
        return ""
    if not re.search(r"[A-Z0-9]", text):
        return ""
    if text.startswith(("STARTED ", "DIRECTED ", "WHEN ", "IF ", "SEE ", "REPLACE ", "CHECK ")):
        return ""
    if any(
        phrase in text
        for phrase in (
            "SIGNAL FROM", "BOOK NO", "PAGE ", "ALARM ID", "DESCRIPTION",
            "TOO COLD", "TOO HOT",
        )
    ):
        return ""
    return text


def _engineering_title_block_metadata(
    markdown: str,
    require_article_table: bool = True,
) -> dict[str, str]:
    """Extract labelled hierarchy metadata from an engineering drawing title block.

    The logic is manufacturer-neutral and deliberately label-aware. Explicit
    ``Document No./Drawing No./DWG No.`` and ``Title/Assembly/Equipment`` fields
    outrank nearby dates, revisions, sheet numbers, chapter headings, and OCR
    reading-order noise. Unlabelled standalone numbers are never accepted as a
    title-block code.
    """
    if require_article_table and not _is_engineering_article_table(markdown):
        return {}

    raw = str(markdown or "")

    def value_text(value: Any) -> str:
        return clean_text(value).strip(" :;|-\t")

    code_labels = {
        "documentno", "documentnumber", "docno", "docnumber",
        "drawingno", "drawingnumber", "dwgno", "dwgnumber",
    }
    title_labels = {"title", "assembly", "equipment", "drawingtitle"}
    metadata_labels = code_labels | title_labels | {
        "date", "sheet", "sheetno", "sheetnumber", "noofsheets",
        "numberofsheets", "revision", "revisionno", "revisionnumber", "rev",
        "responsibledepartment", "department", "creator", "createdby",
        "approved", "approvedby", "checker", "checkedby", "scale",
        "format", "project", "projectno", "bookno", "pageno", "page",
    }

    def key_of(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", value_text(value).lower())

    def code_candidate(value: Any) -> str:
        candidate = value_text(value)
        if not candidate or key_of(candidate) in metadata_labels:
            return ""
        candidate = re.sub(r"\s+", "", candidate).upper()
        if not re.fullmatch(r"[A-Z0-9][A-Z0-9._/-]{3,29}", candidate, flags=re.I):
            return ""
        if not re.search(r"\d", candidate):
            return ""
        if _is_date_like_section_code(candidate):
            return ""
        # Prevent ordinary section numbers and obvious sheet/revision values from
        # masquerading as drawing/document identifiers.
        if re.fullmatch(r"\d{1,3}(?:[._/-]\d{1,3}){0,2}", candidate):
            return ""
        return candidate

    def title_candidate(value: Any) -> str:
        candidate = value_text(value)
        if not candidate or key_of(candidate) in metadata_labels:
            return ""
        if _is_date_like_section_code(candidate):
            return ""
        if not re.search(r"[A-Za-zÀ-ÿ]", candidate):
            return ""
        if len(candidate) > 140:
            return ""
        if re.fullmatch(r"\d+(?:[ ._/-]\d+)*", candidate):
            return ""
        return candidate

    # Candidate tuples are (score, value, evidence). Highest score wins.
    code_candidates: list[tuple[int, str, str]] = []
    title_candidates: list[tuple[int, str, str]] = []

    blocks = _markdown_table_blocks(markdown)
    for block in blocks:
        for row_index, row in enumerate(block):
            cells = [value_text(cell) for cell in row]
            keys = [key_of(cell) for cell in cells]
            for column_index, label_key in enumerate(keys):
                is_code = label_key in code_labels
                is_title = label_key in title_labels
                if not (is_code or is_title):
                    continue

                target = code_candidates if is_code else title_candidates
                validator = code_candidate if is_code else title_candidate

                # 1) Label and value in adjacent cells on the same row.
                for offset, score in ((1, 130), (2, 112)):
                    candidate_index = column_index + offset
                    if candidate_index >= len(cells):
                        continue
                    # Stop once another known field label starts; do not jump across
                    # Date/Sheet/Revision cells to steal their values.
                    if keys[candidate_index] in metadata_labels:
                        break
                    candidate = validator(cells[candidate_index])
                    if candidate:
                        target.append((score, candidate, "same-row labelled field"))

                # 2) Common title-block layout: labels in one row, values directly below.
                if row_index + 1 < len(block):
                    below = block[row_index + 1]
                    if column_index < len(below):
                        candidate = validator(below[column_index])
                        if candidate:
                            target.append((126, candidate, "label-over-value field"))
                    if column_index + 1 < len(below):
                        candidate = validator(below[column_index + 1])
                        if candidate:
                            target.append((105, candidate, "offset label-over-value field"))

    # 3) Inline / line-oriented OCR. Restrict the captured value to the current
    # line so a label cannot consume a following Date or Revision field.
    code_pattern = re.compile(
        r"(?im)^\s*(?:Document|Drawing|DWG)\s*(?:No\.?|Number)\s*[:|\-]?\s*([^\n|]{1,40})"
    )
    for match in code_pattern.finditer(raw):
        candidate = code_candidate(match.group(1))
        if candidate:
            code_candidates.append((118, candidate, "inline labelled field"))

    title_pattern = re.compile(
        r"(?im)^\s*(?:Title|Assembly|Equipment|Drawing\s+Title)\s*[:|\-]?\s*([^\n|]{1,140})"
    )
    for match in title_pattern.finditer(raw):
        candidate = title_candidate(match.group(1))
        if candidate:
            title_candidates.append((118, candidate, "inline labelled field"))

    # 4) Column-major OCR often prints the label on one line and the value on the
    # next. Scan only a very small window and stop at the next recognized label.
    lines = [value_text(line) for line in raw.splitlines() if value_text(line)]
    for index, line in enumerate(lines):
        label_key = key_of(line)
        if label_key not in code_labels and label_key not in title_labels:
            continue
        target = code_candidates if label_key in code_labels else title_candidates
        validator = code_candidate if label_key in code_labels else title_candidate
        for distance in (1, 2):
            if index + distance >= len(lines):
                break
            next_line = lines[index + distance]
            if key_of(next_line) in metadata_labels:
                break
            candidate = validator(next_line)
            if candidate:
                target.append((116 - distance * 4, candidate, "next-line labelled field"))
                break

    # Repeated Name/Designation values provide useful corroboration when OCR has
    # placed more than one text value beside the explicit Title label. They never
    # create a title by themselves; they only help choose among labelled candidates.
    description_counts: dict[str, tuple[int, str]] = {}
    for block in blocks:
        header_index = _find_data_header_index(block)
        if header_index is None:
            continue
        headers = block[header_index]
        header_keys = {
            re.sub(r"[^a-z0-9]+", "", clean_text(header).lower())
            for header in headers
        }
        if not header_keys.intersection({"articleno", "articlenumber", "articlenr"}):
            continue
        mappings = [_canonical_source_header(header) for header in headers]
        try:
            ident_index = mappings.index("ident_no")
            description_index = mappings.index("description_raw")
        except ValueError:
            continue
        for row in block[header_index + 1:]:
            if ident_index >= len(row) or description_index >= len(row):
                continue
            identifier = clean_text(row[ident_index])
            description = clean_text(row[description_index])
            if not identifier or not description or not re.search(r"[A-Za-zÀ-ÿ]", description):
                continue
            if not re.search(r"\d{5}", identifier):
                continue
            key = normalize_key(description)
            if not key:
                continue
            count, representative = description_counts.get(key, (0, description))
            description_counts[key] = (count + 1, representative)

    dominant_description_key = ""
    if description_counts:
        key, (count, _) = max(description_counts.items(), key=lambda item: item[1][0])
        if count >= 2:
            dominant_description_key = key

    def best(candidates: list[tuple[int, str, str]], *, title_field: bool = False) -> tuple[str, str]:
        if not candidates:
            return "", ""
        # Prefer repeated agreement across extraction views when scores tie. For a
        # title field, repeated orderable-row descriptions can corroborate the same
        # explicitly labelled title (for example CABLE) without inventing a title.
        counts: dict[str, int] = {}
        for _, value, _ in candidates:
            counts[normalize_key(value)] = counts.get(normalize_key(value), 0) + 1

        def rank(item: tuple[int, str, str]) -> tuple[int, int, int]:
            base_score, value, _ = item
            value_key = normalize_key(value)
            corroboration = 24 if title_field and dominant_description_key and value_key == dominant_description_key else 0
            return (
                base_score + corroboration,
                counts.get(value_key, 0),
                -len(value),
            )

        score, value, evidence = max(candidates, key=rank)
        return value, evidence

    document_no, code_evidence = best(code_candidates)
    title, title_evidence = best(title_candidates, title_field=True)

    # Cross-check against the orderable article identifiers. On many engineering
    # drawings the spare article numbers share the drawing/document stem. This is
    # supportive evidence only; it never overrides a valid explicitly labelled code.
    article_identifiers = _engineering_article_identifier_evidence(markdown, "")
    article_stems = [
        re.split(r"\s+", clean_text(identifier), maxsplit=1)[0].upper()
        for identifier in article_identifiers
        if clean_text(identifier)
    ]
    repeated_stem = ""
    if article_stems:
        stem_counts = {stem: article_stems.count(stem) for stem in set(article_stems)}
        stem, count = max(stem_counts.items(), key=lambda item: item[1])
        if count >= 2 and re.search(r"\d{5}", stem) and not _is_date_like_section_code(stem):
            repeated_stem = stem

    if not document_no and repeated_stem:
        document_no = repeated_stem
        code_evidence = "repeated Article-No. stem fallback"

    if not document_no:
        return {}

    result = {
        "section_code": document_no,
        "section_code_source": "engineering drawing title block",
        "section_code_evidence": code_evidence,
    }
    if title:
        result.update(
            {
                "section_name_raw": title,
                "section_name_english": title.upper(),
                "section_name_evidence": title_evidence,
            }
        )
    return result


# Compatibility alias retained for previous app imports/tests.
def _alfa_laval_title_block_metadata(markdown: str) -> dict[str, str]:
    return _engineering_title_block_metadata(markdown)


def _page_metadata(page_number: int, markdown: str) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "page": int(page_number),
        "section_code": "",
        "section_name_raw": "",
        "section_name_english": "",
        "maker": "",
        "model": "",
        "header_text": "",
        "section_code_source": "",
    }
    classic_catalog = _is_multilingual_order_catalog(markdown)
    blocks = _markdown_table_blocks(markdown)
    for block in blocks:
        header_index = _find_data_header_index(block)
        if header_index is None or header_index <= 0:
            continue
        preheader_rows = block[:header_index]
        meta_row = max(
            preheader_rows,
            key=lambda row: sum(bool(clean_text(value)) for value in row),
        )
        nonempty = [value for value in meta_row if clean_text(value)]
        if nonempty:
            metadata["maker"] = clean_text(nonempty[0]).upper()

        tail = nonempty[-1] if nonempty else ""
        tail_tokens = _section_code_tokens(tail, permissive=True)
        if tail_tokens:
            metadata["section_code"] = _join_section_codes(tail_tokens)
            metadata["section_code_source"] = "page header cell"

        middle_candidates = nonempty[1:-1] if len(nonempty) >= 3 else nonempty[1:]
        title_candidates = []
        for value in middle_candidates:
            cleaned = _clean_markdown_cell(value)
            flat = clean_text(cleaned)
            if not re.search(r"[A-Za-zÀ-ÿ]", flat):
                continue
            if re.fullmatch(r"[A-Z]?\s*\d+(?:[ ._/-]\d+)*", flat, flags=re.I):
                continue
            if re.search(r"\b(?:TAFEL|TABLE|PLANCHE|DRAWING|DWG)\b", flat, flags=re.I):
                continue
            title_candidates.append(cleaned)
        if title_candidates:
            # Prefer the richest multilingual title cell rather than a short model label.
            metadata["section_name_raw"] = max(
                title_candidates,
                key=lambda value: (len(_language_segments(value)), len(clean_text(value))),
            )

        model_part = re.split(
            r"\b(?:TAFEL|TABLE|PLANCHE|DRAWING|DWG)\b",
            clean_text(tail),
            maxsplit=1,
            flags=re.I,
        )[0]
        metadata["model"] = clean_text(model_part).upper()
        break

    header_text = _page_header_text(markdown)
    metadata["header_text"] = header_text
    if not metadata["section_code"] and not classic_catalog:
        # Never search the spare-parts table body for a section code. Identifiers and
        # cross-references in the body caused the previous section to remain active.
        tokens = _section_code_tokens(header_text, permissive=True)
        if tokens:
            metadata["section_code"] = _join_section_codes(tokens)
            metadata["section_code_source"] = "page header area"

    text = str(markdown or "")
    if not metadata["maker"]:
        bold_candidates = re.findall(r"\*\*([^*]{2,60})\*\*", header_text or text)
        for candidate in bold_candidates:
            candidate_clean = clean_text(candidate).upper()
            if re.search(r"[A-Z]", candidate_clean) and not re.search(r"\d", candidate_clean):
                if candidate_clean not in {"DESCRIPTION", "DESIGNATION", "QUANTITY", "TABLE"}:
                    metadata["maker"] = candidate_clean
                    break

    if not metadata["section_name_raw"]:
        lines = [
            _clean_markdown_cell(line.lstrip("# "))
            for line in (header_text or text).splitlines()
            if clean_text(line.lstrip("# ")) and not line.lstrip().startswith("![")
        ]
        stop_index = next(
            (idx for idx, line in enumerate(lines) if re.search(r"\b(?:TAFEL|TABLE|PLANCHE)\b", clean_text(line), flags=re.I)),
            None,
        )
        title_lines = lines[:stop_index] if stop_index is not None else lines[:6]
        title_lines = [line for line in title_lines if not _section_code_tokens(line)]
        title_lines = [
            line for line in title_lines
            if not re.fullmatch(r"[A-Z]?\s*\d+(?:\s*-\s*[A-Z]?\s*\d+)*", clean_text(line), flags=re.I)
        ]
        if len(title_lines) >= 3:
            metadata["section_name_raw"] = "\n".join(title_lines[:3])
        elif title_lines:
            metadata["section_name_raw"] = title_lines[0]

    if _is_date_like_section_code(metadata.get("section_code", "")):
        metadata["section_code"] = ""
        metadata["section_code_source"] = ""

    metadata["section_name_english"] = _english_variant(metadata["section_name_raw"])
    if classic_catalog:
        classic_section = _classic_catalog_section_metadata(markdown)
        # In this layout, only a simple numbered heading is authoritative. Clear
        # false codes such as PG11/ASZ12 or an Order-No. read from the table body.
        metadata["section_code"] = classic_section.get("section_code", "")
        metadata["section_code_source"] = (
            "numbered catalogue section heading" if classic_section else ""
        )
        if classic_section:
            metadata["section_name_raw"] = classic_section["section_name_raw"]
            metadata["section_name_english"] = classic_section[
                "section_name_english"
            ]
        else:
            metadata["section_name_raw"] = ""
            metadata["section_name_english"] = ""
        # Words such as Brennermotor, Luftregelung and Stellantrieb are German
        # assembly titles in this layout, not manufacturer/model metadata. The
        # verified main-machinery maker/model are applied later as the fallback.
        metadata["maker"] = ""
        metadata["model"] = ""
    drawing_title_block = _engineering_title_block_metadata(markdown)
    if drawing_title_block:
        metadata.update(drawing_title_block)

    if not metadata["model"] and not classic_catalog:
        model_match = re.search(
            r"\b(?:COMPRESSOR|KOMPRESSOR|ENGINE|GENERATOR|PUMP)\s+([A-Z0-9][A-Z0-9 ._/-]{1,30})",
            header_text or text,
            flags=re.I,
        )
        if model_match:
            metadata["model"] = clean_text(model_match.group(1)).upper()
    return metadata


def _catalog_cell_lines(value: Any) -> list[str]:
    raw = _clean_markdown_cell(value)
    return [clean_text(line) for line in raw.splitlines() if clean_text(line)]


def _catalog_order_number_lines(value: Any) -> list[str]:
    """Return individual Order-No. values from a vertically merged catalogue cell."""
    result: list[str] = []
    for line in _catalog_cell_lines(value):
        candidate = re.sub(r"\s+[A-Z]$", "", line.strip(), flags=re.IGNORECASE)
        candidate = re.sub(r"\s+", " ", candidate).strip(" ;,|")
        # OCR can flatten several visual Order-No. lines into one text line. Split
        # the known manufacturer formats before applying the generic fallback.
        tokens = re.findall(
            r"(?<![\d/])(?:"
            # The third printed group is normally four digits, but this manual
            # also contains a genuine 4-3-3/slash family such as
            # ``0511 210 002/2``. Keep the slash mandatory for the three-digit
            # third group so flattened short codes are not joined accidentally.
            r"\d{3,4}\s+\d{3}\s+(?:\d{4}(?:/\d{1,2})?|\d{3}/\d{1,2})"
            r"|\d{9,12}(?:/\d{1,2})?"
            r"|\d{3,4}[ .]\d{3}"
            r"|\d{6}"
            r")(?![\d/])",
            candidate,
        )
        if tokens:
            for token in tokens:
                cleaned_token = re.sub(r"\s+", " ", token).strip(" .;,|")
                if cleaned_token and cleaned_token not in result:
                    result.append(cleaned_token)
            continue
        # Do not use a loose numeric fallback for this catalogue family. Values
        # such as ``5-8 9-11``, ``7 8/2`` and ``1102/7`` are applicability or
        # damaged OCR fragments, not Order-No. identifiers. The explicit formats
        # above cover the printed Weishaupt identifiers used by the source.
    return result


def _catalog_order_number_evidence_count(
    extracted_pages: Sequence[tuple[int, str]],
) -> int:
    """Count source Order-No. evidence without requiring a complete parsed row.

    A sparse AI response used to look acceptable when the deterministic parser
    also missed the description or hierarchy on a difficult page. Counting the
    identifier column independently gives recovery a source-backed coverage floor.
    """
    evidence: set[tuple[int, str]] = set()
    for page_number, markdown in extracted_pages:
        if not _is_multilingual_order_catalog(markdown):
            continue
        for block in _markdown_table_blocks(markdown):
            header_index = _find_data_header_index(block)
            if header_index is None:
                continue
            mappings = [
                _canonical_source_header(header) for header in block[header_index]
            ]
            ident_indexes = [
                index for index, canonical in enumerate(mappings)
                if canonical == "ident_no"
            ]
            for row in block[header_index + 1 :]:
                for index in ident_indexes:
                    if index >= len(row):
                        continue
                    for identifier in _catalog_order_number_lines(row[index]):
                        evidence.add((int(page_number), normalize_key(identifier)))
    return len(evidence)


def _classic_english_description_index(
    headers: Sequence[str], mappings: Sequence[str | None]
) -> int | None:
    description_indexes = [
        index for index, canonical in enumerate(mappings) if canonical == "description_raw"
    ]
    if not description_indexes:
        return None
    # German / English / French are printed from left to right. Bezeichnung is
    # normally the first description column and the English Designation is second.
    if len(description_indexes) >= 3:
        return description_indexes[1]
    if len(description_indexes) == 2:
        # The German Bezeichnung often shares the applicability column and is
        # therefore not classified as a description column. The two remaining
        # Designation columns are English first, French second.
        return description_indexes[0]
    return description_indexes[0]


def _is_engineering_article_identifier_value(value: Any, section_code: Any = "") -> bool:
    """Validate one orderable Article-No. value without relying on table layout."""
    text = re.sub(r"\s+", " ", clean_text(value)).strip(" .;,|:")
    if not text or normalize_key(text) == normalize_key(section_code):
        return False
    upper = text.upper()
    if any(token in upper for token in (
        "PAGE", "SHEET", "REV", "BOOK", "DATE", "RESCUE", "ARTICLE",
        "POSITION", " POS", "TITLE", "DOCUMENT", "DRAWING", "CREATOR",
        "APPROVED", "DEPARTMENT", "MATERIAL", "NOTE",
    )):
        return False
    if _is_date_like_section_code(text):
        return False
    first_token = re.split(r"\s+", text, maxsplit=1)[0]
    digits = sum(character.isdigit() for character in first_token)
    letters = sum(character.isalpha() for character in first_token)
    if digits < 5 or letters > digits:
        return False
    return bool(
        re.fullmatch(
            r"[A-Z0-9][A-Z0-9._/-]{5,19}(?:\s+[A-Z0-9._/-]{1,5})?",
            text,
            flags=re.I,
        )
    )


def _headerless_engineering_variant_rows(
    page_number: int,
    markdown: Any,
    metadata: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    """Recover repeated two-column article variants without semantic headers.

    Some equipment drawings print only ``article-family + size/revision`` beside
    the English designation, for example ``9007170 72/92 | Valve DN500 / A500``.
    There is no Article-No. header for the normal semantic parser to recognize.
    To keep this recovery high precision, at least three rows must share the same
    long source identifier stem; isolated numbers, drawing dimensions and callouts
    are ignored. The caller still requires a source-confirmed equipment drawing,
    so these rows inherit that page's exact title-block parent.
    """
    raw = str(markdown or "")
    if not raw.strip():
        return []
    active_metadata = dict(metadata or {})
    candidates: list[tuple[str, str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()

    def add_candidate(identifier: Any, description: Any) -> None:
        ident = re.sub(r"\s+", " ", clean_text(identifier)).strip(" .;,|:").upper()
        desc = re.sub(r"\s+", " ", clean_text(description)).strip(" .;,|:")
        if not _is_engineering_article_identifier_value(ident):
            return
        match = re.fullmatch(
            r"(?P<stem>[A-Z0-9][A-Z0-9._/-]{5,19})\s+"
            r"(?P<variant>[A-Z0-9._/-]{1,5})",
            ident,
            flags=re.I,
        )
        if not match:
            return
        stem = match.group("stem").upper()
        if not re.search(r"\d{5}", stem) or _is_date_like_section_code(stem):
            return
        if (
            len(desc) < 3
            or len(desc) > 100
            or not re.search(r"[A-Za-z]", desc)
            or re.match(
                r"^(?:article|document|drawing|sheet|revision|title|material|note|"
                r"dimension|weight|approved|creator)\b",
                desc,
                flags=re.I,
            )
        ):
            return
        pair = (normalize_key(ident), normalize_key(desc))
        if pair in seen_pairs:
            return
        seen_pairs.add(pair)
        candidates.append((stem, ident, desc))

    for raw_line in raw.splitlines():
        line = re.sub(r"^[#>*+\-\s]+", "", str(raw_line)).strip()
        if not line or _is_markdown_separator(line):
            continue

        # Prefer explicit cell boundaries emitted by OCR/Markdown.
        cells = [
            clean_text(cell).strip(" |:;-–—")
            for cell in re.split(r"\s*\|\s*|\t+|\s{2,}", line.strip(" |"))
            if clean_text(cell).strip(" |:;-–—")
        ]
        if len(cells) >= 2:
            add_candidate(cells[0], cells[1])

        # Aligned OCR occasionally collapses the two cells to single spaces.
        match = re.match(
            r"^\s*(?P<identifier>[A-Z0-9][A-Z0-9._/-]{5,19}\s+"
            r"[A-Z0-9._/-]{1,5})\s+(?P<description>[A-Za-z].+?)\s*$",
            line,
            flags=re.I,
        )
        if match:
            add_candidate(match.group("identifier"), match.group("description"))

    grouped: dict[str, list[tuple[str, str]]] = {}
    for stem, identifier, description in candidates:
        grouped.setdefault(stem, []).append((identifier, description))

    # Every accepted family must contain enough repeated structure to distinguish
    # an orderable variant list from incidental drawing labels.
    accepted_groups = {
        stem: values
        for stem, values in grouped.items()
        if len(values) >= 3
    }
    rows: list[dict[str, Any]] = []
    for stem, values in sorted(accepted_groups.items()):
        for identifier, description in values:
            rows.append(
                {
                    "source_page": int(page_number),
                    "section_code": active_metadata.get("section_code", ""),
                    "section_name_english": active_metadata.get(
                        "section_name_english", ""
                    ),
                    "section_maker": active_metadata.get("maker", ""),
                    "section_model": active_metadata.get("model", ""),
                    "table_title": active_metadata.get("section_name_raw", ""),
                    "ident_no": identifier,
                    "description_raw": description,
                    "quantity": None,
                    "item_no": "",
                    "confidence": 0.92,
                    "source_layout": (
                        "headerless two-column engineering variant table"
                    ),
                    "article_family": stem,
                }
            )
    return rows


def _is_headerless_engineering_variant_table(markdown: Any) -> bool:
    return bool(_headerless_engineering_variant_rows(0, markdown, {}))


def _engineering_article_identifier_evidence(
    markdown: Any,
    section_code: Any = "",
) -> list[str]:
    """Return source-backed Article-No. candidates from an engineering table page.

    This is deliberately layout-agnostic. OCR may emit rows horizontally, columns
    vertically, or flatten the whole table. A genuine Article-No. table is already
    required by the caller, so identifiers can be collected from the whole OCR text.
    The title-block document number itself is excluded.
    """
    if not _is_engineering_article_table(markdown):
        return []
    raw = str(markdown or "")
    section_key = normalize_key(section_code)
    candidates: list[str] = []
    seen: set[str] = set()

    # Prefer compound article numbers such as ``9007280 06``. The whitespace is
    # meaningful and prevents a standalone drawing/document number from being
    # mistaken for a spare identifier.
    patterns = (
        r"(?<![A-Z0-9])([A-Z0-9][A-Z0-9._/-]{5,19}\s+[A-Z0-9._/-]{1,5})(?![A-Z0-9])",
        r"(?<![A-Z0-9])([A-Z]{1,4}[-/]?\d{5,12}[A-Z0-9._/-]*)(?![A-Z0-9])",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, raw, flags=re.I):
            value = re.sub(r"\s+", " ", clean_text(match.group(1))).strip(" .;,|:")
            key = normalize_key(value)
            if not key or key == section_key or key in seen:
                continue
            if not _is_engineering_article_identifier_value(value, section_code):
                continue
            seen.add(key)
            candidates.append(value)
    return candidates


def _engineering_article_identifier_evidence_count(
    extracted_pages: Sequence[tuple[int, str]],
) -> int:
    evidence: set[tuple[int, str]] = set()
    for page_number, markdown in extracted_pages:
        if not _is_engineering_article_table(markdown):
            continue
        metadata = _engineering_title_block_metadata(markdown)
        for identifier in _engineering_article_identifier_evidence(
            markdown, metadata.get("section_code", "")
        ):
            evidence.add((int(page_number), normalize_key(identifier)))
    return len(evidence)


def _columnar_orderable_article_rows(
    page_number: int,
    markdown: str,
    metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    """Recover an Article-No. table when OCR emits columns instead of rows.

    Rotated engineering drawings are often OCR'd in column-major reading order:
    all Article No. values first, then revisions, then Name/Designation values.
    This routine reconstructs only when a semantic Article-No. table is present and
    there is enough source evidence to align identifiers with descriptions safely.
    It is manufacturer-neutral.
    """
    if not _is_engineering_article_table(markdown):
        return []

    section_code = clean_text(metadata.get("section_code", ""))
    identifiers = _engineering_article_identifier_evidence(markdown, section_code)
    if not identifiers:
        return []

    raw = str(markdown or "")
    lines = [_clean_markdown_cell(line) for line in raw.splitlines()]
    lines = [line for line in lines if line]

    header_patterns = {
        "article": re.compile(r"^\s*article\s*(?:no\.?|number|nr\.?)\s*$", re.I),
        "revision": re.compile(r"^\s*art\.?\s*rev\.?\s*$", re.I),
        "name": re.compile(r"^\s*name\s*/?\s*designation\s*$", re.I),
        "material": re.compile(r"^\s*material\s*/?\s*blank\s*$", re.I),
        "note": re.compile(r"^\s*note\s*$", re.I),
    }

    def header_kind(line: str) -> str:
        cleaned = clean_text(line)
        for kind, pattern in header_patterns.items():
            if pattern.fullmatch(cleaned):
                return kind
        return ""

    # Collect values that follow the Name/Designation header until the next known
    # semantic header. This handles the common column-major OCR representation.
    name_values: list[str] = []
    in_name_column = False
    for line in lines:
        kind = header_kind(line)
        if kind:
            in_name_column = kind == "name"
            continue
        if not in_name_column:
            continue
        value = clean_text(line).strip(" |;:")
        if not value or not re.search(r"[A-Za-zÀ-ÿ]", value):
            continue
        compact = normalize_key(value)
        if compact in {
            "PURCHASEDARTICLE", "BOUGHTOUTITEM", "STANDARDARTICLE",
            "MATERIALBLANK", "NOTE", "ARTREV",
        }:
            continue
        if re.fullmatch(r"[A-Z0-9][A-Z0-9._/-]{5,19}(?:\s+[A-Z0-9._/-]{1,5})?", value, re.I):
            continue
        name_values.append(value)

    # If OCR kept repeated values, align them directly. If it collapsed repeated
    # names to one value, repeat it only when the title block independently prints
    # the same title. This remains source-backed and avoids inventing descriptions.
    descriptions: list[str] = []
    if len(name_values) >= len(identifiers):
        descriptions = name_values[: len(identifiers)]
    elif len(name_values) == 1:
        title_name = clean_text(metadata.get("section_name_english", ""))
        if title_name and normalize_key(title_name) == normalize_key(name_values[0]):
            descriptions = [name_values[0]] * len(identifiers)

    if len(descriptions) != len(identifiers):
        return []

    rows: list[dict[str, Any]] = []
    for identifier, description in zip(identifiers, descriptions):
        rows.append(
            {
                "source_page": int(page_number),
                "section_code": section_code,
                "section_name_english": clean_text(metadata.get("section_name_english", "")).upper(),
                "section_maker": clean_text(metadata.get("maker", "")).upper(),
                "section_model": clean_text(metadata.get("model", "")).upper(),
                "table_title": clean_text(metadata.get("section_name_raw", "")),
                "section_start_page": int(page_number),
                "ident_no": clean_text(identifier),
                "source_part_no": "",
                "item_no": "",
                "description_raw": clean_text(description),
                "quantity": None,
                "unit": "",
                "confidence": 0.86,
                "source_layout": "column-major Article-No. table",
            }
        )
    return rows


def _plain_orderable_article_rows(
    page_number: int,
    markdown: str,
    metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    """Recover Article-No. rows when OCR emits aligned text instead of a table.

    This fallback is semantic rather than manufacturer-specific. It activates only
    when the page contains the Article No. + Name/Designation drawing-table pattern.
    It intentionally ignores revision, material/classification, note/length, and
    composition rows as import fields.
    """
    if not _is_engineering_article_table(markdown):
        return []
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    raw_lines = [line.rstrip() for line in str(markdown or "").splitlines()]
    article_re = re.compile(r"^\s*([A-Z0-9][A-Z0-9._/-]{5,19}(?:\s+[A-Z0-9._/-]{1,5})?)\s*$", re.I)

    def add(article: str, description: str) -> None:
        ident = clean_text(article)
        desc = clean_text(description)
        key = normalize_key(ident)
        if not ident or not desc or key in seen or not re.search(r"[A-Za-z]", desc):
            return
        if not re.search(r"\d", ident):
            return
        seen.add(key)
        rows.append({
            "source_page": int(page_number),
            "section_code": clean_text(metadata.get("section_code", "")),
            "section_name_english": clean_text(metadata.get("section_name_english", "")).upper(),
            "section_maker": clean_text(metadata.get("maker", "")).upper(),
            "section_model": clean_text(metadata.get("model", "")).upper(),
            "table_title": clean_text(metadata.get("section_name_raw", "")),
            "section_start_page": int(page_number),
            "ident_no": ident,
            "source_part_no": "",
            "item_no": "",
            "description_raw": desc,
            "quantity": None,
            "unit": "",
            "confidence": 0.88,
            "source_layout": "aligned text Article-No. table",
        })

    # Prefer column spacing preserved by OCR.
    for line in raw_lines:
        compact = re.sub(r"[^a-z0-9]+", "", line.lower())
        if any(token in compact for token in ("articleno", "namedesignation", "materialblank", "artrev")):
            continue
        cells = [clean_text(cell) for cell in re.split(r"\t+|\s{2,}", line.strip()) if clean_text(cell)]
        if len(cells) >= 3 and article_re.fullmatch(cells[0]):
            # The second cell is commonly Art. Rev.; the third is Name/Designation.
            description_index = 2 if len(cells) >= 3 and re.fullmatch(r"\d{1,4}", cells[1]) else 1
            if description_index < len(cells):
                add(cells[0], cells[description_index])

    # Fallback for OCR that collapses the row to single spaces but retains a clear
    # uppercase material/classification marker such as PURCHASED ARTICLE.
    if not rows:
        collapsed_re = re.compile(
            r"^\s*(?P<article>[A-Z0-9][A-Z0-9._/-]{5,19}(?:\s+[A-Z0-9._/-]{1,5})?)"
            r"\s+(?P<rev>\d{1,4})\s+(?P<name>.+?)\s+"
            r"(?P<class>PURCHASED\s+ARTICLE|BOUGHT[- ]OUT\s+ITEM|STANDARD\s+ARTICLE)"
            r"(?:\s+(?P<note>.*))?$",
            re.I,
        )
        for line in raw_lines:
            match = collapsed_re.match(line)
            if match:
                add(match.group("article"), match.group("name"))
    return rows



def _engineering_article_consensus_metadata(
    rows: Sequence[dict[str, Any]],
    page_number: int | None = None,
    markdown: str = "",
) -> dict[str, Any]:
    """Derive parent hierarchy from a coherent orderable Article-No. table.

    This is a generic evidence-reconciliation rule for engineering drawings. OCR
    can recover the orderable rows while failing to pair the nearby ``Document
    No.`` and ``Title`` labels. When several source-backed Article-No. values share
    one stable prefix, that prefix is strong evidence for the parent drawing code.
    A repeated spare description may identify the title only when it dominates the
    same source table. The rule never activates for ordinary parts catalogues,
    varied descriptions, dates, or a single isolated article.
    """
    source_rows = [dict(row) for row in rows if isinstance(row, dict)]
    if not source_rows:
        return {}

    def article_stem(value: Any) -> str:
        text_value = re.sub(r"\s+", " ", clean_text(value)).strip(" .;,|:").upper()
        match = re.fullmatch(
            r"([A-Z0-9][A-Z0-9._/-]{4,24})\s+([A-Z0-9._/-]{1,6})",
            text_value,
            flags=re.I,
        )
        if not match:
            return ""
        prefix = match.group(1).strip(" .;,|:").upper()
        if _is_date_like_section_code(prefix):
            return ""
        if not re.search(r"\d{5}", prefix):
            return ""
        if not _valid_automatic_section_code(prefix):
            return ""
        return prefix

    stem_rows: dict[str, list[dict[str, Any]]] = {}
    for row in source_rows:
        identifier = clean_text(
            row.get("ident_no", row.get("code", row.get("part_no", "")))
        )
        stem = article_stem(identifier)
        if stem:
            stem_rows.setdefault(stem, []).append(row)

    if not stem_rows:
        return {}
    stem, matching_rows = max(
        stem_rows.items(),
        key=lambda item: (len(item[1]), len(item[0])),
    )
    identifiable_rows = sum(len(values) for values in stem_rows.values())
    required = max(2, (identifiable_rows * 3 + 4) // 5)  # at least 60% agreement
    if len(matching_rows) < required:
        return {}

    # Prefer an explicitly labelled title if the title-block parser managed to
    # recover one. The consensus code can still replace an unrelated chapter code.
    labelled = _engineering_title_block_metadata(markdown) if markdown else {}
    labelled_title = clean_text(labelled.get("section_name_english", "")).upper()

    excluded_descriptions = {
        "PURCHASED ARTICLE", "BOUGHT OUT ITEM", "BOUGHT-OUT ITEM",
        "STANDARD ARTICLE", "MATERIAL", "BLANK", "MATERIAL/BLANK",
        "NOTE", "ART. REV.", "ARTICLE",
    }
    description_groups: dict[str, list[str]] = {}
    for row in matching_rows:
        description = clean_text(
            row.get(
                "description_raw",
                row.get("description_english", row.get("description", "")),
            )
        ).upper()
        if not description or description in excluded_descriptions:
            continue
        if not re.search(r"[A-Z]", description):
            continue
        key = normalize_key(description)
        if key:
            description_groups.setdefault(key, []).append(description)

    consensus_title = ""
    if description_groups:
        _, values = max(
            description_groups.items(),
            key=lambda item: (len(item[1]), -len(item[1][0])),
        )
        title_required = max(2, (len(matching_rows) * 3 + 4) // 5)
        if len(values) >= title_required:
            consensus_title = max(set(values), key=values.count)

    title = labelled_title or consensus_title
    # When a labelled title is long navigation/chapter text but every orderable
    # row independently repeats one concise designation, the table consensus is
    # more specific to the parent article family.
    if labelled_title and consensus_title:
        if normalize_key(labelled_title) != normalize_key(consensus_title):
            if len(consensus_title) < len(labelled_title) and len(matching_rows) >= 3:
                title = consensus_title

    result: dict[str, Any] = {
        "section_code": stem,
        "section_code_source": "engineering article-table consensus",
        "section_code_evidence": (
            f"{len(matching_rows)} source Article-No. rows share prefix {stem}"
        ),
        "confidence": 0.92,
    }
    if page_number is not None:
        result["page"] = int(page_number)
    if title:
        result.update(
            {
                "section_name_raw": title,
                "section_name_english": title,
                "section_name_evidence": (
                    "dominant Name/Designation value"
                    if consensus_title and normalize_key(title) == normalize_key(consensus_title)
                    else "labelled engineering title"
                ),
            }
        )
    return result

def _direct_table_rows(extracted_pages: Sequence[tuple[int, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for page_number, markdown in extracted_pages:
        page_row_start = len(rows)
        classic_catalog = _is_multilingual_order_catalog(markdown)
        article_drawing = _is_engineering_article_table(markdown)
        metadata = _page_metadata(int(page_number), markdown)
        for block in _markdown_table_blocks(markdown):
            header_index = _find_data_header_index(block)
            if header_index is None:
                continue
            headers = block[header_index]
            table_role = _semantic_table_role(headers)
            if table_role == "MATERIAL_COMPOSITION":
                continue
            if article_drawing:
                header_keys = {
                    re.sub(r"[^a-z0-9]+", "", clean_text(header).lower())
                    for header in headers
                }
                # On this drawing family only the Article No. table is orderable.
                # Pos./Designation/Composition describes cable/material construction.
                if not header_keys.intersection({"articleno", "articlenumber", "articlenr"}):
                    continue
            mappings = [_canonical_source_header(header) for header in headers]

            # Some OEM parts lists expose only Ref. No. / Part Name / Quantity.
            # When Ref. No. values share a stable composite prefix (03036-55,
            # 03036-58A, ...), the full reference is the source identifier and the
            # final suffix is the drawing ITEM NO.
            reference_identifier_columns = _reference_identifier_columns(
                block, header_index, headers
            )
            for column_index in reference_identifier_columns:
                mappings[column_index] = "ident_no"

            if classic_catalog:
                english_index = _classic_english_description_index(headers, mappings)
                mappings = [
                    (
                        canonical
                        if canonical != "description_raw" or index == english_index
                        else None
                    )
                    for index, canonical in enumerate(mappings)
                ]
            # When Ident-No. exists, a separate Part-No. column containing drawing
            # callouts is ITEM NO by the user's import mapping rule.
            if "ident_no" in mappings and "item_no" not in mappings and "source_part_no" in mappings:
                mappings[mappings.index("source_part_no")] = "item_no"

            column_groups = (
                [(0, len(mappings))]
                if classic_catalog
                else _logical_table_column_groups(mappings)
            )
            active_metadata = dict(metadata)
            last_english_description = ""
            last_description_item_no = ""

            for values in block[header_index + 1 :]:
                if classic_catalog:
                    row_section = _classic_catalog_section_from_row(values, mappings)
                    if row_section:
                        active_metadata.update(row_section)
                        last_english_description = ""
                        last_description_item_no = ""
                        ident_indexes = [
                            index
                            for index, canonical in enumerate(mappings)
                            if canonical == "ident_no"
                        ]
                        if not any(
                            index < len(values) and bool(clean_text(values[index]))
                            for index in ident_indexes
                        ):
                            continue
                        values = _catalog_row_without_heading(
                            values, mappings, row_section
                        )

                padded = list(values) + [""] * max(0, len(headers) - len(values))

                for group_start, group_end in column_groups:
                    record: dict[str, Any] = {
                        "source_page": int(page_number),
                        "section_code": active_metadata.get("section_code", ""),
                        "section_name_english": active_metadata.get("section_name_english", ""),
                        "section_maker": active_metadata.get("maker", ""),
                        "section_model": active_metadata.get("model", ""),
                        "table_title": active_metadata.get("section_name_raw", ""),
                        "confidence": 0.88,
                    }
                    for column_index in range(group_start, group_end):
                        canonical = mappings[column_index]
                        if not canonical or column_index >= len(padded):
                            continue
                        value = padded[column_index]
                        record[canonical] = (
                            _clean_markdown_cell(value)
                            if canonical in {"description_raw", "ident_no"}
                            else clean_text(value)
                        )
                        if column_index in reference_identifier_columns:
                            reference_parts = _reference_identifier_parts(value)
                            if reference_parts is not None:
                                full_identifier, _, item_suffix = reference_parts
                                record["ident_no"] = full_identifier
                                record["item_no"] = item_suffix
                                record["item_no_from_reference"] = True

                    if classic_catalog:
                        description_raw = _clean_markdown_cell(
                            record.get("description_raw", "")
                        )
                        item_no = clean_text(record.get("item_no", ""))
                        identifiers = _catalog_order_number_lines(
                            record.get("ident_no", "")
                        )
                        # Visually merged catalogue rows may print the English
                        # designation once beside several Order-No. lines. Carry it
                        # only within the same table/section and the same (or merged)
                        # drawing position; a new heading resets this state above.
                        if (
                            not description_raw
                            and identifiers
                            and last_english_description
                            and (
                                not item_no
                                or not last_description_item_no
                                or normalize_key(item_no)
                                == normalize_key(last_description_item_no)
                            )
                        ):
                            description_raw = last_english_description
                            record["description_raw"] = description_raw
                            record["description_inherited"] = True
                        if not description_raw or not identifiers:
                            continue
                        last_english_description = description_raw
                        if item_no:
                            last_description_item_no = item_no
                        for identifier in identifiers:
                            variant = dict(record)
                            variant["ident_no"] = identifier
                            variant["item_no"] = item_no
                            variant["quantity"] = None
                            variant["confidence"] = 0.92
                            rows.append(variant)
                        continue

                    description_raw = clean_text(record.get("description_raw", ""))
                    ident_no = clean_text(record.get("ident_no", ""))
                    item_no = clean_text(record.get("item_no", ""))
                    if article_drawing and ident_no and not _is_engineering_article_identifier_value(
                        ident_no, active_metadata.get("section_code", "")
                    ):
                        continue
                    if (
                        description_raw
                        and description_raw.upper() not in {"NOT USED", "NOT USED."}
                        and (ident_no or item_no)
                        and re.search(r"[A-Za-zÀ-ÿ]", description_raw)
                    ):
                        record["quantity"] = quantity_to_number(record.get("quantity"))
                        rows.append(record)

        # OCR sometimes exposes the correct semantic headers but not a formal
        # Markdown/HTML table. Recover aligned Article-No. rows deterministically.
        if article_drawing and not any(int(row.get("source_page") or 0) == int(page_number) for row in rows):
            rows.extend(_plain_orderable_article_rows(int(page_number), markdown, metadata))
        if article_drawing and not any(int(row.get("source_page") or 0) == int(page_number) for row in rows):
            rows.extend(_columnar_orderable_article_rows(int(page_number), markdown, metadata))

        # Reconcile the hierarchy only after the orderable rows are available. A
        # document-wide section heading such as 4.3.21 is navigation context, while
        # seven Article-No. values sharing 9007280 and the same CABLE designation
        # are direct source evidence for the drawing parent. This rule is generic
        # and applies to any coherent engineering Article-No. table.
        if article_drawing and len(rows) > page_row_start:
            page_rows = rows[page_row_start:]
            consensus = _engineering_article_consensus_metadata(
                page_rows, int(page_number), markdown
            )
            if consensus:
                for row in page_rows:
                    row["section_code"] = consensus.get("section_code", row.get("section_code", ""))
                    if clean_text(consensus.get("section_name_english", "")):
                        row["section_name_english"] = consensus["section_name_english"]
                        row["table_title"] = consensus.get(
                            "section_name_raw", consensus["section_name_english"]
                        )
                    row["section_start_page"] = int(page_number)
                    row["section_context_source"] = consensus.get(
                        "section_code_source", "engineering article-table consensus"
                    )
                    row["confidence"] = max(
                        clamp_confidence(row.get("confidence", 0.88)),
                        float(consensus.get("confidence", 0.92)),
                    )
    return rows

def _index_sections(extracted_pages: Sequence[tuple[int, str]]) -> tuple[list[dict[str, Any]], set[int]]:
    sections: list[dict[str, Any]] = []
    index_pages: set[int] = set()
    for page_number, markdown in extracted_pages:
        page_entries: list[dict[str, Any]] = []
        for block in _markdown_table_blocks(markdown):
            for row in block:
                if len(row) < 2:
                    continue
                code_tokens = _section_code_tokens(row[-1], permissive=True)
                title = clean_text(row[0])
                if not code_tokens or not re.search(r"[A-Za-zÀ-ÿ]", title):
                    continue
                page_entries.append(
                    {
                        "code": _join_section_codes(code_tokens),
                        "aliases": code_tokens,
                        "name": _english_variant(title),
                        "maker": "",
                        "model": "",
                        "pages": set(),
                    }
                )
        if len(page_entries) >= 4:
            index_pages.add(int(page_number))
            sections.extend(page_entries)
    unique: dict[str, dict[str, Any]] = {}
    for section in sections:
        key = normalize_key(section["code"])
        if key and key not in unique:
            unique[key] = section
    return list(unique.values()), index_pages


def _global_manual_maker_model(extracted_pages: Sequence[tuple[int, str]]) -> tuple[str, str]:
    maker = ""
    model = ""
    for page_number, markdown in extracted_pages[:10]:
        metadata = _page_metadata(int(page_number), markdown)
        if not maker and metadata.get("maker"):
            maker = clean_text(metadata["maker"]).upper()
        if not model and metadata.get("model"):
            model = clean_text(metadata["model"]).upper()
        if maker and model:
            break
    return maker, model


_DRAWING_SECTION_HEADING_RE = re.compile(
    r"^\s*\d{1,3}(?:\.\d{1,3}){1,5}\s+(.+?)\s*$",
    flags=re.IGNORECASE,
)
_DRAWING_CODE_PREFIX_RE = re.compile(
    r"^[A-Z0-9][A-Z0-9._/ -]{3,28}$",
    flags=re.IGNORECASE,
)
_DRAWING_COMPONENT_NOUNS = {
    "ASSEMBLY", "ACTUATOR", "BOILER", "CABINET", "CABLE", "CASE",
    "COMPRESSOR", "CONTROLLER", "COOLER", "COUPLING", "CYLINDER",
    "DEVICE", "DRIVE", "ENGINE", "EQUIPMENT", "FAN", "FILTER", "GAUGE",
    "GEARBOX", "GENERATOR", "GOVERNOR", "HEAD", "HEATER", "MACHINERY",
    "METER", "MODULE", "MONITOR", "MOTOR", "PANEL", "PISTON", "PUMP",
    "PURIFIER", "REACTOR", "REGULATOR", "SENSOR", "SEPARATOR", "SKID",
    "SLEEVE", "SWITCH", "SYSTEM", "TRANSMITTER", "TRANSFORMER", "TURBINE",
    "UNIT", "VALVE",
}


def _clean_drawing_component_title(value: Any) -> str:
    text = clean_text(value).strip(" -:;|")
    if not text:
        return ""
    # Parentheticals describing pagination/applicability are navigation metadata,
    # while short equipment acronyms such as (LDC) are part of the useful title.
    text = re.sub(
        r"\s*\((?:[^)]*\b(?:PAGES?|PREVIOUSLY|OPTIONAL|INCLUDING|INCL\.?)\b[^)]*)\)\s*",
        " ",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\s+", " ", text).strip(" -:;|")
    return _clean_machinery_name(text)


def _valid_drawing_heading_code(value: Any) -> str:
    """Return a high-precision source drawing/document code or blank.

    Drawing headings are hierarchy evidence only when the nearby token looks like a
    real engineering document identifier. Requiring at least five digits prevents TOC
    section numbers, item tags, words and OCR fragments from becoming machinery codes.
    """
    raw_text = ("" if value is None else str(value)).upper().strip(" .,:;|")
    # A native searchable-text layer may expose the internal page label rather
    # than the title-block Document No. (for example ``9004714_p1``). Do not accept
    # that compound token as a genuine machinery code. The drawing-heading parser
    # below may recover the base only when the same heading explicitly confirms a
    # multi-page drawing.
    if re.fullmatch(
        r"[A-Z0-9][A-Z0-9./-]{4,24}(?:_|\s+)P(?:AGE)?(?:_|\s*)\d{1,3}",
        raw_text,
        flags=re.I,
    ):
        return ""
    text = clean_text(value).upper().strip(" .,:;|")
    # Native PDF text occasionally splits a seven-digit drawing number after its
    # fifth digit (``90121 03``). Conversely, a six-or-more digit drawing number
    # followed by one isolated digit is normally the revision printed beside it.
    split_code = re.fullmatch(r"(\d{5})\s+(\d{2})", text)
    if split_code:
        text = "".join(split_code.groups())
    else:
        code_with_revision = re.fullmatch(r"(\d{6,10})\s+(\d)", text)
        if code_with_revision:
            text = code_with_revision.group(1)
    if not text or _is_date_like_section_code(text):
        return ""
    compact = normalize_key(text)
    if sum(character.isdigit() for character in compact) < 5:
        return ""
    if re.search(
        r"\b(?:BOOK|PAGE|SHEET|REV(?:ISION)?|DATE|DN\s*\d+|PN\s*\d+)\b",
        text,
        flags=re.I,
    ):
        return ""
    if re.fullmatch(
        r"(?:P|V|PT|TT|TS|LS|QT|PI|FIT|FIC|SV|XV|KS|M)\d{2,4}[-–—]\d+(?:\.X)?",
        text,
        flags=re.I,
    ):
        return ""
    words = re.findall(r"[A-Z]{2,}", text)
    if len(words) >= 2 or text.count("/") >= 2:
        return ""
    candidate = _valid_automatic_section_code(text)
    if not candidate or _chapter_like_section_code(candidate):
        return ""
    return candidate


def _drawing_heading_document_code(value: Any, heading_context: Any = "") -> str:
    """Return the source Document No. while excluding native page labels.

    ``<document>_p1`` is reduced to ``<document>`` only when the nearby component
    heading explicitly says that the drawing has multiple pages. A compact code
    such as ``9004714P1`` remains valid because it may genuinely be printed in a
    title block; there is no separator proving that ``P1`` is pagination.
    """
    text = ("" if value is None else str(value)).upper().strip(" .,:;|")
    page_label = re.fullmatch(
        r"(?P<base>[A-Z0-9][A-Z0-9./-]{4,24})(?:_|\s+)"
        r"P(?:AGE)?(?:_|\s*)?(?P<page>\d{1,3})",
        text,
        flags=re.I,
    )
    if not page_label:
        return _valid_drawing_heading_code(text)
    heading = clean_text(heading_context)
    if not re.search(r"\(\s*\d+\s+PAGES?\s*\)", heading, flags=re.I):
        return ""
    return _valid_drawing_heading_code(page_label.group("base"))


def _drawing_heading_sections(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    """Detect source-backed component drawing headings without manufacturer rules.

    A section is accepted only when a numbered component heading appears on the same
    drawing-oriented page as a nearby, code-like document/drawing identifier. TOC rows
    are rejected by their trailing page number and short chapter numbers cannot qualify
    as drawing codes. The function is intentionally conservative because a missed code
    can remain reviewable, while a fabricated hierarchy code would corrupt assignments.
    """
    results: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for page_number, raw_text in extracted_pages:
        text = str(raw_text or "")
        # Preserve the separator that proves ``_p1`` is a page label before the
        # general Markdown cleaner removes underscores. A genuine compact ``P1``
        # suffix is deliberately unchanged.
        text = re.sub(
            r"(?i)(?<![A-Z0-9])([A-Z0-9./-]*\d[A-Z0-9./-]*)_"
            r"P(?:AGE)?_?(\d{1,3})(?![A-Z0-9])",
            r"\1 P\2",
            text,
        )
        lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
        if not lines:
            continue
        context = " ".join(lines[:12]).upper()
        if not any(
            token in context
            for token in (
                "DRAWING", "DRAWINGS", "DIMENSION", "SCHEMATIC",
                "GENERAL ARRANGEMENT", "ASSEMBLY DRAWING", "PARTS DRAWING",
            )
        ):
            continue

        for index, line in enumerate(lines):
            heading_match = _DRAWING_SECTION_HEADING_RE.match(line)
            if not heading_match:
                continue
            raw_title = clean_text(heading_match.group(1))
            # Contents/index rows normally end with their destination page number.
            # A local drawing-page heading does not.
            if re.search(r"\s\d{1,4}\s*$", raw_title):
                continue
            # Native PDF text often wraps the final qualifier in a drawing heading.
            # Join only the immediately following lines until parentheses balance;
            # an unresolved heading remains rejected rather than being truncated.
            title_end = index
            while (
                raw_title.count("(") > raw_title.count(")")
                and title_end + 1 < min(len(lines), index + 3)
            ):
                title_end += 1
                raw_title = clean_text(f"{raw_title} {lines[title_end]}")
            if raw_title.count("(") != raw_title.count(")"):
                continue
            title = _clean_drawing_component_title(raw_title)
            if not title or _is_generic_machinery_name(title):
                continue
            title_key = normalize_key(title)
            if title_key in {
                "DRAWING", "DRAWINGS", "DIMENSIONDRAWING", "DIMENSIONDRAWINGS",
                "ELECTRICALDRAWING", "ELECTRICALDRAWINGS", "TECHNICALDATA",
            }:
                continue
            title_tokens = set(re.findall(r"[A-Z]+", title.upper()))
            if not (title_tokens & _DRAWING_COMPONENT_NOUNS):
                continue

            code = ""
            for following in lines[title_end + 1:title_end + 5]:
                # PDF text extraction can concatenate the drawing number directly
                # with "Book No." (e.g. 9024508Book No.9028195...). Strip it even
                # when there is no word boundary before BOOK.
                candidate_line = re.split(
                    r"BOOK\s+NO\.?", following, maxsplit=1, flags=re.IGNORECASE
                )[0].strip()
                # The source-code line itself must be code-like. This rejects wrapped
                # component prose such as "PT201-16, PI201-18, ..." even if its first
                # token looks like an identifier.
                if _DRAWING_SECTION_HEADING_RE.match(candidate_line):
                    continue
                if not _DRAWING_CODE_PREFIX_RE.fullmatch(candidate_line):
                    continue
                candidate = _drawing_heading_document_code(
                    candidate_line, raw_title
                )
                if not candidate:
                    continue
                code = candidate
                break
            if not code:
                continue

            key = (normalize_key(code), normalize_key(title))
            if key in seen:
                continue
            seen.add(key)
            results.append(
                {
                    "code": code,
                    "aliases": [code],
                    "name": title,
                    "maker": "",
                    "model": "",
                    "pages": {int(page_number)},
                    "source": "numbered component drawing heading",
                }
            )
    return results


def _unnumbered_equipment_drawing_sections(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    """Detect an equipment drawing whose page heading is not chapter-numbered.

    Some manuals print the concrete equipment heading above the drawing and place
    the genuine document number on the next line, while the rotated drawing title
    block is only partly available to OCR. This remains high-precision by requiring
    all three independent source signals on the same page:

    * explicit drawing/dimension context near the top of the page;
    * a concrete equipment/component heading (not a document-level drawing title);
    * a nearby standalone engineering code containing at least five digits.

    Equipment tags, dimensions, DN sizes, legend rows and drawing callouts cannot
    qualify because they are neither a concrete heading paired with an independently
    valid standalone document code nor accepted by ``_valid_drawing_heading_code``.
    """
    results: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    for page_number, raw_text in extracted_pages:
        source_text = re.sub(
            r"(?i)(?<![A-Z0-9])([A-Z0-9./-]*\d[A-Z0-9./-]*)_"
            r"P(?:AGE)?_?(\d{1,3})(?![A-Z0-9])",
            r"\1 P\2",
            str(raw_text or ""),
        )
        lines = [
            clean_text(line)
            for line in source_text.splitlines()
            if clean_text(line)
        ]
        if not lines:
            continue

        header_lines = lines[:18]
        context = " ".join(header_lines).upper()
        if not any(
            token in context
            for token in (
                "DRAWING", "DRAWINGS", "DIMENSION", "SCHEMATIC",
                "GENERAL ARRANGEMENT", "ASSEMBLY DRAWING", "PARTS DRAWING",
            )
        ):
            continue

        for index, line in enumerate(header_lines):
            # Numbered headings are handled by the established detector above.
            if _DRAWING_SECTION_HEADING_RE.match(line):
                continue
            if _DRAWING_CODE_PREFIX_RE.fullmatch(line) and _valid_drawing_heading_code(line):
                continue

            title = _clean_drawing_component_title(line)
            if not _is_equipment_component_title(title):
                continue

            code = ""
            # The printed document number normally follows the heading. Keep the
            # window deliberately narrow so a footer Book No. or unrelated drawing
            # reference elsewhere on the page cannot be borrowed as the parent code.
            for following in header_lines[index + 1:index + 6]:
                candidate_line = re.split(
                    r"BOOK\s+NO\.?", following, maxsplit=1, flags=re.IGNORECASE
                )[0].strip()
                if not candidate_line or _DRAWING_SECTION_HEADING_RE.match(candidate_line):
                    continue
                if not _DRAWING_CODE_PREFIX_RE.fullmatch(candidate_line):
                    continue
                candidate = _drawing_heading_document_code(
                    candidate_line, line
                )
                if candidate:
                    code = candidate
                    break
            if not code:
                continue

            key = (normalize_key(code), normalize_key(title))
            if key in seen:
                continue
            seen.add(key)
            results.append(
                {
                    "code": code,
                    "aliases": [code],
                    "name": title,
                    "maker": "",
                    "model": "",
                    "pages": {int(page_number)},
                    "source": "unnumbered equipment drawing heading",
                    "confidence": 0.93,
                }
            )
    return results


def _looks_like_spare_table_page(markdown: str) -> bool:
    text = clean_text(markdown).lower()
    has_description = any(
        token in text for token in ("description", "designation", "benennung", "part name")
    )
    has_standard_identifier = any(
        token in text
        for token in (
            "ident-nr", "ident-no", "ident no", "item no", "position",
            "part no", "part-no", "article no", "article-no", "article number",
            "code", "order-no", "order no",
            "bestell-nr", "bestell nr", "no de commande", "pict.",
        )
    )
    has_reference_no = bool(
        re.search(r"\b(?:ref(?:erence)?\.?\s*no\.?)\b", text, flags=re.I)
    )
    compact_text = re.sub(r"\s+", "", text)
    reference_parts_layout = has_reference_no and (
        "part name" in text or "partname" in compact_text
    )
    article_parts_layout = _is_engineering_article_table(markdown)
    has_identifier = has_standard_identifier or reference_parts_layout or article_parts_layout
    has_quantity = any(token in text for token in ("qty", "quantity", "menge", "qnt"))
    illustrated_ref_list = (
        "parts list for figure" in text and has_reference_no and has_quantity
    )
    if illustrated_ref_list:
        return True
    return has_description and has_identifier and (
        has_quantity
        or (reference_parts_layout and "parts list" in text)
        or article_parts_layout
        or _is_multilingual_order_catalog(markdown)
        or (has_standard_identifier and "|" in str(markdown or ""))
    )


def select_spare_table_pages(
    extracted_pages: Sequence[tuple[int, str]],
    fallback_pages: Sequence[tuple[int, str]] | None = None,
) -> list[tuple[int, str]]:
    """Select genuine parts-table pages while keeping OCR of all pages for section context."""
    _, index_pages = _index_sections(extracted_pages)
    selected = [
        item
        for item in extracted_pages
        if int(item[0]) not in index_pages and _looks_like_spare_table_page(item[1])
    ]
    if selected:
        return selected
    return list(fallback_pages if fallback_pages is not None else extracted_pages)



_EXPLICIT_SPARE_LABEL_RE = re.compile(
    r"\b(?:spare\s*part|sparepart|replacement\s*part)\s*(?:number|no\.?|nr\.?)\s*:",
    flags=re.IGNORECASE,
)
_RECOMMENDED_SPARES_RE = re.compile(
    r"\b(?:recommended\s+spare\s+parts?|spares?\s+recommended|spare\s+parts?\s+on\s+board)\b",
    flags=re.IGNORECASE,
)


def _looks_like_explicit_spare_source(markdown: str) -> bool:
    """Return True when a page explicitly declares orderable spare identifiers.

    This is intentionally independent of table syntax. Technical manuals often place
    genuine spares in maintenance prose/list layouts using labels such as
    ``Spare part number:``. Those pages are source candidates even when no Markdown
    table exists.
    """
    text = clean_text(markdown)
    if not text:
        return False
    label_count = len(_EXPLICIT_SPARE_LABEL_RE.findall(text))
    if label_count >= 2:
        return True
    if label_count >= 1 and _RECOMMENDED_SPARES_RE.search(text):
        return True
    # A single explicit spare number in a maintenance instruction is still useful,
    # provided the surrounding page speaks about replacement/maintenance rather than
    # merely mentioning a catalogue.
    lowered = text.lower()
    return label_count >= 1 and any(
        token in lowered
        for token in ("replace", "replacement", "maintenance", "service", "recommended")
    )


def classify_spare_source_pages(
    extracted_pages: Sequence[tuple[int, str]],
    fallback_pages: Sequence[tuple[int, str]] | None = None,
) -> tuple[list[tuple[int, str]], dict[str, list[int]]]:
    """Classify page-level spare evidence without forcing one document-wide pattern."""
    _, index_pages = _index_sections(extracted_pages)
    roles: dict[str, list[int]] = {
        "table_or_drawing": [],
        "explicit_spare_number_list": [],
        "fallback": [],
    }
    selected: list[tuple[int, str]] = []
    selected_pages: set[int] = set()
    for page, markdown in extracted_pages:
        page_int = int(page)
        if page_int in index_pages:
            continue
        is_table = _looks_like_spare_table_page(markdown)
        is_explicit = _looks_like_explicit_spare_source(markdown)
        if not (is_table or is_explicit):
            continue
        selected.append((page_int, markdown))
        selected_pages.add(page_int)
        if is_table:
            roles["table_or_drawing"].append(page_int)
        if is_explicit:
            roles["explicit_spare_number_list"].append(page_int)
    if not selected:
        selected = list(fallback_pages if fallback_pages is not None else extracted_pages)
        roles["fallback"] = [int(page) for page, _ in selected]
    return selected, roles

def _looks_like_spare_table_page(markdown: str) -> bool:
    if _looks_like_explicit_spare_source(markdown):
        return True
    text = clean_text(markdown).lower()
    has_description = any(
        token in text for token in ("description", "designation", "benennung", "part name")
    )
    has_standard_identifier = any(
        token in text
        for token in (
            "ident-nr", "ident-no", "ident no", "item no", "position",
            "part no", "part-no", "article no", "article-no", "article number",
            "code", "order-no", "order no",
            "bestell-nr", "bestell nr", "no de commande", "pict.",
        )
    )
    has_reference_no = bool(
        re.search(r"\b(?:ref(?:erence)?\.?\s*no\.?)\b", text, flags=re.I)
    )
    compact_text = re.sub(r"\s+", "", text)
    reference_parts_layout = has_reference_no and (
        "part name" in text or "partname" in compact_text
    )
    article_parts_layout = _is_engineering_article_table(markdown)
    has_identifier = has_standard_identifier or reference_parts_layout or article_parts_layout
    has_quantity = any(token in text for token in ("qty", "quantity", "menge", "qnt"))
    illustrated_ref_list = (
        "parts list for figure" in text and has_reference_no and has_quantity
    )
    if illustrated_ref_list:
        return True
    return has_description and has_identifier and (
        has_quantity
        or (reference_parts_layout and "parts list" in text)
        or article_parts_layout
        or _is_multilingual_order_catalog(markdown)
        or (has_standard_identifier and "|" in str(markdown or ""))
    )


def select_spare_table_pages(
    extracted_pages: Sequence[tuple[int, str]],
    fallback_pages: Sequence[tuple[int, str]] | None = None,
) -> list[tuple[int, str]]:
    """Backward-compatible selector for every page carrying spare-source evidence."""
    selected, _ = classify_spare_source_pages(extracted_pages, fallback_pages=fallback_pages)
    return selected


_EXPLICIT_SPARE_LABEL_RE = re.compile(
    r"\b(?:spare\s*part|sparepart|replacement\s+part)\s*(?:number|no\.?|nr\.?|code)\b\s*[:#-]?",
    flags=re.IGNORECASE,
)
_PROCUREMENT_PART_LABEL_RE = re.compile(
    r"\bpart\s*(?:number|no\.?)\b\s*[:#-]?",
    flags=re.IGNORECASE,
)
_EXPLICIT_SPARE_IDENTIFIER_RE = re.compile(
    r"(?<![A-Z0-9])([A-Z0-9][A-Z0-9./_-]{3,}(?:[ \t]+[A-Z0-9][A-Z0-9./_-]{0,5})?)(?![A-Z0-9])",
    flags=re.IGNORECASE,
)
_EXPLICIT_PARENT_NOUNS = {
    "ASSEMBLY", "BOILER", "CABINET", "COMPRESSOR", "CONTROLLER", "COOLER",
    "ENGINE", "EQUIPMENT", "FAN", "FILTER", "GEARBOX", "GENERATOR", "GOVERNOR",
    "HEATER", "MACHINERY", "MOTOR", "MODULE", "PANEL", "PUMP", "PURIFIER",
    "REACTOR", "SEPARATOR", "SYSTEM", "TURBINE", "UNIT", "VALVE",
}
_EXPLICIT_DESCRIPTION_SKIP_PREFIXES = (
    "BOOK NO", "PAGE ", "CONTENT:", "CONTENTS:", "NOTE:", "TIME INTERVAL",
    "INSTRUCTIONS", "MAINTENANCE SCHEDULE", "RECOMMENDED SPARE PARTS",
    "ONE OF EACH SPARE PART", "SPARE PART NUMBER", "SPAREPART NUMBER",
    "IMO REQUIREMENT", "EPA REQUIREMENT", "EVERY ", "ONCE ", "AFTER ",
    "RECOMMENDED TO", "SEE ", "IF ", "WHEN ", "RULE OF THUMB",
)


def _valid_explicit_spare_identifier(value: Any) -> str:
    text = clean_text(value).upper().strip(" .,:;|#")
    if not text or _is_date_like_section_code(text):
        return ""
    if re.search(r"\b(?:BOOK|PAGE|SHEET|REV(?:ISION)?)\b", text, flags=re.I):
        return ""
    compact = normalize_key(text)
    if len(compact) < 5 or len(compact) > 24:
        return ""
    if sum(character.isdigit() for character in compact) < 4:
        return ""
    if re.fullmatch(r"\d{1,4}", compact):
        return ""
    return text


def _explicit_spare_identifier_from_text(value: Any) -> str:
    text = clean_text(value).upper().lstrip(" :#-")
    if not text:
        return ""
    for match in _EXPLICIT_SPARE_IDENTIFIER_RE.finditer(text[:80]):
        candidate = _valid_explicit_spare_identifier(match.group(1))
        if candidate:
            return candidate
    return ""


def _looks_like_explicit_parent_heading(value: Any) -> bool:
    text = clean_text(value).strip(" |:-")
    if not text or len(text) > 90 or len(text.split()) > 8:
        return False
    if any(mark in text for mark in (",", ".", ";", "!", "?")):
        return False
    upper = text.upper()
    if any(upper.startswith(prefix) for prefix in _EXPLICIT_DESCRIPTION_SKIP_PREFIXES):
        return False
    if re.search(r"\b(?:HAS|HAVE|BEEN|WHEN|WHICHEVER|MUST|SHOULD)\b", upper):
        return False
    if _EXPLICIT_SPARE_LABEL_RE.search(text):
        return False
    if re.search(r"\b(?:CONTENT|QUANTITY|QTY|ARTICLE|PART\s+NO|ITEM\s+NO)\b", upper):
        return False
    tokens = {token for token in re.findall(r"[A-Z]+", upper)}
    return bool(tokens & _EXPLICIT_PARENT_NOUNS)


def _clean_explicit_spare_description(value: Any) -> str:
    text = clean_text(value).strip(" |:-")
    if not text:
        return ""
    text = re.split(r"\b(?:CONTENT|CONTENTS)\s*:", text, maxsplit=1, flags=re.I)[0]
    text = re.sub(r"^\d+\s+", "", text).strip()
    text = re.sub(
        r"^(?:REPLACE(?:MENT)?(?:\s+OF)?|CHANGE|RENEW|INSTALL)\s+",
        "",
        text,
        flags=re.I,
    ).strip()
    text = re.sub(r"\(\s*[A-Z]{1,5}\d{2,}[A-Z0-9–—._/-]*\s*\)\s*$", "", text, flags=re.I).strip()
    if any(text.upper().startswith(prefix) for prefix in _EXPLICIT_DESCRIPTION_SKIP_PREFIXES):
        return ""
    if not re.search(r"[A-Za-z]", text):
        return ""
    return clean_text(text)


def _explicit_description_from_fragments(fragments: Sequence[str]) -> str:
    cleaned: list[str] = []
    for fragment in fragments[-4:]:
        value = _clean_explicit_spare_description(fragment)
        if value:
            cleaned.append(value)
    if not cleaned:
        return ""
    # OCR often splits short noun phrases over two lines: "Quartz sleeve" / "set"
    # or "Temperature" / "transmitter". Join only a compact tail so surrounding
    # maintenance prose cannot leak into the spare description.
    while len(cleaned) > 1 and len(" ".join(cleaned)) > 90:
        cleaned.pop(0)
    combined = clean_text(" ".join(cleaned))
    combined = re.sub(r"\b(?:REPLACE|REPLACEMENT OF)\s+", "", combined, flags=re.I).strip()
    return combined


def _looks_like_explicit_spare_number_page(markdown: Any) -> bool:
    lines = [clean_text(line) for line in str(markdown or "").splitlines()]
    for index, line in enumerate(lines):
        label = _EXPLICIT_SPARE_LABEL_RE.search(line)
        if not label:
            bare_label = _PROCUREMENT_PART_LABEL_RE.search(line)
            context = clean_text(" ".join(lines[max(0, index - 6):index + 1]))
            if bare_label and re.search(
                r"\b(?:RECOMMEND(?:ATION|ED)?\s+TO\s+ORDER|ORDER\s+\d+|REPLACEMENT|REPLACE|RENEW)\b",
                context,
                flags=re.IGNORECASE,
            ):
                label = bare_label
        if not label:
            continue
        identifier = _explicit_spare_identifier_from_text(line[label.end():])
        if not identifier:
            for next_line in lines[index + 1:index + 3]:
                identifier = _explicit_spare_identifier_from_text(next_line)
                if identifier:
                    break
        if identifier:
            return True
    return False


def select_explicit_spare_number_pages(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[tuple[int, str]]:
    """Return pages containing clearly labelled spare-part identifiers outside tables."""
    return [
        (int(page), text)
        for page, text in extracted_pages
        if _looks_like_explicit_spare_number_page(text)
    ]


def select_spare_source_pages(
    extracted_pages: Sequence[tuple[int, str]],
    fallback_pages: Sequence[tuple[int, str]] | None = None,
) -> list[tuple[int, str]]:
    """Select all source-page families that can legitimately create spare rows.

    This is intentionally broader than ``select_spare_table_pages``. A technical
    manual may contain formal tables, illustrated lists, engineering drawings, and
    prose/list sections with an explicit ``Spare part number`` label in the same PDF.
    """
    table_pages = select_spare_table_pages(extracted_pages, fallback_pages=[])
    explicit_pages = select_explicit_spare_number_pages(extracted_pages)
    selected_by_page: dict[int, str] = {}
    for page, text in [*table_pages, *explicit_pages]:
        selected_by_page[int(page)] = text
    if selected_by_page:
        return sorted(selected_by_page.items(), key=lambda item: item[0])
    return list(fallback_pages if fallback_pages is not None else extracted_pages)


def _deduplicate_explicit_spare_rows(
    rows: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    """Deduplicate repeated references while preserving genuinely different parents."""
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        identifier_key = normalize_key(row.get("ident_no", ""))
        if identifier_key:
            grouped.setdefault(identifier_key, []).append(dict(row))

    output: list[dict[str, Any]] = []
    removed = 0
    for identifier_key, candidates in grouped.items():
        # A dedicated recommended-spares/procurement list is stronger evidence than
        # a maintenance-schedule reference to the same identifier. When such a
        # high-priority source exists, use it to define the row/parent and treat
        # lower-priority occurrences as corroborating evidence rather than a second
        # import row. This prevents one part from duplicating merely because an
        # earlier schedule page had ambiguous column reading order.
        strongest_priority = max(int(row.get("source_priority", 0)) for row in candidates)
        strongest = [
            row for row in candidates
            if int(row.get("source_priority", 0)) == strongest_priority
        ]
        if strongest_priority >= 3:
            removed += len(candidates) - len(strongest)
            candidates = strongest

        parent_groups: dict[str, list[dict[str, Any]]] = {}
        blank_parent: list[dict[str, Any]] = []
        for row in candidates:
            parent_key = normalize_key(row.get("section_name_english", ""))
            if parent_key:
                parent_groups.setdefault(parent_key, []).append(row)
            else:
                blank_parent.append(row)

        if len(parent_groups) <= 1:
            combined = candidates
            best = max(
                combined,
                key=lambda row: (
                    int(row.get("source_priority", 0)),
                    bool(clean_text(row.get("section_name_english", ""))),
                    bool(clean_text(row.get("description_english", ""))),
                    -int(row.get("source_page") or 10**9),
                ),
            )
            evidence_pages = sorted({int(row.get("source_page")) for row in combined if row.get("source_page")})
            best["evidence_pages"] = evidence_pages
            output.append(best)
            removed += max(0, len(combined) - 1)
            continue

        # The same orderable part may legitimately be used under more than one
        # parent component. Retain one best record per explicit parent. Parent-less
        # duplicates are evidence only and are absorbed into the strongest group.
        for parent_rows in parent_groups.values():
            best = max(
                parent_rows,
                key=lambda row: (
                    int(row.get("source_priority", 0)),
                    bool(clean_text(row.get("description_english", ""))),
                    -int(row.get("source_page") or 10**9),
                ),
            )
            best["evidence_pages"] = sorted({int(row.get("source_page")) for row in parent_rows if row.get("source_page")})
            output.append(best)
            removed += max(0, len(parent_rows) - 1)
        removed += len(blank_parent)

    output.sort(key=lambda row: (int(row.get("source_page") or 10**9), clean_text(row.get("ident_no", ""))))
    return output, removed


def extract_explicit_spare_number_rows(
    extracted_pages: Sequence[tuple[int, str]],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Recover non-tabular rows explicitly labelled as spare-part numbers.

    The parser is manufacturer-neutral and deliberately high precision: a row is
    created only when a source phrase such as ``Spare part number`` or ``Replacement
    part No.`` is paired with a plausible identifier. Nearby short component headings
    are retained as hierarchy context but never converted into fabricated codes.
    """
    rows: list[dict[str, Any]] = []
    current_parent = ""
    previous_page: int | None = None

    for page_number, raw_text in sorted(
        ((int(page), str(text or "")) for page, text in extracted_pages),
        key=lambda item: item[0],
    ):
        if previous_page is not None and page_number - previous_page > 1:
            current_parent = ""
        previous_page = page_number
        lines = [clean_text(line) for line in raw_text.splitlines()]
        item_fragments: list[str] = []
        page_recommended = "recommended spare parts" in clean_text(raw_text).lower()

        for index, line in enumerate(lines):
            if not line:
                continue
            label = _EXPLICIT_SPARE_LABEL_RE.search(line)
            procurement_label = False
            if not label:
                bare_label = _PROCUREMENT_PART_LABEL_RE.search(line)
                procurement_context = clean_text(
                    " ".join(lines[max(0, index - 8):index + 1])
                )
                if bare_label and re.search(
                    r"\b(?:RECOMMEND(?:ATION|ED)?\s+TO\s+ORDER|ORDER\s+\d+|REPLACEMENT|REPLACE|RENEW)\b",
                    procurement_context,
                    flags=re.IGNORECASE,
                ):
                    label = bare_label
                    procurement_label = True
            if label:
                identifier = _explicit_spare_identifier_from_text(line[label.end():])
                identifier_line_index = index
                if not identifier:
                    for offset, next_line in enumerate(lines[index + 1:index + 3], start=1):
                        identifier = _explicit_spare_identifier_from_text(next_line)
                        if identifier:
                            identifier_line_index = index + offset
                            break
                if not identifier:
                    item_fragments = []
                    continue

                row_parent = current_parent

                prefix = _clean_explicit_spare_description(line[:label.start()])
                fragments = list(item_fragments)
                if prefix:
                    fragments.append(prefix)
                description = _explicit_description_from_fragments(fragments)
                if not description:
                    description = "SPARE PART"

                quantity = 1 if page_recommended else None
                if procurement_label:
                    # In maintenance schedules the order recommendation is commonly
                    # printed before ``Part number`` and the actual consumable name on
                    # the following line. Prefer that short post-label description to
                    # fragments of the recommendation sentence.
                    following_description = next(
                        (
                            candidate
                            for candidate in (
                                _clean_explicit_spare_description(value)
                                for value in lines[
                                    identifier_line_index + 1:identifier_line_index + 4
                                ]
                            )
                            if candidate
                            and not _looks_like_explicit_parent_heading(candidate)
                            and not _explicit_spare_identifier_from_text(candidate)
                        ),
                        "",
                    )
                    if following_description:
                        description = following_description
                    action_description = next(
                        (
                            _clean_explicit_spare_description(fragment)
                            for fragment in reversed([*item_fragments[-8:], line[:label.start()]])
                            if re.search(
                                r"\b(?:REPLACEMENT\s+OF|REPLACE|RENEW)\b",
                                clean_text(fragment),
                                flags=re.IGNORECASE,
                            )
                        ),
                        "",
                    )
                    if action_description and not following_description:
                        description = action_description
                    quantity_match = re.search(
                        r"\bORDER\s+(\d+)\s+(?:CANS?|PCS|PIECES?|SETS?)\b",
                        procurement_context,
                        flags=re.IGNORECASE,
                    )
                    if quantity_match:
                        quantity = int(quantity_match.group(1))

                    # A maintenance action can name a consumable while nearby prose
                    # identifies its owning module/unit. Preserve that hierarchy.
                    consumable_match = re.search(
                        r"\b([A-Z][A-Z0-9-]{1,15})\s+(?:LIQUID|FLUID|OIL)\b",
                        description.upper(),
                    )
                    if consumable_match:
                        owner_match = re.search(
                            rf"\b{re.escape(consumable_match.group(1))}\s+(MODULE|UNIT|SYSTEM)\b",
                            raw_text,
                            flags=re.IGNORECASE,
                        )
                        if owner_match:
                            row_parent = (
                                f"{consumable_match.group(1)} {owner_match.group(1)}"
                            )

                confidence = 0.92 if page_recommended else 0.86
                source_priority = 3 if page_recommended else 2
                rows.append(
                    {
                        "section_code": "",
                        "section_name_english": clean_text(row_parent).upper(),
                        "detected_machinery": clean_text(row_parent).upper(),
                        "section_maker": "",
                        "section_model": "",
                        "table_title": (
                            f"Recommended spare parts - {row_parent}"
                            if page_recommended and row_parent
                            else ("Recommended spare parts" if page_recommended else "Explicit spare-number list")
                        ),
                        "section_start_page": page_number,
                        "source_part_no": "",
                        "ident_no": identifier,
                        "part_no": identifier,
                        "code": identifier,
                        "item_no": "",
                        "description_english": description.upper(),
                        "description": description.upper(),
                        "unit": (
                            "SET"
                            if re.search(r"\b(?:SET|KIT)\b", description, flags=re.IGNORECASE)
                            else "PCS"
                        ),
                        "quantity": quantity,
                        "source_page": page_number,
                        "confidence": confidence,
                        "source_pattern": "explicit spare-number list",
                        "source_priority": source_priority,
                        "explicit_spare_number": True,
                    }
                )
                item_fragments = []
                continue

            cleaned_for_parent = _clean_explicit_spare_description(line)
            if _looks_like_explicit_parent_heading(cleaned_for_parent):
                current_parent = cleaned_for_parent
                item_fragments = []
                continue

            upper = line.upper()
            if any(upper.startswith(prefix) for prefix in _EXPLICIT_DESCRIPTION_SKIP_PREFIXES):
                continue
            if _explicit_spare_identifier_from_text(line) and not re.search(r"[A-Za-z]", line):
                continue
            fragment = _clean_explicit_spare_description(line)
            if fragment:
                item_fragments.append(fragment)
                item_fragments = item_fragments[-4:]

    deduplicated, duplicate_count = _deduplicate_explicit_spare_rows(rows)
    messages: list[str] = []
    if deduplicated:
        pages = sorted({int(row["source_page"]) for row in deduplicated})
        messages.append(
            f"Explicit spare-number recovery found {len(deduplicated)} unique row(s) "
            f"on source page(s) {', '.join(map(str, pages))}."
        )
    if duplicate_count:
        messages.append(
            f"Collapsed {duplicate_count} repeated explicit spare-number reference(s) "
            "using identifier and compatible parent context."
        )
    return deduplicated, messages


def extract_explicit_spares_from_pdf(
    pdf_bytes: bytes,
    page_indexes: Sequence[int] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Text-layer safety net for explicit spare-number lists in PDF manuals."""
    if not pdf_bytes:
        return [], []
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return [], []
    indexes = list(page_indexes) if page_indexes is not None else list(range(len(reader.pages)))
    pages: list[tuple[int, str]] = []
    for index in indexes:
        if not 0 <= int(index) < len(reader.pages):
            continue
        page = reader.pages[int(index)]
        try:
            text = page.extract_text() or page.extract_text(extraction_mode="layout") or ""
        except Exception:
            try:
                text = page.extract_text(extraction_mode="layout") or ""
            except Exception:
                text = ""
        pages.append((int(index) + 1, text))
    rows, messages = extract_explicit_spare_number_rows(pages)
    if rows:
        messages.insert(0, "PDF text-layer explicit-spare verification completed.")
    return rows, messages


def apply_authoritative_explicit_spare_rows(
    rows: Sequence[dict[str, Any]],
    authoritative_rows: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int, int]:
    """Replace OCR/AI duplicates with canonical PDF text-layer spare rows.

    Explicitly labelled source rows are stronger than page-level OCR hierarchy.
    This prevents a maintenance table's preceding section from overwriting a row
    whose printed description and surrounding source text prove another owner.
    It also restores a source row if a page-role pass accidentally dropped it.
    """
    canonical_by_identifier: dict[str, dict[str, Any]] = {}
    for raw in authoritative_rows:
        if not isinstance(raw, dict):
            continue
        identifier = normalize_key(
            raw.get("ident_no", raw.get("code", raw.get("part_no", "")))
        )
        if not identifier:
            continue
        existing = canonical_by_identifier.get(identifier)
        if existing is None or int(raw.get("source_priority", 0)) >= int(
            existing.get("source_priority", 0)
        ):
            canonical_by_identifier[identifier] = dict(raw)

    if not canonical_by_identifier:
        return [dict(row) for row in rows if isinstance(row, dict)], 0, 0

    retained: list[dict[str, Any]] = []
    replaced = 0
    present_before = {
        normalize_key(
            row.get("ident_no", row.get("code", row.get("part_no", "")))
        )
        for row in rows
        if isinstance(row, dict)
    }
    for raw in rows:
        if not isinstance(raw, dict):
            continue
        identifier = normalize_key(
            raw.get("ident_no", raw.get("code", raw.get("part_no", "")))
        )
        if identifier in canonical_by_identifier:
            replaced += 1
            continue
        retained.append(dict(raw))

    restored = sum(
        1 for identifier in canonical_by_identifier if identifier not in present_before
    )
    retained.extend(dict(row) for row in canonical_by_identifier.values())
    retained.sort(
        key=lambda row: (
            int(row.get("source_page") or 10**9),
            clean_text(row.get("ident_no", row.get("code", ""))),
        )
    )
    return retained, replaced, restored



_PARTS_LIST_FIGURE_RE = re.compile(
    r"\bPARTS\s+LIST\s+FOR\s+FIGURE\s+([A-Z0-9]+(?:[.-][A-Z0-9]+)*)",
    flags=re.IGNORECASE,
)
_ILLUSTRATED_PARTS_CAPTION_RE = re.compile(
    r"\bFIGURE\s+([A-Z0-9]+(?:[.-][A-Z0-9]+)*)\.?\s+ILLUSTRATED\s+PARTS\s+FOR\s+(?:THE\s+)?(.+?)(?:\s*\(\s*SHEET\s*\d+\s*\))?(?:$|\n)",
    flags=re.IGNORECASE,
)
_MANUAL_REFERENCE_RE = re.compile(
    r"\bMANUAL\s+([A-Z0-9][A-Z0-9._/-]{2,})\b",
    flags=re.IGNORECASE,
)


def _figure_family_key(value: Any) -> str:
    text = clean_text(value).upper().replace(".", "-")
    parts = [part for part in text.split("-") if part]
    if len(parts) >= 2 and parts[-1].isdigit():
        return "-".join(parts[:-1])
    return text


def _illustrated_parts_series_sections(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    """Detect one equipment catalogue split across successive illustrated sheets.

    Some illustrated-parts manuals alternate a parts-list page with an exploded
    drawing page: Figure 6-1, 6-2, 6-3, etc. Those list pages are continuations of
    the same equipment and must not become separate sub-machineries merely because
    an illustration page occurs between them.
    """
    page_texts = [(int(page), str(markdown or "")) for page, markdown in extracted_pages]
    captions: dict[str, list[str]] = {}
    for _, text in page_texts:
        for figure, raw_name in _ILLUSTRATED_PARTS_CAPTION_RE.findall(text):
            family = _figure_family_key(figure)
            raw_name = re.sub(
                r"\s*\(\s*SHEET\s*\d+\s*\)\s*$",
                "",
                clean_text(raw_name),
                flags=re.IGNORECASE,
            )
            name = _clean_machinery_name(raw_name)
            if family and name and not _is_generic_machinery_name(name):
                captions.setdefault(family, []).append(name)

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for page, text in page_texts:
        figure_match = _PARTS_LIST_FIGURE_RE.search(text)
        if not figure_match:
            continue
        family = _figure_family_key(figure_match.group(1))
        manual_match = _MANUAL_REFERENCE_RE.search(text)
        manual_code = clean_text(manual_match.group(1)).upper() if manual_match else ""
        if not family or not manual_code:
            continue

        name_candidates = captions.get(family, [])
        name = ""
        if name_candidates:
            counts: dict[str, int] = {}
            for candidate in name_candidates:
                counts[candidate] = counts.get(candidate, 0) + 1
            name = sorted(
                counts,
                key=lambda candidate: (-counts[candidate], len(candidate), candidate),
            )[0]
        if not name:
            # Fallback to the repeated running header around "Manual <code>".
            for line in text.splitlines()[:20]:
                flat = clean_text(line)
                if not flat or "manual" not in flat.lower():
                    continue
                cleaned = re.sub(
                    rf"\bMANUAL\s+{re.escape(manual_code)}\b",
                    "",
                    flat,
                    flags=re.IGNORECASE,
                ).strip(" -:;|")
                if re.search(r"[A-Za-z]", cleaned) and not _is_generic_machinery_name(cleaned):
                    name = _clean_machinery_name(cleaned)
                    break
        if not name:
            continue

        key = (normalize_key(manual_code), family)
        group = grouped.setdefault(
            key,
            {
                "code": manual_code,
                "aliases": [manual_code],
                "name": name,
                "maker": "",
                "model": "",
                "pages": set(),
                "figure_family": family,
            },
        )
        group["pages"].add(page)
        if len(name) < len(clean_text(group.get("name", ""))) and not _is_generic_machinery_name(name):
            group["name"] = name

    # Require at least two parts-list sheets before applying cross-page series
    # continuation. A single ordinary parts-list page keeps the normal logic.
    return [group for group in grouped.values() if len(group.get("pages", set())) >= 2]



def extract_reference_parts_from_pdf(pdf_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Supplement OCR with exact text-layer rows from Ref. No. parts lists.

    This is a fail-safe for text-based manuals whose replacement-parts chapter is
    laid out as ``Ref. No. / Part Name / Quantity``. It does nothing for scanned
    PDFs without a text layer. The source PDF remains authoritative: full composite
    references populate PART NO/CODE and only the final suffix becomes ITEM NO.
    """
    if not pdf_bytes:
        return [], []
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return [], []

    text_pages: list[tuple[int, str]] = []
    for index, page in enumerate(reader.pages):
        try:
            text = page.extract_text(extraction_mode="layout") or page.extract_text() or ""
        except Exception:
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
        text_pages.append((index + 1, text))

    series_sections = _illustrated_parts_series_sections(text_pages)
    page_to_series: dict[int, dict[str, Any]] = {}
    for series in series_sections:
        for page_number in series.get("pages", set()):
            page_to_series[int(page_number)] = series

    rows: list[dict[str, Any]] = []
    seen: set[tuple[int, str]] = set()
    generic_reference_re = re.compile(
        r"(?<![A-Z0-9])([A-Z0-9][A-Z0-9._/]{2,}-\d+[A-Z]?)(?![A-Z0-9])",
        re.I,
    )
    quantity_re = re.compile(r"\.{2,}\s*(\d+(?:[.,]\d+)?)\b")

    for page_number, text in text_pages:
        lowered = clean_text(text).lower()
        if (
            "parts list for figure" not in lowered
            or "ref. no" not in lowered
            or "part name" not in lowered
        ):
            continue
        series = page_to_series.get(page_number)
        manual_code = clean_text((series or {}).get("code", "")).upper()
        if manual_code:
            reference_re = re.compile(
                rf"(?<![A-Z0-9])({re.escape(manual_code)}-\d+[A-Z]?)(?![A-Z0-9])",
                re.I,
            )
        else:
            reference_re = generic_reference_re
        matches = list(reference_re.finditer(text))
        for match_index, match in enumerate(matches):
            identifier = clean_text(match.group(1)).upper()
            reference_parts = _reference_identifier_parts(identifier)
            if reference_parts is None:
                continue
            _, _, item_no = reference_parts
            segment_end = matches[match_index + 1].start() if match_index + 1 < len(matches) else len(text)
            segment = text[match.end():segment_end]
            quantity_matches = list(quantity_re.finditer(segment))
            if not quantity_matches:
                continue
            quantity_match = quantity_matches[-1]
            description = clean_text(segment[:quantity_match.start()].replace("\n", " "))
            description = re.sub(r"\.{2,}\s*$", "", description).strip(" .;:-")
            if not description or description.upper() in {"NOT USED", "NOT USED."}:
                continue
            # Protect against the footer becoming attached to the last row.
            if len(description) > 500:
                continue
            quantity = quantity_to_number(quantity_match.group(1))
            key = (page_number, normalize_key(identifier))
            if key in seen:
                # A repeated Ref. No. on the same source page is ambiguous. Keep
                # the first exact source occurrence; AI/drawing evidence may add a
                # corrected distinct identifier separately (for example 03036-38).
                continue
            seen.add(key)
            rows.append(
                {
                    "source_page": page_number,
                    "section_start_page": min(series.get("pages", {page_number})) if series else page_number,
                    "section_code": clean_text((series or {}).get("code", "")),
                    "section_name_english": clean_text((series or {}).get("name", "")).upper(),
                    "detected_machinery": clean_text((series or {}).get("name", "")).upper(),
                    "section_maker": "",
                    "section_model": "",
                    "table_title": clean_text((series or {}).get("name", "")).upper(),
                    "ident_no": identifier,
                    "part_no": identifier,
                    "code": identifier,
                    "item_no": item_no,
                    "description_english": description.upper(),
                    "description": description.upper(),
                    "source_description_raw": description,
                    "unit": "PCS",
                    "quantity": quantity,
                    "confidence": 0.99,
                    "language_review": False,
                    "description_source": "printed PDF text layer",
                    "description_source_mismatch": False,
                    "section_review": False,
                    "section_assignment_source": "illustrated parts-list series" if series else "PDF text layer",
                }
            )

    messages: list[str] = []
    if rows:
        messages.append(
            f"PDF text-layer verification recovered {len(rows)} Ref. No. spare-part row(s) "
            "and derived ITEM NO from the printed reference suffix."
        )
        if series_sections:
            messages.append(
                f"Detected {len(series_sections)} illustrated-parts series spanning "
                f"{sum(len(section.get('pages', set())) for section in series_sections)} parts-list page(s)."
            )
    return rows, messages

def build_section_catalog(
    extracted_pages: Sequence[tuple[int, str]],
    extracted_rows: Sequence[dict[str, Any]] | None = None,
    main_row: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build a section catalogue anchored to codes printed in each page header.

    Version 4.7 continues to avoid scanning the spare-parts table body for section
    codes. Body identifiers and cross-references previously produced multiple matches,
    leaving the previous ``active`` section stuck while the spare rows moved on.
    """
    extraction_profile = _document_extraction_profile(extracted_pages)
    allow_simple_section_codes = (
        extraction_profile == MULTILINGUAL_ORDER_CATALOG_PROFILE
    )
    index_sections, index_pages = _index_sections(extracted_pages)
    global_maker, global_model = _global_manual_maker_model(extracted_pages)
    main_row = main_row or {}
    sections: list[dict[str, Any]] = []
    by_alias: dict[str, dict[str, Any]] = {}

    markdown_by_page = {int(page): markdown for page, markdown in extracted_pages}
    rows_by_page: dict[int, list[dict[str, Any]]] = {}
    for extracted_row in extracted_rows or []:
        source_page_value = quantity_to_number(extracted_row.get("source_page"))
        if source_page_value is None:
            continue
        rows_by_page.setdefault(int(source_page_value), []).append(dict(extracted_row))

    engineering_consensus_by_page: dict[int, dict[str, Any]] = {}
    for source_page, source_rows in rows_by_page.items():
        page_markdown = markdown_by_page.get(source_page, "")
        if not page_markdown or not _is_engineering_article_table(page_markdown):
            continue
        consensus = _engineering_article_consensus_metadata(
            source_rows, source_page, page_markdown
        )
        if consensus:
            engineering_consensus_by_page[source_page] = consensus

    def catalog_code_tokens(value: Any) -> list[str]:
        text = clean_text(value).upper().strip().rstrip(".")
        if _is_date_like_section_code(text):
            return []
        if allow_simple_section_codes:
            # This profile prints numeric hierarchy codes. Model names such as
            # SQM10 and ASZ12 are applicability data, never sub-machinery codes.
            return (
                [text]
                if re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", text)
                else []
            )
        return _section_code_tokens(value, permissive=True)

    def register(section: dict[str, Any], priority: int = 50) -> dict[str, Any]:
        aliases = [
            clean_text(value).upper()
            for value in section.get("aliases", [])
            if clean_text(value) and not _is_date_like_section_code(value)
        ]
        raw_code = clean_text(section.get("code", "")).upper()
        code = ("" if _is_date_like_section_code(raw_code) else raw_code) or _join_section_codes(aliases)
        if not aliases and code:
            aliases = catalog_code_tokens(code) or [code]
        aliases = list(dict.fromkeys(aliases))
        name = _clean_catalog_section_name(section.get("name", ""), code)
        existing = next(
            (
                by_alias.get(normalize_key(alias))
                for alias in aliases
                if by_alias.get(normalize_key(alias)) is not None
            ),
            None,
        )
        if existing is None:
            existing = {
                "code": code,
                "aliases": aliases or ([code] if code else []),
                "name": name,
                "maker": _clean_catalog_maker(section.get("maker", ""))
                if allow_simple_section_codes
                else clean_text(section.get("maker", "")).upper(),
                "model": clean_text(section.get("model", "")).upper(),
                "pages": set(section.get("pages", set())),
                "_name_priority": int(priority if name else -1),
                "_maker_priority": int(
                    priority
                    if (
                        _clean_catalog_maker(section.get("maker", ""))
                        if allow_simple_section_codes
                        else clean_text(section.get("maker", ""))
                    )
                    else -1
                ),
                "_model_priority": int(priority if clean_text(section.get("model", "")) else -1),
            }
            sections.append(existing)
        else:
            if name and int(priority) > int(existing.get("_name_priority", -1)):
                existing["name"] = name
                existing["_name_priority"] = int(priority)
            maker = (
                _clean_catalog_maker(section.get("maker", ""))
                if allow_simple_section_codes
                else clean_text(section.get("maker", "")).upper()
            )
            if maker and int(priority) > int(existing.get("_maker_priority", -1)):
                existing["maker"] = maker
                existing["_maker_priority"] = int(priority)
            model = clean_text(section.get("model", "")).upper()
            if model and int(priority) > int(existing.get("_model_priority", -1)):
                existing["model"] = model
                existing["_model_priority"] = int(priority)
            existing["pages"].update(section.get("pages", set()))
            for alias in aliases:
                if alias not in existing["aliases"]:
                    existing["aliases"].append(alias)
            if not existing.get("code") and code:
                existing["code"] = code
        for alias in existing.get("aliases", []):
            by_alias[normalize_key(alias)] = existing
        return existing

    # Printed index/catalogue entries are reliable, but an exact page header is
    # stronger and is therefore registered with a higher priority below.
    for section in index_sections:
        register(section, priority=80)

    for drawing_section in _drawing_heading_sections(extracted_pages):
        register(drawing_section, priority=120)

    illustrated_parts_page_map: dict[int, dict[str, Any]] = {}
    for series in _illustrated_parts_series_sections(extracted_pages):
        registered_series = register(series, priority=125)
        registered_series["_illustrated_series_pages"] = {
            int(series_page) for series_page in series.get("pages", set())
        }
        registered_series["_illustrated_parts_series"] = True
        for series_page in series.get("pages", set()):
            illustrated_parts_page_map[int(series_page)] = registered_series

    page_metadata: dict[int, dict[str, Any]] = {}
    page_heading_sections: dict[int, list[dict[str, Any]]] = {}
    for page_number, markdown in extracted_pages:
        page = int(page_number)
        metadata = _page_metadata(page, markdown)
        consensus = engineering_consensus_by_page.get(page)
        if consensus:
            metadata["section_code"] = consensus.get("section_code", "")
            metadata["section_code_source"] = consensus.get(
                "section_code_source", "engineering article-table consensus"
            )
            metadata["section_code_evidence"] = consensus.get("section_code_evidence", "")
            if clean_text(consensus.get("section_name_english", "")):
                metadata["section_name_raw"] = consensus.get(
                    "section_name_raw", consensus["section_name_english"]
                )
                metadata["section_name_english"] = consensus["section_name_english"]
        page_metadata[page] = metadata
        if page in index_pages:
            continue
        if allow_simple_section_codes:
            registered_headings: list[dict[str, Any]] = []
            for heading in _classic_catalog_sections(markdown):
                registered_headings.append(
                    register(
                        {
                            "code": heading.get("section_code", ""),
                            "aliases": [heading.get("section_code", "")],
                            "name": heading.get("section_name_english", ""),
                            "maker": metadata.get("maker", ""),
                            "model": metadata.get("model", ""),
                            "pages": {page},
                        },
                        priority=110,
                    )
                )
            if registered_headings:
                page_heading_sections[page] = registered_headings
                continue
        if clean_text(metadata.get("section_code_source", "")) in {"engineering drawing title block", "engineering article-table consensus"}:
            printed_code = clean_text(metadata.get("section_code", "")).upper()
            code_tokens = [printed_code] if printed_code else []
        else:
            code_tokens = catalog_code_tokens(metadata.get("section_code", ""))
        if code_tokens:
            is_title_block = (
                clean_text(metadata.get("section_code_source", ""))
                in {"engineering drawing title block", "engineering article-table consensus"}
            )
            registered = register(
                {
                    "code": _join_section_codes(code_tokens),
                    "aliases": code_tokens,
                    "name": metadata.get("section_name_english", ""),
                    "maker": metadata.get("maker", ""),
                    "model": metadata.get("model", ""),
                    "pages": {page},
                },
                priority=145 if is_title_block else 100,
            )
            registered["pages"].add(page)
            if is_title_block:
                registered.setdefault("_authoritative_start_pages", set()).add(page)

    # AI catalogue values can fill gaps but must never overwrite an exact printed
    # header title such as RESILIENT MOUNTING with a paraphrase such as ELASTIC MOUNTING.
    #
    # In a multilingual Order-No. catalogue, an item's drawing position (for
    # example 7.15 or 15.3) looks exactly like a legitimate hierarchy code.  The
    # AI therefore must not be allowed to create a section from an unconfirmed
    # numeric value: that turns spare parts into false sub-machineries.  It may
    # enrich only a section that the source itself has already confirmed.
    for row in extracted_rows or []:
        source_page_value = quantity_to_number(row.get("source_page"))
        source_page = int(source_page_value) if source_page_value is not None else None
        source_metadata = page_metadata.get(source_page, {}) if source_page is not None else {}
        if clean_text(source_metadata.get("section_code_source", "")) in {"engineering drawing title block", "engineering article-table consensus"}:
            printed_code = clean_text(source_metadata.get("section_code", "")).upper()
            code_tokens = [printed_code] if printed_code else []
        else:
            code_tokens = catalog_code_tokens(row.get("section_code", ""))
        if not code_tokens:
            continue
        if allow_simple_section_codes:
            confirmed: list[dict[str, Any]] = []
            for token in code_tokens:
                section = by_alias.get(normalize_key(token))
                if section is not None and all(
                    id(section) != id(existing) for existing in confirmed
                ):
                    confirmed.append(section)
            if len(confirmed) != 1:
                continue
        source_page = quantity_to_number(row.get("source_page"))
        register(
            {
                "code": _join_section_codes(code_tokens),
                "aliases": code_tokens,
                "name": clean_text(
                    row.get("section_name_english", row.get("detected_machinery", ""))
                ).upper(),
                "maker": clean_text(row.get("section_maker", "")).upper(),
                "model": clean_text(row.get("section_model", "")).upper(),
                # AI rows may contain stale section context. They can define a
                # missing catalogue alias, but never claim authoritative pages.
                "pages": set(),
            },
            priority=30,
        )

    def unique_sections(values: Sequence[dict[str, Any] | None]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen_ids: set[int] = set()
        for value in values:
            if value is not None and id(value) not in seen_ids:
                result.append(value)
                seen_ids.add(id(value))
        return result

    def sections_for_codes(value: Any) -> list[dict[str, Any]]:
        direct = by_alias.get(normalize_key(value))
        if direct is not None:
            return [direct]
        return unique_sections(
            [
                by_alias.get(normalize_key(token))
                for token in catalog_code_tokens(value)
            ]
        )

    def normalized_code_text(value: Any) -> str:
        text_value = clean_text(value).upper()
        return re.sub(r"\s*([./_-])\s*", r"\1", text_value)

    def alias_in_header(alias: str, header_text: str) -> bool:
        alias_text = normalized_code_text(alias)
        source_text = normalized_code_text(header_text)
        if not alias_text or not source_text:
            return False
        return bool(
            re.search(
                rf"(?<![A-Z0-9]){re.escape(alias_text)}(?![A-Z0-9])",
                source_text,
            )
        )

    parts_pages = {
        int(page)
        for page, markdown in extracted_pages
        if int(page) not in index_pages
        and (
            _looks_like_spare_table_page(markdown)
            or _looks_like_explicit_spare_number_page(markdown)
        )
    }
    # Every source page explicitly titled "Parts List for Figure ..." in a
    # confirmed illustrated-parts series is a parts page even when OCR splits
    # the header text (for example "Part Na me").
    parts_pages.update(
        page for page in illustrated_parts_page_map if page not in index_pages
    )
    page_map: dict[int, dict[str, Any]] = {}
    page_map_sources: dict[int, str] = {}
    ambiguous_pages: set[int] = set()
    unmapped_parts_pages: set[int] = set()
    active: dict[str, Any] | None = None
    active_anchor_page: int | None = None

    for page in sorted(page_metadata):
        if page in index_pages:
            continue
        metadata = page_metadata[page]
        explicit_matches = page_heading_sections.get(page) or sections_for_codes(
            metadata.get("section_code", "")
        )
        multiple_numbered_headings = bool(
            allow_simple_section_codes and len(explicit_matches) > 1
        )
        resolved: dict[str, Any] | None = None
        resolved_source = ""

        series_section = illustrated_parts_page_map.get(page)
        if series_section is not None:
            resolved = series_section
            resolved_source = "illustrated parts-list series"
        elif len(explicit_matches) == 1:
            resolved = explicit_matches[0]
            resolved_source = "exact page-header code"
        elif len(explicit_matches) > 1:
            # Historical catalogues can legitimately start two numbered sections
            # on the same PDF page. Row-level parsing assigns that page's records;
            # retain the last heading only as context for the following page.
            ambiguous_pages.add(page)
            if allow_simple_section_codes:
                active = explicit_matches[-1]
                active_anchor_page = page
            else:
                active = None
                active_anchor_page = None
        else:
            header_text = metadata.get("header_text", "")
            header_matches = unique_sections(
                [
                    section
                    for section in sections
                    if (
                        not section.get("_illustrated_series_pages")
                        or page in section.get("_illustrated_series_pages", set())
                    )
                    and any(
                        alias_in_header(alias, header_text)
                        for alias in section.get("aliases", [])
                    )
                ]
            )
            if len(header_matches) == 1:
                resolved = header_matches[0]
                resolved_source = "exact code in page header area"
            elif len(header_matches) > 1:
                ambiguous_pages.add(page)
                active = None
                active_anchor_page = None
            else:
                # Use a very strong title match only when no code is available.
                name_key = normalize_key(metadata.get("section_name_english", ""))
                if name_key:
                    scored = sorted(
                        (
                            (
                                SequenceMatcher(
                                    None,
                                    name_key,
                                    normalize_key(section.get("name", "")),
                                ).ratio(),
                                section,
                            )
                            for section in sections
                            if clean_text(section.get("name", ""))
                        ),
                        key=lambda pair: pair[0],
                        reverse=True,
                    )
                    if scored and scored[0][0] >= 0.96:
                        second_score = scored[1][0] if len(scored) > 1 else 0.0
                        if scored[0][0] - second_score >= 0.03:
                            resolved = scored[0][1]
                            resolved_source = "exact page-header title"

        if resolved is not None:
            active = resolved
            active_anchor_page = page
            resolved["pages"].add(page)
        elif (
            page in parts_pages
            and active is not None
            and active_anchor_page is not None
            and page - active_anchor_page == 1
            and page not in ambiguous_pages
            and not clean_text(metadata.get("section_code", ""))
        ):
            # Carry forward only to an immediately consecutive continuation page
            # that has no new header code. This cannot jump across ambiguous pages.
            resolved = active
            resolved_source = "confirmed consecutive continuation"
            active_anchor_page = page
            resolved["pages"].add(page)
        elif page in parts_pages and not multiple_numbered_headings:
            unmapped_parts_pages.add(page)
            active = None
            active_anchor_page = None

        if page in parts_pages and resolved is not None:
            page_map[page] = resolved
            page_map_sources[page] = resolved_source

    main_maker = _credible_section_maker(main_row.get("MAKER", ""))
    main_model = _credible_section_model(main_row.get("MODEL", ""))
    fallback_maker = main_maker or _credible_section_maker(global_maker)
    fallback_model = main_model or _credible_section_model(global_model)
    final_sections: list[dict[str, Any]] = []
    seen_section_ids: set[int] = set()
    for section in sections:
        if id(section) in seen_section_ids:
            continue
        seen_section_ids.add(id(section))
        pages = sorted(
            int(page)
            for page in section.get("pages", set())
            if int(page) not in index_pages
        )
        section["name"] = _clean_catalog_section_name(
            section.get("name", ""), section.get("code", "")
        )
        if section.get("_illustrated_parts_series"):
            section["maker"] = (
                _credible_section_maker(section.get("maker", "")) or fallback_maker
            )
            section["model"] = (
                _credible_section_model(section.get("model", "")) or fallback_model
            )
        else:
            section["maker"] = _credible_section_maker(section.get("maker", "")) or fallback_maker
            section["model"] = _credible_section_model(section.get("model", "")) or fallback_model
        authoritative_start_pages = sorted(
            int(page)
            for page in section.get("_authoritative_start_pages", set())
            if page is not None
        )
        section["first_page"] = (
            min(authoritative_start_pages)
            if authoritative_start_pages
            else (min(pages) if pages else None)
        )
        section["last_page"] = max(pages) if pages else section["first_page"]
        section["pages"] = pages
        section.pop("_name_priority", None)
        section.pop("_maker_priority", None)
        section.pop("_model_priority", None)
        section.pop("_illustrated_series_pages", None)
        section.pop("_illustrated_parts_series", None)
        section.pop("_authoritative_start_pages", None)
        if section.get("code") or section.get("name"):
            final_sections.append(section)
    def section_sort_code(value: Any) -> tuple[Any, ...]:
        code = clean_text(value).strip().rstrip(".")
        if re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", code):
            return (0, *(int(part) for part in code.split(".")))
        return (1, normalize_key(code))

    final_sections.sort(
        key=lambda item: (
            item.get("first_page") is None,
            item.get("first_page") or 10**9,
            section_sort_code(item.get("code", "")),
            item.get("name", ""),
        )
    )
    return {
        "sections": final_sections,
        "page_map": page_map,
        "page_map_sources": page_map_sources,
        "page_metadata": page_metadata,
        "index_pages": index_pages,
        "parts_pages": parts_pages,
        "ambiguous_pages": ambiguous_pages,
        "unmapped_parts_pages": unmapped_parts_pages,
        "illustrated_parts_page_map": illustrated_parts_page_map,
    }

def _catalog_prompt_hint(sections: Sequence[dict[str, Any]]) -> str:
    lines: list[str] = []
    for section in sections[:120]:
        code = clean_text(section.get("code", ""))
        name = clean_text(section.get("name", ""))
        maker = clean_text(section.get("maker", ""))
        model = clean_text(section.get("model", ""))
        pages = ",".join(str(page) for page in section.get("pages", []) if page)
        if code or name:
            lines.append(
                f"- CODE={code}; NAME={name}; PAGES={pages}; "
                f"MAKER={maker}; MODEL={model}"
            )
    return "\n".join(lines)


def _normalized_ai_row(row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    ident = clean_text(item.get("ident_no", item.get("code", item.get("part_no", ""))))
    section_code = clean_text(item.get("section_code", ""))
    if ident and section_code and normalize_key(ident) == normalize_key(section_code):
        ident = ""
    item_no = clean_text(item.get("item_no", ""))
    reference_parts = _reference_identifier_parts(ident)
    if reference_parts is not None and section_code:
        _, reference_prefix, reference_item = reference_parts
        if normalize_key(reference_prefix) == normalize_key(section_code) and (
            not item_no or normalize_key(item_no) == normalize_key(ident)
        ):
            item_no = reference_item
    explicit_source = normalize_key(item.get("source_pattern", "")) == normalize_key("explicit spare-number list")
    if explicit_source:
        description = clean_text(item.get("description_english", item.get("description", ""))).upper()
        language_review = False
    else:
        description, language_review = _best_effort_english_description(
            item.get("description_english", item.get("description", ""))
        )
    return {
        **item,
        "ident_no": ident,
        "item_no": item_no,
        "description_english": description,
        "language_review": bool(language_review),
        "section_code": section_code,
        "section_name_english": clean_text(item.get("section_name_english", item.get("detected_machinery", ""))).upper(),
        "section_maker": clean_text(item.get("section_maker", "")).upper(),
        "section_model": clean_text(item.get("section_model", "")).upper(),
        "source_page": int(quantity_to_number(item.get("source_page"))) if quantity_to_number(item.get("source_page")) is not None else None,
        "section_start_page": int(quantity_to_number(item.get("section_start_page"))) if quantity_to_number(item.get("section_start_page")) is not None else None,
        "quantity": quantity_to_number(item.get("quantity", item.get("QNT"))),
        "confidence": clamp_confidence(item.get("confidence", 0.70)),
    }


def _carry_forward_merged_item_numbers(
    rows: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    """Repeat a merged drawing position on immediate continuation records.

    Some manuals print one ITEM NO in a vertically merged cell while listing two
    or more identifiers/descriptions beside it. OCR/Markdown may therefore leave
    the continuation record's item number blank. Carry forward only within the
    same source page, section code, and table title, and only for a genuine spare
    row with its own identifier and description.
    """
    result: list[dict[str, Any]] = []
    last_key: tuple[str, str, str] | None = None
    last_item_no = ""
    inherited_count = 0

    for raw in rows:
        row = dict(raw)
        key = (
            str(row.get("source_page", "")),
            normalize_key(row.get("section_code", "")),
            normalize_key(row.get("table_title", "")),
        )
        item_no = clean_text(row.get("item_no", ""))
        identifier = clean_text(
            row.get("ident_no", row.get("code", row.get("part_no", "")))
        )
        description = clean_text(
            row.get("description_english", row.get("description", ""))
        )

        if item_no:
            last_key = key
            last_item_no = item_no
        elif (
            last_key == key
            and last_item_no
            and identifier
            and description
        ):
            row["item_no"] = last_item_no
            row["item_no_inherited"] = True
            inherited_count += 1
        else:
            # Do not carry a position across another page, section, or table.
            last_key = key
            last_item_no = ""

        result.append(row)

    return result, inherited_count



def _engineering_article_parent_stem(value: Any) -> str:
    """Return the parent/drawing stem from a structured Article No. value.

    This deliberately requires a clear parent + variant suffix pattern. It does
    not split ordinary one-piece part numbers and therefore cannot create a
    hierarchy merely because several unrelated identifiers share leading digits.
    """
    text = clean_text(value).upper().strip(" .;,:|-")
    match = re.fullmatch(
        r"([A-Z0-9]{5,12})[\s._/-]+([A-Z0-9]{1,4})",
        text,
        flags=re.I,
    )
    if not match:
        return ""
    parent = _valid_automatic_section_code(match.group(1))
    suffix = clean_text(match.group(2))
    if not parent or not suffix or normalize_key(parent) == normalize_key(suffix):
        return ""
    return parent


def _chapter_like_section_code(value: Any) -> bool:
    """Return True for navigation/section numbering rather than drawing identity."""
    text = clean_text(value).strip().rstrip(".")
    return bool(re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){1,5}", text))


def _reconcile_engineering_article_parent_identity(
    rows: Sequence[dict[str, Any]],
    extracted_pages: Sequence[tuple[int, str]],
) -> tuple[list[dict[str, Any]], int, list[str]]:
    """Correct parent hierarchy from cumulative, source-backed article evidence.

    OCR reading order can expose a chapter heading (for example 4.3.21) while the
    actual labelled drawing number/title remain difficult to pair. Once the page
    has yielded several genuine Article-No. rows, their common structured stem and
    repeated short Name/Designation provide independent source evidence for the
    parent drawing. This reconciliation is generic and runs only on confirmed
    engineering Article-No. tables.
    """
    result = [dict(row) for row in rows if isinstance(row, dict)]
    if not result:
        return result, 0, []

    page_lookup = {int(page): str(markdown or "") for page, markdown in extracted_pages}
    row_indexes_by_page: dict[int, list[int]] = {}
    for index, row in enumerate(result):
        page_value = quantity_to_number(row.get("source_page"))
        if page_value is None:
            continue
        row_indexes_by_page.setdefault(int(page_value), []).append(index)

    changed_rows = 0
    messages: list[str] = []
    for page_number, indexes in sorted(row_indexes_by_page.items()):
        markdown = page_lookup.get(page_number, "")
        if not markdown or not _is_engineering_article_table(markdown):
            continue
        if len(indexes) < 3:
            continue

        title_block = _engineering_title_block_metadata(markdown)
        explicit_code = _valid_automatic_section_code(
            title_block.get("section_code", "")
        )
        explicit_name = clean_text(
            title_block.get("section_name_english", "")
        ).upper()

        identifiers = [
            clean_text(result[index].get("ident_no", result[index].get("code", "")))
            for index in indexes
        ]
        stems = [
            stem
            for stem in (_engineering_article_parent_stem(value) for value in identifiers)
            if stem
        ]
        stem_counts = {stem: stems.count(stem) for stem in set(stems)}
        repeated_stem = ""
        repeated_count = 0
        if stem_counts:
            repeated_stem, repeated_count = max(
                stem_counts.items(), key=lambda item: item[1]
            )
        strong_stem = bool(
            repeated_stem
            and repeated_count >= 3
            and repeated_count / max(1, len(indexes)) >= 0.75
        )

        descriptions = [
            clean_text(
                result[index].get(
                    "description_english", result[index].get("description", "")
                )
            ).upper()
            for index in indexes
        ]
        descriptions = [value for value in descriptions if value]
        description_counts = {
            value: descriptions.count(value) for value in set(descriptions)
        }
        dominant_name = ""
        dominant_count = 0
        if description_counts:
            dominant_name, dominant_count = max(
                description_counts.items(), key=lambda item: item[1]
            )
        strong_name = bool(
            dominant_name
            and dominant_count >= 3
            and dominant_count / max(1, len(indexes)) >= 0.75
            and len(dominant_name) <= 80
            and not _is_generic_machinery_name(dominant_name)
        )

        resolved_code = explicit_code or (repeated_stem if strong_stem else "")
        resolved_name = explicit_name or (dominant_name if strong_name else "")
        if explicit_name and strong_name and normalize_key(explicit_name) != normalize_key(dominant_name):
            evidence = clean_text(title_block.get("section_name_evidence", "")).lower()
            weak_pairing = (
                not evidence
                or "next-line" in evidence
                or "offset" in evidence
            )
            dominant_pattern = re.compile(
                rf"(?<![A-Z0-9]){re.escape(dominant_name)}(?![A-Z0-9])",
                flags=re.IGNORECASE,
            )
            broad_navigation = bool(
                dominant_pattern.search(explicit_name)
                and len(explicit_name) >= len(dominant_name) + 8
                and (
                    "/" in explicit_name
                    or " AND " in f" {explicit_name} "
                    or "DRAWING" in explicit_name
                    or "TECHNICAL DATA" in explicit_name
                )
            )
            if weak_pairing and broad_navigation:
                resolved_name = dominant_name
        if not resolved_code:
            continue

        page_changes = 0
        for index in indexes:
            row = result[index]
            current_code = clean_text(row.get("section_code", "")).upper()
            current_name = clean_text(
                row.get("section_name_english", row.get("detected_machinery", ""))
            ).upper()

            replace_code = bool(
                explicit_code
                or not current_code
                or _is_date_like_section_code(current_code)
                or _chapter_like_section_code(current_code)
                or normalize_key(current_code) == normalize_key(repeated_stem)
            )
            if replace_code and normalize_key(current_code) != normalize_key(resolved_code):
                row["section_code"] = resolved_code
                page_changes += 1

            navigation_like_name = bool(
                "/" in current_name
                or "DRAWING" in current_name
                or "DIMENSION" in current_name
                or "TECHNICAL DATA" in current_name
            )
            replace_name = bool(
                resolved_name
                and (
                    explicit_name
                    or not current_name
                    or navigation_like_name
                    or _is_generic_machinery_name(current_name)
                )
            )
            if replace_name and normalize_key(current_name) != normalize_key(resolved_name):
                row["section_name_english"] = resolved_name
                row["detected_machinery"] = resolved_name
                row["table_title"] = resolved_name
                page_changes += 1

            row["section_start_page"] = int(page_number)
            row["section_assignment_source"] = "engineering title-block/article evidence"
            row["section_review"] = False
            if resolved_name:
                row["confidence"] = max(
                    clamp_confidence(row.get("confidence", 0.0), fallback=0.0),
                    0.90,
                )

        if page_changes:
            changed_rows += len(indexes)
            source_description = (
                "labelled title block"
                if explicit_code
                else "repeated Article-No. parent stem"
            )
            name_description = (
                "labelled title"
                if explicit_name
                else ("repeated Name/Designation" if resolved_name else "existing title")
            )
            messages.append(
                f"PDF page {page_number}: reconciled {len(indexes)} engineering article "
                f"row(s) to parent {resolved_code} / {resolved_name or 'existing name'} "
                f"using {source_description} and {name_description}."
            )

    return result, changed_rows, messages

def prepare_benefit_rows(
    ai_rows: Sequence[dict[str, Any]],
    extracted_pages: Sequence[tuple[int, str]],
    source_document_name: str,
    main_row: dict[str, Any],
    default_unit: str = "PCS",
    catalog_pages: Sequence[tuple[int, str]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Repair and normalize extraction using exact PDF table structure and page headers."""
    direct_rows = _direct_table_rows(extracted_pages)
    section_context_pages = list(catalog_pages) if catalog_pages is not None else list(extracted_pages)
    extraction_profile = _document_extraction_profile(section_context_pages)
    allow_simple_section_codes = (
        extraction_profile == MULTILINGUAL_ORDER_CATALOG_PROFILE
    )

    def safe_section_code(value: Any) -> str:
        text_value = clean_text(value).upper().strip().rstrip(".")
        if allow_simple_section_codes and re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", text_value):
            return text_value
        return _valid_automatic_section_code(text_value)

    page_markdown_lookup = {int(page): markdown for page, markdown in extracted_pages}
    normalized_ai: list[dict[str, Any]] = []
    rejected_ai_identifiers = 0
    expanded_ai_identifiers = 0
    rejected_non_article_ai_rows = 0
    for raw in ai_rows:
        if not isinstance(raw, dict):
            continue
        normalized = _normalized_ai_row(raw)
        source_page = normalized.get("source_page")
        page_markdown = page_markdown_lookup.get(int(source_page)) if source_page is not None else ""
        if page_markdown and _is_engineering_article_table(page_markdown):
            title_block = _engineering_title_block_metadata(page_markdown)
            if title_block:
                normalized["section_code"] = safe_section_code(title_block.get("section_code", ""))
                if clean_text(title_block.get("section_name_english", "")):
                    normalized["section_name_english"] = title_block["section_name_english"]
                    normalized["detected_machinery"] = title_block["section_name_english"]
            # This page pattern has no drawing-position or quantity field in the
            # Article No. table. Art. Rev. and Note lengths must never leak into them.
            normalized["item_no"] = ""
            normalized["quantity"] = None
            if not clean_text(normalized.get("ident_no", "")):
                rejected_non_article_ai_rows += 1
                continue
        if not allow_simple_section_codes:
            normalized_ai.append(normalized)
            continue
        raw_identifier = clean_text(normalized.get("ident_no", ""))
        identifiers = _catalog_order_number_lines(raw_identifier)
        if not identifiers:
            if raw_identifier:
                rejected_ai_identifiers += 1
            # Retain the row only as section/description context. It cannot become
            # an included spare without a source-valid Order-No.
            normalized["ident_no"] = ""
            normalized["identifier_rejected"] = bool(raw_identifier)
            normalized_ai.append(normalized)
            continue
        if len(identifiers) > 1:
            expanded_ai_identifiers += len(identifiers) - 1
        for identifier in identifiers:
            variant = dict(normalized)
            variant["ident_no"] = identifier
            variant["code"] = identifier
            variant["part_no"] = identifier
            variant["identifier_rejected"] = False
            normalized_ai.append(variant)

    catalog = build_section_catalog(section_context_pages, [*normalized_ai, *direct_rows], main_row)
    sections = catalog["sections"]
    page_map = catalog["page_map"]
    page_map_sources = catalog.get("page_map_sources", {})
    index_pages = catalog["index_pages"]
    parts_pages = catalog["parts_pages"]
    ambiguous_pages = catalog.get("ambiguous_pages", set())
    unmapped_parts_pages = catalog.get("unmapped_parts_pages", set())

    by_alias: dict[str, dict[str, Any]] = {}
    for section in sections:
        for alias in section.get("aliases", []):
            if normalize_key(alias):
                by_alias[normalize_key(alias)] = section

    def benefit_code_tokens(value: Any) -> list[str]:
        text = clean_text(value).upper().strip().rstrip(".")
        if normalize_key(text) in by_alias:
            return [text]
        if allow_simple_section_codes:
            return (
                [text]
                if re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", text)
                else []
            )
        return _section_code_tokens(value, permissive=True)

    def exact_sections_for_code(value: Any) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        seen_ids: set[int] = set()
        for token in benefit_code_tokens(value):
            section = by_alias.get(normalize_key(token))
            if section is not None and id(section) not in seen_ids:
                matches.append(section)
                seen_ids.add(id(section))
        return matches

    def fuzzy_section_for_name(value: Any) -> dict[str, Any] | None:
        name_key = normalize_key(value)
        if not name_key:
            return None
        scored = sorted(
            (
                (
                    SequenceMatcher(
                        None,
                        name_key,
                        normalize_key(section.get("name", "")),
                    ).ratio(),
                    section,
                )
                for section in sections
                if clean_text(section.get("name", ""))
            ),
            key=lambda pair: pair[0],
            reverse=True,
        )
        if not scored or scored[0][0] < 0.94:
            return None
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        return scored[0][1] if scored[0][0] - second_score >= 0.03 else None

    def section_for(
        page: int | None,
        direct_row: dict[str, Any] | None = None,
        ai_row: dict[str, Any] | None = None,
    ) -> tuple[dict[str, Any] | None, str, bool]:
        """Resolve a section with printed page context ahead of AI context."""
        direct_row = direct_row or {}
        ai_row = ai_row or {}
        direct_matches = exact_sections_for_code(direct_row.get("section_code", ""))
        page_section = page_map.get(page) if page is not None else None
        ai_matches = exact_sections_for_code(ai_row.get("section_code", ""))
        conflict = False
        explicit_spare_row = (
            normalize_key(ai_row.get("source_pattern", ""))
            == normalize_key("explicit spare-number list")
        )

        # Maintenance/procurement prose often contains component item tags such as
        # QT201-50, PT201-16, etc. They are not automatically the parent machinery
        # code. For an explicitly labelled spare-number row, prefer a strong parent
        # NAME match from the document catalogue and otherwise keep the row on its
        # source parent text/main machinery rather than borrowing a page-map code.
        if explicit_spare_row and not direct_matches and not ai_matches:
            explicit_name_match = fuzzy_section_for_name(
                ai_row.get("section_name_english", "")
            )
            if explicit_name_match is not None:
                return explicit_name_match, "explicit spare parent-title match", False
            return None, "explicit spare parent unconfirmed", True

        if len(direct_matches) == 1:
            chosen = direct_matches[0]
            if page_section is not None and id(page_section) != id(chosen):
                conflict = True
            if len(ai_matches) == 1 and id(ai_matches[0]) != id(chosen):
                conflict = True
            return chosen, "exact printed page-header code", conflict
        if len(direct_matches) > 1:
            conflict = True

        if page_section is not None:
            if len(ai_matches) == 1 and id(ai_matches[0]) != id(page_section):
                conflict = True
            return page_section, page_map_sources.get(page, "page-header map"), conflict

        # AI section codes are only a fallback when no printed page/header mapping
        # exists. They can no longer override a deterministic source code.
        if len(ai_matches) == 1:
            return ai_matches[0], "AI section code fallback", conflict
        if len(ai_matches) > 1:
            conflict = True

        direct_name_match = fuzzy_section_for_name(direct_row.get("section_name_english", ""))
        if direct_name_match is not None:
            return direct_name_match, "printed page-title match", conflict
        ai_name_match = fuzzy_section_for_name(ai_row.get("section_name_english", ""))
        if ai_name_match is not None:
            return ai_name_match, "AI title fallback", conflict
        return None, "unconfirmed", True

    # Match AI rows with the strongest available row identity. ITEM NO alone is not
    # globally unique and could previously select a different row on the same page.
    ai_by_full_key: dict[tuple[int | None, str, str], dict[str, Any]] = {}
    ai_by_page_ident_lists: dict[tuple[int | None, str], list[dict[str, Any]]] = {}
    ai_by_page_item_lists: dict[tuple[int | None, str], list[dict[str, Any]]] = {}
    for row in normalized_ai:
        page = row.get("source_page")
        item_key = normalize_key(row.get("item_no", ""))
        ident_key = normalize_key(row.get("ident_no", ""))
        if item_key or ident_key:
            ai_by_full_key[(page, item_key, ident_key)] = row
        if ident_key:
            ai_by_page_ident_lists.setdefault((page, ident_key), []).append(row)
        if item_key:
            ai_by_page_item_lists.setdefault((page, item_key), []).append(row)

    def matched_ai_row(page: int, item_key: str, ident_key: str) -> dict[str, Any] | None:
        exact = ai_by_full_key.get((page, item_key, ident_key))
        if exact is not None:
            return exact
        by_ident = ai_by_page_ident_lists.get((page, ident_key), []) if ident_key else []
        if len(by_ident) == 1:
            return by_ident[0]
        by_item = ai_by_page_item_lists.get((page, item_key), []) if item_key else []
        if len(by_item) == 1:
            return by_item[0]
        return None

    output: list[dict[str, Any]] = []
    matched_ai_ids: set[int] = set()
    language_exceptions = 0
    paraphrases_prevented = 0
    section_conflicts = 0
    unconfirmed_sections = 0

    if direct_rows:
        for direct in direct_rows:
            page = int(direct["source_page"])
            item_key = normalize_key(direct.get("item_no", ""))
            ident_key = normalize_key(direct.get("ident_no", ""))
            ai = matched_ai_row(page, item_key, ident_key)
            if ai is not None:
                matched_ai_ids.add(id(ai))

            ai_description_source = ""
            if ai is not None and not bool(ai.get("local_fallback", False)):
                ai_description_source = ai.get("description_english", "")
            description, language_review, ai_disagreed, description_source = (
                _select_source_faithful_description(
                    direct.get("description_raw", ""),
                    ai_description_source,
                )
            )
            if language_review:
                language_exceptions += 1
            if ai_disagreed:
                paraphrases_prevented += 1

            section, section_source, section_conflict = section_for(page, direct, ai)
            if section_conflict:
                section_conflicts += 1
            if section is None:
                unconfirmed_sections += 1

            # Printed page data wins over AI fallbacks for section identity and title.
            section_code = next(
                (
                    candidate
                    for candidate in (
                        safe_section_code((section or {}).get("code", "")),
                        safe_section_code(direct.get("section_code", "")),
                        safe_section_code((ai or {}).get("section_code", "")),
                    )
                    if candidate
                ),
                "",
            )
            section_name = (
                clean_text((section or {}).get("name", ""))
                or clean_text(direct.get("section_name_english", ""))
                or clean_text((ai or {}).get("section_name_english", ""))
            )
            section_maker = (
                clean_text((section or {}).get("maker", ""))
                or clean_text(direct.get("section_maker", ""))
                or clean_text((ai or {}).get("section_maker", ""))
            )
            section_model = (
                clean_text((section or {}).get("model", ""))
                or clean_text(direct.get("section_model", ""))
                or clean_text((ai or {}).get("section_model", ""))
            )
            first_page = (section or {}).get("first_page") or page

            confidence = max(0.88, clamp_confidence(direct.get("confidence", 0.88)))
            if description_source != "printed English":
                confidence = min(
                    confidence,
                    clamp_confidence((ai or {}).get("confidence", 0.78), 0.78),
                    0.82,
                )
            if language_review:
                confidence = min(confidence, 0.60)
            if section_conflict:
                confidence = min(confidence, 0.62)
            if section is None:
                confidence = min(confidence, 0.50)
            source_backed_engineering_row = bool(
                extraction_profile == ENGINEERING_ARTICLE_DRAWING_PROFILE
                and clean_text(direct.get("ident_no", ""))
                and clean_text(direct.get("section_context_source", ""))
                == "engineering article-table consensus"
            )
            if source_backed_engineering_row and not language_review and not section_conflict:
                confidence = max(confidence, 0.90)

            ident_no = clean_text(direct.get("ident_no", ""))
            resolved_item_no = clean_text(direct.get("item_no", "")) or clean_text(
                (ai or {}).get("item_no", "")
            )
            output.append(
                {
                    "source_page": page,
                    "section_start_page": int(first_page),
                    "section_code": section_code,
                    "section_name_english": section_name.upper(),
                    "detected_machinery": section_name.upper(),
                    "section_maker": section_maker.upper(),
                    "section_model": section_model.upper(),
                    "table_title": clean_text(direct.get("table_title", section_name)),
                    "ident_no": ident_no,
                    "part_no": ident_no,
                    "code": ident_no,
                    "item_no": resolved_item_no,
                    "description_english": description.upper(),
                    "description": description.upper(),
                    "source_description_raw": _clean_markdown_cell(direct.get("description_raw", "")),
                    "source_section_name_raw": _clean_markdown_cell(direct.get("table_title", "")),
                    "unit": normalize_unit((ai or {}).get("unit", ""), default_unit),
                    # This catalogue prints approximate weight, not quantity.
                    # QNT must therefore remain blank even if AI interpreted a
                    # neighbouring numeric column as quantity.
                    "quantity": None if allow_simple_section_codes else quantity_to_number(direct.get("quantity")),
                    "confidence": confidence,
                    "language_review": language_review,
                    "description_source": description_source,
                    "description_source_mismatch": ai_disagreed,
                    "section_review": bool(section_conflict or section is None),
                    "section_assignment_source": section_source,
                }
            )
    else:
        # Irregular/non-Markdown manuals still use the strict AI schema, but section
        # assignment remains anchored to the page map whenever one exists.
        for ai in normalized_ai:
            page = ai.get("source_page")
            if page in index_pages:
                continue
            explicit_spare_row = normalize_key(ai.get("source_pattern", "")) == normalize_key("explicit spare-number list")
            if not explicit_spare_row and page not in parts_pages:
                continue
            section, section_source, section_conflict = section_for(page, None, ai)
            ident_no = clean_text(ai.get("ident_no", ""))
            if explicit_spare_row:
                description = clean_text(ai.get("description_english", "")).upper()
                language_review = False
            else:
                description, language_review = _best_effort_english_description(
                    ai.get("description_english", "")
                )
            if not description or not (ident_no or clean_text(ai.get("item_no", ""))):
                continue
            if allow_simple_section_codes and not ident_no:
                continue
            confidence = clamp_confidence(ai.get("confidence", 0.70))
            if language_review:
                confidence = min(confidence, 0.60)
                language_exceptions += 1
            if section_conflict:
                if not explicit_spare_row:
                    confidence = min(confidence, 0.62)
                section_conflicts += 1
            if section is None:
                if not explicit_spare_row:
                    confidence = min(confidence, 0.50)
                unconfirmed_sections += 1
            output.append(
                {
                    **ai,
                    "section_code": safe_section_code((section or {}).get("code", "")) or safe_section_code(ai.get("section_code", "")),
                    "section_name_english": clean_text((section or {}).get("name", ai.get("section_name_english", ""))).upper(),
                    "detected_machinery": clean_text((section or {}).get("name", ai.get("section_name_english", ""))).upper(),
                    "section_maker": clean_text((section or {}).get("maker", "")) or clean_text(ai.get("section_maker", "")),
                    "section_model": clean_text((section or {}).get("model", "")) or clean_text(ai.get("section_model", "")),
                    "section_start_page": (section or {}).get("first_page") or ai.get("section_start_page") or page,
                    "part_no": ident_no,
                    "code": ident_no,
                    "description_english": description,
                    "description": description,
                    "source_description_raw": clean_text(ai.get("description_english", "")),
                    "source_section_name_raw": clean_text(ai.get("section_name_english", "")),
                    "confidence": confidence,
                    "language_review": language_review,
                    "description_source": "explicit printed spare-part number" if explicit_spare_row else "AI English/translation",
                    "description_source_mismatch": False,
                    "section_review": bool(section_conflict or section is None),
                    "section_assignment_source": section_source,
                    "quantity": None if allow_simple_section_codes else quantity_to_number(ai.get("quantity")),
                }
            )

    # AI-only/source-safety-net rows are accepted only on confirmed spare-source pages.
    # This includes formal parts tables and explicit labelled spare-number lists.
    if direct_rows:
        existing_keys = {
            (
                int(row.get("source_page") or 0),
                normalize_key(row.get("item_no", "")),
                normalize_key(row.get("ident_no", "")),
            )
            for row in output
        }
        for ai in normalized_ai:
            if id(ai) in matched_ai_ids or ai.get("source_page") not in parts_pages:
                continue
            key = (
                int(ai.get("source_page") or 0),
                normalize_key(ai.get("item_no", "")),
                normalize_key(ai.get("ident_no", "")),
            )
            if key in existing_keys:
                continue
            explicit_spare_row = normalize_key(ai.get("source_pattern", "")) == normalize_key("explicit spare-number list")
            if explicit_spare_row:
                description = clean_text(ai.get("description_english", "")).upper()
                language_review = False
            else:
                description, language_review = _best_effort_english_description(
                    ai.get("description_english", "")
                )
            if description and (
                clean_text(ai.get("ident_no", ""))
                or (
                    not allow_simple_section_codes
                    and clean_text(ai.get("item_no", ""))
                )
            ):
                section, section_source, section_conflict = section_for(
                    ai.get("source_page"), None, ai
                )
                ident_no = clean_text(ai.get("ident_no", ""))
                confidence = clamp_confidence(ai.get("confidence", 0.70)) if explicit_spare_row else min(clamp_confidence(ai.get("confidence", 0.70)), 0.82)
                if language_review:
                    confidence = min(confidence, 0.60)
                    language_exceptions += 1
                if section_conflict:
                    if not explicit_spare_row:
                        confidence = min(confidence, 0.62)
                    section_conflicts += 1
                if section is None:
                    if not explicit_spare_row:
                        confidence = min(confidence, 0.50)
                    unconfirmed_sections += 1
                output.append(
                    {
                        **ai,
                        "section_code": safe_section_code((section or {}).get("code", "")) or safe_section_code(ai.get("section_code", "")),
                        "section_name_english": clean_text((section or {}).get("name", ai.get("section_name_english", ""))).upper(),
                        "detected_machinery": clean_text((section or {}).get("name", ai.get("section_name_english", ""))).upper(),
                        "section_maker": clean_text((section or {}).get("maker", "")) or clean_text(ai.get("section_maker", "")),
                        "section_model": clean_text((section or {}).get("model", "")) or clean_text(ai.get("section_model", "")),
                        "section_start_page": (section or {}).get("first_page") or ai.get("source_page"),
                        "part_no": ident_no,
                        "code": ident_no,
                        "description_english": description,
                        "description": description,
                        "source_description_raw": clean_text(ai.get("description_english", "")),
                        "source_section_name_raw": clean_text(ai.get("section_name_english", "")),
                        "confidence": confidence,
                        "language_review": language_review,
                        "description_source": "explicit printed spare-part number" if explicit_spare_row else "AI-only row",
                        "description_source_mismatch": False,
                        "section_review": bool(section_conflict or section is None),
                        "section_assignment_source": section_source,
                        "quantity": None if allow_simple_section_codes else quantity_to_number(ai.get("quantity")),
                    }
                )

    output, reconciled_parent_rows, parent_identity_messages = (
        _reconcile_engineering_article_parent_identity(output, extracted_pages)
    )

    output, inherited_item_count = _carry_forward_merged_item_numbers(output)

    # Deterministic de-duplication by source page + item + identifier. In the
    # multilingual Order-No. catalogue, the Order-No. is globally unique. When
    # the same code is printed for several exact drawing positions/applicability
    # rows, retain one import record and aggregate those positions instead of
    # silently discarding every occurrence after the first.
    deduplicated: list[dict[str, Any]] = []
    seen: dict[tuple[str, str, str], dict[str, Any]] = {}
    consolidated_catalog_rows = 0

    def item_position_sort_key(value: str) -> tuple[Any, ...]:
        parts = re.split(r"(\d+)", clean_text(value))
        return tuple(int(part) if part.isdigit() else part.upper() for part in parts)

    for row in output:
        identifier_key = normalize_key(row.get("ident_no", row.get("code", "")))
        explicit_row = bool(row.get("explicit_spare_number", False)) or normalize_key(row.get("source_pattern", "")) == normalize_key("explicit spare-number list")
        if identifier_key and explicit_row:
            key = ("EXPLICIT", identifier_key, normalize_key(row.get("description_english", row.get("description", ""))))
        elif allow_simple_section_codes and identifier_key:
            # Order-No. is the unique catalogue identifier. If the same code is
            # printed again for another applicability line, keep one import row.
            key = ("ORDERNO", "", identifier_key)
        else:
            key = (
                str(row.get("source_page", "")),
                normalize_key(row.get("item_no", "")),
                identifier_key,
            )
        existing = seen.get(key)
        if existing is not None:
            if allow_simple_section_codes:
                positions = {
                    clean_text(value)
                    for source in (
                        existing.get("item_no", ""),
                        row.get("item_no", ""),
                    )
                    for value in re.split(r"\s*;\s*", clean_text(source))
                    if clean_text(value)
                }
                existing["item_no"] = "; ".join(
                    sorted(positions, key=item_position_sort_key)
                )
                existing["item_no_aggregated"] = len(positions) > 1
                existing["source_page"] = min(
                    int(existing.get("source_page") or 10**9),
                    int(row.get("source_page") or 10**9),
                )
                consolidated_catalog_rows += 1
            continue
        seen[key] = row
        deduplicated.append(row)

    messages = [
        f"Deterministic table verification found {len(direct_rows)} table-derived spare-part row(s).",
        f"Detected {len(sections)} source-coded sub-machinery section(s).",
    ]
    messages.extend(parent_identity_messages)
    if rejected_non_article_ai_rows:
        messages.append(
            f"Rejected {rejected_non_article_ai_rows} AI row(s) from construction/composition tables because no genuine Article No. was present."
        )
    if rejected_ai_identifiers:
        messages.append(
            f"Rejected {rejected_ai_identifiers} AI identifier value(s) that did not "
            "match a source Order-No. format; applicability, size, and damaged OCR "
            "fragments were not exported as spare-part codes."
        )
    if expanded_ai_identifiers:
        messages.append(
            f"Split {expanded_ai_identifiers} additional source Order-No. value(s) "
            "from AI cells that contained multiple printed identifiers."
        )
    if inherited_item_count:
        messages.append(
            f"Repeated {inherited_item_count} merged/continuation ITEM NO value(s) on linked spare-part rows."
        )
    if consolidated_catalog_rows:
        messages.append(
            f"Consolidated {consolidated_catalog_rows} repeated Order-No. occurrence(s) "
            "into unique spare-part codes while retaining every distinct ITEM NO."
        )
    if paraphrases_prevented:
        messages.append(
            f"Kept the exact printed English wording for {paraphrases_prevented} row(s) where the AI proposed a different phrase."
        )
    if language_exceptions:
        messages.append(
            f"{language_exceptions} description(s) could not be confidently isolated as English and remain low-confidence review items."
        )
    if section_conflicts:
        messages.append(
            f"Resolved {section_conflicts} section conflict(s) by giving the printed page header priority over AI context."
        )
    if ambiguous_pages:
        messages.append(
            "Ambiguous section headers detected on PDF page(s): "
            + ", ".join(str(page) for page in sorted(ambiguous_pages))
            + ". Previous section context was not carried across those pages."
        )
    if unmapped_parts_pages:
        messages.append(
            "No confirmed section mapping was available for parts-table page(s): "
            + ", ".join(str(page) for page in sorted(unmapped_parts_pages))
            + ". Those rows were kept at low confidence instead of inheriting a stale section."
        )
    if unconfirmed_sections:
        messages.append(
            f"{unconfirmed_sections} row(s) remain without a fully confirmed source section."
        )
    represented_section_codes = {
        normalize_key(row.get("section_code", ""))
        for row in deduplicated
        if normalize_key(row.get("section_code", ""))
    }
    empty_source_sections = [
        section
        for section in sections
        if section.get("pages")
        and normalize_key(section.get("code", "")) not in represented_section_codes
    ]
    if empty_source_sections:
        labels = ", ".join(
            f"{clean_text(section.get('code', ''))} {clean_text(section.get('name', ''))}".strip()
            for section in empty_source_sections
        )
        messages.append(
            "COMPLETENESS WARNING: source sub-machinery heading(s) were detected "
            f"without any exported spare rows: {labels}. Review or reprocess their pages."
        )
    return deduplicated, messages


# Backward-compatible local parser used by the app when AI extraction is disabled.
def extract_spare_parts_from_markdown_tables(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    rows, _ = prepare_benefit_rows(
        ai_rows=[],
        extracted_pages=extracted_pages,
        source_document_name="",
        main_row={},
        default_unit="PCS",
    )
    for row in rows:
        row["local_fallback"] = True
    return rows


# ---------------------------------------------------------------------------
# Automatic sub-machinery detection and assignment
# ---------------------------------------------------------------------------


_GENERIC_MACHINERY_NAMES = {
    "SPARE PARTS", "SPARE PARTS LIST", "PARTS LIST", "LIST OF PARTS",
    "PARTS CATALOGUE", "PARTS CATALOG", "CATALOGUE", "CATALOG",
    "DESCRIPTION", "DESIGNATION", "ITEM NO", "ITEM NUMBER", "PART NO",
    "PART NUMBER", "QUANTITY", "DRAWING", "DRAWING NO", "TABLE",
    "CONTINUED", "LIST OF SPARE PARTS",
    "DIMENSION DRAWINGS INCLUDING TECHNICAL DATA",
    "DRAWINGS", "GENERAL DRAWINGS", "RECOMMENDED SPARE PARTS ON BOARD",
    "DETAILED PAGE DESCRIPTIONS", "SECTION ON PAGE DESCRIPTION",
}


def _clean_machinery_name(value: Any) -> str:
    text = clean_text(value)
    text = re.sub(r"^\s*#{1,6}\s*", "", text).strip(" -:;|/")
    text = re.sub(r"^\d{1,3}(?:\.\d{1,3}){1,5}\s+", "", text)
    text = re.sub(
        r"^(?:spare\s+parts(?:\s+list)?|parts\s+list|list\s+of\s+parts)\s*(?:for|of)?\s*[:\-]*\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\s+(?:continued|cont\.?)(?:\s*\(.*?\))?$", "", text, flags=re.I)
    text = re.sub(r"\s+page\s+\d+(?:\s+of\s+\d+)?$", "", text, flags=re.I)
    return clean_text(text).strip(" -:;|/").upper()


def _is_generic_machinery_name(value: Any) -> bool:
    text = _clean_machinery_name(value)
    if not text or len(text) < 3:
        return True
    normalized = re.sub(r"[^A-Z0-9 ]+", " ", text.upper())
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if normalized in _GENERIC_MACHINERY_NAMES:
        return True
    if normalized.startswith(("REPLACE ", "RENEW ", "INSTALL ", "CHECK ", "REMOVE ")):
        return True
    if re.search(r"\|\s*\d{1,4}\s*$", clean_text(value)):
        return True
    words = re.findall(r"[A-Z0-9]+", normalized)
    if len(text) > 72 and len(words) > 9:
        return True
    if any(
        phrase in normalized
        for phrase in (
            "CAN EASILY BECOME DAMAGED", "EXERCISE GREAT CARE",
            "SECTION ON PAGE DESCRIPTION",
        )
    ):
        return True
    if any(
        phrase in normalized
        for phrase in (
            "INTERCONNECTION DIAGRAM", "CONNECTION DIAGRAM", "WIRING DIAGRAM",
            "CIRCUIT DIAGRAM", "FLOW DIAGRAM", "FLOW CHART", "CABLE LIST",
            "GENERAL ARRANGEMENT", "OPERATING SYSTEM PLAN",
        )
    ):
        return True
    if re.fullmatch(r"(?:FIG(?:URE)?|DWG|DRAWING|PAGE|TABLE)\s*[A-Z0-9./_-]*", normalized):
        return True
    if not re.search(r"[A-Z]", normalized):
        return True
    return False


def _machinery_name_similarity(left: str, right: str) -> float:
    left_clean = _clean_machinery_name(left).upper()
    right_clean = _clean_machinery_name(right).upper()
    if not left_clean or not right_clean:
        return 0.0
    left_key = normalize_key(left_clean)
    right_key = normalize_key(right_clean)
    if left_key == right_key:
        return 1.0
    sequence = SequenceMatcher(None, left_key, right_key).ratio()
    left_tokens = {token for token in re.findall(r"[A-Z0-9]+", left_clean) if len(token) > 1}
    right_tokens = {token for token in re.findall(r"[A-Z0-9]+", right_clean) if len(token) > 1}
    token_score = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    generic_suffixes = {"ASSEMBLY", "ASSY", "UNIT", "COMPLETE"}
    containment = 0.0
    if min(len(left_key), len(right_key)) >= 8 and (left_key in right_key or right_key in left_key):
        extra_tokens = (left_tokens | right_tokens) - (left_tokens & right_tokens)
        if extra_tokens and extra_tokens <= generic_suffixes:
            containment = 0.94
    return max(sequence, token_score, containment)


def _split_detection_keys(value: Any) -> set[str]:
    return {key.strip() for key in clean_text(value).split("|") if key.strip()}


def _generated_submachinery_code(position: int, existing_codes: set[str]) -> str:
    counter = max(1, int(position))
    while True:
        candidate = f"SUB-{counter:03d}"
        if normalize_key(candidate) not in existing_codes:
            existing_codes.add(normalize_key(candidate))
            return candidate
        counter += 1


def _submachinery_hierarchy_confidence(
    observations: Sequence[dict[str, Any]],
    resolved_code: str,
    resolved_name: str,
) -> float:
    """Return confidence in the *parent hierarchy*, not raw row OCR quality.

    Spare-row confidence and hierarchy confidence are different signals. A difficult
    rotated scan can legitimately give individual OCR rows very low confidence while
    the parent sub-machinery is still strongly proven by cumulative source evidence.

    The only strong automatic uplift is source-backed: at least three spare identifiers
    must independently expose the same structured parent stem as ``resolved_code`` and
    the detected hierarchy name must also agree across the group. Merely repeating the
    same low-confidence AI guess is not enough to manufacture a high confidence score.
    """
    if not observations:
        return 0.70

    row_confidences = [
        clamp_confidence(observation.get("confidence", 0.70))
        for observation in observations
    ]
    baseline = sum(row_confidences) / max(1, len(row_confidences))

    code_key = normalize_key(resolved_code)
    name_key = normalize_key(resolved_name)
    count = len(observations)
    if not code_key or not name_key:
        return baseline

    code_support = sum(
        1 for observation in observations
        if normalize_key(observation.get("code", "")) == code_key
    ) / count
    hierarchy_name_support = sum(
        1 for observation in observations
        if normalize_key(observation.get("name", "")) == name_key
    ) / count
    source_description_support = sum(
        1 for observation in observations
        if normalize_key(observation.get("part_description", "")) == name_key
    ) / count
    # A repeated source Name/Designation is independent evidence for the hierarchy
    # name on engineering Article-No. families. This lets a difficult title-block
    # OCR remain low quality without dragging a strongly corroborated parent name
    # down with it.
    name_support = max(hierarchy_name_support, source_description_support)

    # The strongest generic evidence is that the orderable spare identifiers
    # themselves encode the same parent/drawing stem. Examples include
    # ``9007280 99`` -> ``9007280`` and ``03036-58A`` -> ``03036``.
    stem_matches = 0
    stem_candidates = 0
    for observation in observations:
        stem = _engineering_article_parent_stem(
            observation.get("part_identifier", "")
        )
        if not stem:
            continue
        stem_candidates += 1
        if normalize_key(stem) == code_key:
            stem_matches += 1

    stem_support_all_rows = stem_matches / count
    stem_support_candidates = (
        stem_matches / stem_candidates if stem_candidates else 0.0
    )

    section_pages = [
        observation.get("section_page")
        for observation in observations
        if observation.get("section_page") is not None
    ]
    page_consistency = 0.0
    if section_pages:
        dominant_page_count = max(section_pages.count(page) for page in set(section_pages))
        page_consistency = dominant_page_count / len(section_pages)

    strong_source_consensus = bool(
        count >= 3
        and stem_matches >= 3
        and stem_support_all_rows >= 0.75
        and stem_support_candidates >= 0.90
        and code_support >= 0.75
        and name_support >= 0.75
    )
    if strong_source_consensus:
        # 0.92 means strong corroboration rather than perfect certainty. Give a
        # small additional uplift when all source rows agree on one section page.
        evidence_confidence = 0.94 if page_consistency >= 0.90 else 0.92
        return max(baseline, evidence_confidence)

    return baseline


def _submachinery_name_from_source_consensus(
    observations: Sequence[dict[str, Any]],
    resolved_code: str,
    current_name: str,
) -> str:
    """Prefer a strongly corroborated source designation over broad navigation text.

    This is intentionally generic and conservative. It activates only when at least
    three spare identifiers independently encode the same parent/drawing stem and at
    least 75% of those rows repeat the same concise source description. A broad
    chapter/navigation label can then be replaced by that repeated source designation.
    Exact short assembly titles are otherwise preserved.
    """
    if not observations:
        return clean_text(current_name).upper()

    code_key = normalize_key(resolved_code)
    current = clean_text(current_name).upper()
    if not code_key:
        return current

    count = len(observations)
    stems = [
        _engineering_article_parent_stem(obs.get("part_identifier", ""))
        for obs in observations
    ]
    stem_matches = sum(1 for stem in stems if stem and normalize_key(stem) == code_key)
    if count < 3 or stem_matches < 3 or stem_matches / count < 0.75:
        return current

    descriptions = [
        clean_text(obs.get("part_description", "")).upper()
        for obs in observations
        if clean_text(obs.get("part_description", ""))
    ]
    if not descriptions:
        return current
    description_counts = {value: descriptions.count(value) for value in set(descriptions)}
    dominant, dominant_count = max(description_counts.items(), key=lambda item: item[1])
    if (
        dominant_count < 3
        or dominant_count / count < 0.90
        or len(dominant) > 80
        or _is_generic_machinery_name(dominant)
    ):
        return current

    if not current:
        return dominant
    if normalize_key(current) == normalize_key(dominant):
        return current

    # Only replace a broader label when the repeated designation is a complete
    # word/phrase within it. This avoids turning VALVES into VALVE or replacing an
    # unrelated explicit assembly title merely because descriptions are repetitive.
    dominant_pattern = re.compile(
        rf"(?<![A-Z0-9]){re.escape(dominant)}(?![A-Z0-9])",
        flags=re.IGNORECASE,
    )
    dominant_is_component = bool(dominant_pattern.search(current))
    broader_context = (
        len(current) >= len(dominant) + 8
        and len(current.split()) >= 5
    )
    navigation_like = bool(
        "/" in current
        or " AND " in f" {current} "
        or " INCLUDING " in f" {current} "
        or "DRAWING" in current
        or "TECHNICAL DATA" in current
    )
    if dominant_is_component and broader_context and navigation_like:
        return dominant
    return current



_DRAWING_ONLY_EXCLUDED_TITLES = {
    "GENERAL ARRANGEMENT", "GENERAL DRAWING", "GENERAL DRAWINGS",
    "DIMENSION DRAWING", "DIMENSION DRAWINGS", "TECHNICAL DATA",
    "ELECTRICAL DRAWING", "ELECTRICAL DRAWINGS", "FLOW DIAGRAM",
    "FLOW SCHEME", "PIPING DIAGRAM", "CONNECTION LIST", "CABLE LIST",
    "INSTALLATION PRINCIPLE", "SYSTEM SCHEMATIC", "SYSTEM LAYOUT",
    "CIRCUIT DIAGRAM", "WIRING DIAGRAM", "DRAWING", "DRAWINGS",
}


def _clean_equipment_drawing_title(value: Any) -> str:
    """Return a compact equipment name from a drawing-title field.

    Drawing metadata such as ``dim. drw.``, ``dimension drawing`` and pagination /
    optional-scope parentheticals are not part of the machinery name. Short equipment
    acronyms/tags remain when they are genuinely part of the title.
    """
    text = clean_text(value).translate(
        str.maketrans({"\u2010": "-", "\u2011": "-", "\u2012": "-", "\u2013": "-", "\u2014": "-", "\u2212": "-"})
    ).strip(" -:;|,.")
    if not text:
        return ""
    text = re.sub(
        r"\s*\((?:[^)]*\b(?:PAGES?|PREVIOUSLY|OPTIONAL|INCLUDING|INCL\.?)\b[^)]*)\)\s*",
        " ", text, flags=re.IGNORECASE,
    )
    text = re.sub(
        r"(?:[,;:\-]\s*)?(?:DIM(?:ENSION)?\.?\s*)?(?:DRW\.?|DRAWING)(?:\s+INCLUDING\s+TECHNICAL\s+DATA)?\s*$",
        "", text, flags=re.IGNORECASE,
    )
    text = re.sub(r"\s+", " ", text).strip(" -:;|,.")
    # A chapter heading can describe the systems traversed by a drawing rather
    # than the item named in its title box (for example ``UV reactor and LDC /
    # Lamp power cable`` while the labelled Title is simply ``Cable``). Use the
    # terminal concrete component noun only for this narrow navigation pattern;
    # ordinary compound names such as ``Air filter / Filter regulator`` remain
    # intact.
    if "/" in text:
        parent_context, trailing_context = text.rsplit("/", maxsplit=1)
        trailing_tokens = re.findall(r"[A-Za-z]+", trailing_context.upper())
        if (
            " AND " in f" {parent_context.upper()} "
            and 2 <= len(trailing_tokens) <= 4
            and trailing_tokens[-1] in _DRAWING_COMPONENT_NOUNS
        ):
            text = trailing_tokens[-1]
    return _clean_machinery_name(text)


def _is_equipment_component_title(value: Any) -> bool:
    """High-precision gate for machinery-defining drawing titles.

    A source Document/Drawing No. is not enough by itself: the title must describe a
    concrete component/equipment item. General arrangements, schematics, cable lists,
    flow diagrams and other document-level drawings are intentionally rejected.
    """
    title = _clean_equipment_drawing_title(value)
    if not title or _is_generic_machinery_name(title):
        return False
    normalized = re.sub(r"[^A-Z0-9 ]+", " ", title.upper())
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if normalized.startswith(("NOTE ", "IF ", "WHEN ", "WHERE ")):
        return False
    if normalized in _DRAWING_ONLY_EXCLUDED_TITLES:
        return False
    if any(
        phrase in normalized
        for phrase in (
            "GENERAL ARRANGEMENT", "FLOW DIAGRAM", "FLOW SCHEME", "PIPING DIAGRAM",
            "CONNECTION LIST", "CABLE LIST", "SYSTEM SCHEMATIC", "SYSTEM LAYOUT",
            "INSTALLATION PRINCIPLE", "CIRCUIT DIAGRAM", "WIRING DIAGRAM",
            "FLOW CHART", "INTERCONNECTION DIAGRAM", "CONNECTION DIAGRAM",
            "SYSTEM PLAN", "OPERATING SYSTEM PLAN",
        )
    ):
        return False
    title_tokens = set(re.findall(r"[A-Z]+", normalized))
    return bool(title_tokens & _DRAWING_COMPONENT_NOUNS)


def _drawing_title_block_sections(
    extracted_pages: Sequence[tuple[int, str]],
) -> list[dict[str, Any]]:
    """Discover equipment drawings from labelled title blocks, with or without spares."""
    results: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    for page_number, markdown in extracted_pages:
        metadata = _engineering_title_block_metadata(
            markdown,
            require_article_table=False,
        )
        if not metadata:
            continue
        code = _valid_automatic_section_code(metadata.get("section_code", ""))
        title = _clean_equipment_drawing_title(
            metadata.get("section_name_english", metadata.get("section_name_raw", ""))
        )
        if not code or not _is_equipment_component_title(title):
            continue
        key = (normalize_key(code), int(page_number))
        if key in seen:
            continue
        seen.add(key)
        article_table = _is_engineering_article_table(markdown)
        results.append(
            {
                "code": code,
                "aliases": [code],
                "name": title,
                "maker": "",
                "model": "",
                "pages": {int(page_number)},
                "source": (
                    "engineering drawing title block"
                    if article_table
                    else "equipment drawing title block"
                ),
                "confidence": 0.97 if article_table else 0.95,
            }
        )
    return results


def build_component_drawing_candidates(
    extracted_pages: Sequence[tuple[int, str]],
    main_row: dict[str, Any],
    source_document_name: str = "",
) -> pd.DataFrame:
    """Build machinery candidates independently from spare-row discovery.

    A genuine equipment drawing can therefore create a Step-3 sub-machinery with
    ``PARTS FOUND = 0``. Only source-backed title blocks or the existing conservative
    numbered component-drawing detector qualify; generic document drawings are ignored.
    """
    if not extracted_pages:
        return empty_submachinery_review_dataframe()

    discoveries: list[dict[str, Any]] = []
    discoveries.extend(_drawing_title_block_sections(extracted_pages))
    for section in _drawing_heading_sections(extracted_pages):
        item = dict(section)
        item.setdefault("confidence", 0.92)
        discoveries.append(item)
    discoveries.extend(_unnumbered_equipment_drawing_sections(extracted_pages))
    if not discoveries:
        return empty_submachinery_review_dataframe()

    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in discoveries:
        source = clean_text(item.get("source", "")).lower()
        code = _valid_drawing_heading_code(item.get("code", ""))
        name = _clean_equipment_drawing_title(item.get("name", ""))
        if not code or not _is_equipment_component_title(name):
            continue
        grouped.setdefault(normalize_key(code), []).append({**item, "code": code, "name": name})

    records: list[dict[str, Any]] = []
    main_maker = clean_text(main_row.get("MAKER", "")).upper()
    main_model = clean_text(main_row.get("MODEL", "")).upper()
    source_file = clean_text(source_document_name)

    for group in grouped.values():
        if not group:
            continue
        def priority(item: dict[str, Any]) -> tuple[int, float]:
            source = clean_text(item.get("source", "")).lower()
            return (
                3 if "title block" in source else 2,
                clamp_confidence(item.get("confidence", 0.90)),
            )

        best = max(group, key=priority)
        code = clean_text(best.get("code", "")).upper()
        name = _clean_equipment_drawing_title(best.get("name", "")).upper()
        names = [_clean_equipment_drawing_title(item.get("name", "")).upper() for item in group]
        names = [value for value in names if value]
        pages = sorted({
            int(page)
            for item in group
            for page in item.get("pages", set())
            if page is not None
        })
        first_page = pages[0] if pages else None
        last_page = pages[-1] if pages else first_page
        confidence = max(clamp_confidence(item.get("confidence", 0.90)) for item in group)
        detection_keys = {normalize_key(code), normalize_key(name)}
        detection_keys.update(normalize_key(value) for value in names if value)
        aliases = [clean_text(alias) for item in group for alias in item.get("aliases", [])]
        detection_keys.update(normalize_key(value) for value in aliases if value)
        origins = {clean_text(item.get("source", "")) for item in group if clean_text(item.get("source", ""))}

        records.append(
            {
                "INCLUDE": True,
                "CODE": code,
                "NAME": name,
                "MAKER": main_maker,
                "MODEL": main_model,
                "TYPE": "",
                "INSTR.BOOK": f"PDF PAGE.{first_page}" if first_page is not None else "",
                "SPECIFICATIONS": f"PDF FILE: {source_file}" if source_file else "",
                "MCH_TP(M/S/U)": "SubMachinery",
                "FIRST PAGE": first_page,
                "LAST PAGE": last_page,
                "PARTS FOUND": 0,
                "CONFIDENCE": confidence,
                "VARIANTS": " | ".join(dict.fromkeys(names)),
                "DETECTION KEYS": "|".join(sorted(key for key in detection_keys if key)),
                "ORIGIN": (
                    "Auto-detected title-block drawing"
                    if any("title block" in origin.lower() for origin in origins)
                    else "Auto-detected drawing/component"
                ),
            }
        )

    if not records:
        return empty_submachinery_review_dataframe()
    frame = pd.DataFrame(records, columns=SUBMACHINERY_REVIEW_COLUMNS)
    frame["INCLUDE"] = frame["INCLUDE"].astype(bool)
    frame["FIRST PAGE"] = pd.to_numeric(frame["FIRST PAGE"], errors="coerce").astype("Int64")
    frame["LAST PAGE"] = pd.to_numeric(frame["LAST PAGE"], errors="coerce").astype("Int64")
    frame["PARTS FOUND"] = pd.to_numeric(frame["PARTS FOUND"], errors="coerce").fillna(0).astype(int)
    frame["CONFIDENCE"] = pd.to_numeric(frame["CONFIDENCE"], errors="coerce").fillna(0.90)
    return frame.sort_values(["FIRST PAGE", "NAME"], na_position="last").reset_index(drop=True)


def merge_component_candidates_with_native_priority(
    ocr_candidates: pd.DataFrame | None,
    native_candidates: pd.DataFrame | None,
) -> pd.DataFrame:
    """Combine native chapter evidence with exact-code OCR title-block evidence.

    Native searchable text remains authoritative for the drawing/document number
    and suppresses conflicting OCR proposals from that page. When OCR independently
    finds the *same* code on the same page with stronger title-block confidence, its
    compact Title value replaces the broader chapter/section heading. This preserves
    code reliability while correctly preferring ``Title: Cable`` over a heading such
    as ``UV reactor and LDC / Lamp power cable``.
    """
    ocr = (
        ocr_candidates.copy()
        if ocr_candidates is not None
        else empty_submachinery_review_dataframe()
    )
    native = (
        native_candidates.copy()
        if native_candidates is not None
        else empty_submachinery_review_dataframe()
    )
    if native.empty:
        return ocr
    if ocr.empty:
        return native

    reconciled_native = native.copy()
    ocr_pages = pd.to_numeric(ocr["FIRST PAGE"], errors="coerce")
    ocr_codes = ocr["CODE"].map(normalize_key)
    for native_index, native_row in reconciled_native.iterrows():
        native_name = _clean_equipment_drawing_title(
            native_row.get("NAME", "")
        )
        if native_name and normalize_key(native_name) != normalize_key(
            native_row.get("NAME", "")
        ):
            reconciled_native.at[native_index, "NAME"] = native_name.upper()
        native_page = pd.to_numeric(
            pd.Series([native_row.get("FIRST PAGE")]), errors="coerce"
        ).iloc[0]
        native_code = normalize_key(native_row.get("CODE", ""))
        if pd.isna(native_page) or not native_code:
            continue
        exact_matches = ocr.loc[
            ocr_pages.eq(int(native_page)) & ocr_codes.eq(native_code)
        ]
        if exact_matches.empty:
            continue
        native_confidence = clamp_confidence(native_row.get("CONFIDENCE", 0.90))
        best_index = max(
            exact_matches.index,
            key=lambda index: clamp_confidence(
                exact_matches.at[index, "CONFIDENCE"]
            ),
        )
        best_ocr = exact_matches.loc[best_index]
        ocr_confidence = clamp_confidence(best_ocr.get("CONFIDENCE", 0.90))
        ocr_name = _clean_equipment_drawing_title(best_ocr.get("NAME", ""))
        ocr_origin = clean_text(best_ocr.get("ORIGIN", "")).lower()
        labelled_compact_match = bool(
            "title-block" in ocr_origin
            and native_name
            and len(ocr_name) < len(native_name)
            and re.search(
                rf"(?<![A-Z0-9]){re.escape(ocr_name)}(?![A-Z0-9])$",
                native_name,
                flags=re.I,
            )
        )
        if (
            ocr_name
            and _is_equipment_component_title(ocr_name)
            and (ocr_confidence > native_confidence or labelled_compact_match)
        ):
            variants = [
                value
                for value in (
                    ocr_name.upper(),
                    native_name.upper(),
                    clean_text(best_ocr.get("VARIANTS", "")),
                    clean_text(native_row.get("VARIANTS", "")),
                )
                if value
            ]
            detection_keys = {
                key
                for key in (
                    normalize_key(native_code),
                    normalize_key(ocr_name),
                    normalize_key(native_name),
                    *_split_detection_keys(best_ocr.get("DETECTION KEYS", "")),
                    *_split_detection_keys(native_row.get("DETECTION KEYS", "")),
                )
                if key
            }
            reconciled_native.at[native_index, "NAME"] = ocr_name.upper()
            reconciled_native.at[native_index, "CONFIDENCE"] = ocr_confidence
            reconciled_native.at[native_index, "VARIANTS"] = " | ".join(
                dict.fromkeys(variants)
            )
            reconciled_native.at[native_index, "DETECTION KEYS"] = "|".join(
                sorted(detection_keys)
            )
            reconciled_native.at[native_index, "ORIGIN"] = (
                "Auto-detected title-block name with native code confirmation"
            )

    native_pages = {
        int(value)
        for value in pd.to_numeric(
            reconciled_native["FIRST PAGE"], errors="coerce"
        ).dropna()
    }
    ocr = ocr.loc[~ocr_pages.isin(native_pages)].copy()
    return merge_submachinery_candidates(ocr, reconciled_native)


def ensure_component_drawing_detail_spare_rows(
    review_frame: pd.DataFrame | None,
    extracted_pages: Sequence[tuple[int, str]],
    component_candidates: pd.DataFrame | None,
    default_unit: str = "PCS",
) -> tuple[pd.DataFrame, int, int, int]:
    """Add source-backed drawing-table rows and zero-detail legend rows.

    Article-No./Name-Designation tables are always stronger than legends. A
    labelled legend is used only when the exact drawing parent otherwise has no
    detailed spare row. Existing CODE/PART NO values are never duplicated, which
    preserves the application's global unique-code rule.

    Returns ``(review, article_rows_added, legend_rows_added, duplicates_skipped)``.
    """
    review = (
        review_frame.copy()
        if review_frame is not None
        else empty_review_dataframe()
    )
    for column in REVIEW_COLUMNS:
        if column not in review.columns:
            review[column] = False if column in {"INCLUDE", "READY"} else ""
    if (
        component_candidates is None
        or component_candidates.empty
        or not extracted_pages
    ):
        return review[REVIEW_COLUMNS].reset_index(drop=True), 0, 0, 0

    unit = clean_text(default_unit).upper()
    if unit not in UNIT_OPTIONS or not unit:
        unit = "PCS"

    page_text = {int(page): str(text or "") for page, text in extracted_pages}
    direct_by_page: dict[int, list[dict[str, Any]]] = {}
    for row in _direct_table_rows(extracted_pages):
        page_value = quantity_to_number(row.get("source_page"))
        if page_value is not None:
            direct_by_page.setdefault(int(page_value), []).append(dict(row))

    existing_keys: set[str] = set()
    for _, row in review.iterrows():
        for value in (row.get("CODE", ""), row.get("PART NO", "")):
            key = normalize_key(value)
            if key:
                existing_keys.add(key)

    def parent_has_detail(code: str, name: str) -> bool:
        code_key = normalize_key(code)
        name_key = normalize_key(name)
        for _, row in review.iterrows():
            if not bool(row.get("INCLUDE", False)):
                continue
            row_identifier = normalize_key(
                row.get("CODE", row.get("PART NO", ""))
            )
            if not row_identifier or row_identifier == code_key:
                # The whole-drawing assembly row is a fallback, not a detailed spare.
                continue
            row_section = normalize_key(row.get("SECTION CODE", ""))
            row_names = {
                normalize_key(row.get("MACHINERY", "")),
                normalize_key(row.get("DETECTED MACHINERY", "")),
            }
            if (code_key and row_section == code_key) or (name_key and name_key in row_names):
                return True
        return False

    added_rows: list[dict[str, Any]] = []
    article_added = 0
    legend_added = 0
    duplicates_skipped = 0

    candidates = component_candidates.sort_values(
        ["FIRST PAGE", "NAME"], na_position="last"
    )
    # A component drawing can continue on one or two pages without repeating the
    # numbered section heading. Link only continuation pages that independently
    # expose orderable table/legend evidence and stop before the next component.
    # This recovers, for example, a second V212-31 drawing page without creating a
    # new sub-machinery from its document reference.
    candidate_rows = [row.copy() for _, row in candidates.iterrows()]
    expanded_candidate_rows: list[pd.Series] = []
    for position, candidate in enumerate(candidate_rows):
        expanded_candidate_rows.append(candidate)
        first_value = quantity_to_number(candidate.get("FIRST PAGE"))
        if first_value is None:
            continue
        first_page = int(first_value)
        next_page = None
        for following in candidate_rows[position + 1:]:
            following_value = quantity_to_number(following.get("FIRST PAGE"))
            if following_value is not None and int(following_value) > first_page:
                next_page = int(following_value)
                break
        continuation_end = first_page + 2
        if next_page is not None:
            continuation_end = min(continuation_end, next_page - 1)
        for continuation_page in range(first_page + 1, continuation_end + 1):
            continuation_markdown = page_text.get(continuation_page, "")
            if not continuation_markdown:
                continue
            has_source_details = bool(
                direct_by_page.get(continuation_page)
                or _headerless_engineering_variant_rows(
                    continuation_page, continuation_markdown, {}
                )
                or _engineering_legend_entries(continuation_markdown)
            )
            if not has_source_details:
                continue
            continuation_candidate = candidate.copy()
            continuation_candidate["FIRST PAGE"] = continuation_page
            continuation_candidate["LAST PAGE"] = continuation_page
            expanded_candidate_rows.append(continuation_candidate)
    if expanded_candidate_rows:
        candidates = pd.DataFrame(expanded_candidate_rows)
    for _, candidate in candidates.iterrows():
        if "INCLUDE" in component_candidates.columns and not bool(
            candidate.get("INCLUDE", False)
        ):
            continue
        code = _valid_drawing_heading_code(candidate.get("CODE", ""))
        name = _clean_equipment_drawing_title(candidate.get("NAME", "")).upper()
        page_value = quantity_to_number(candidate.get("FIRST PAGE"))
        if not code or not name or page_value is None:
            continue
        page_number = int(page_value)
        markdown = page_text.get(page_number, "")
        if not markdown:
            continue

        def append_detail(
            identifier: Any,
            description: Any,
            *,
            item_no: Any = "",
            quantity: Any = None,
            source_kind: str,
            confidence: float,
        ) -> bool:
            nonlocal duplicates_skipped
            ident = clean_text(identifier).upper()
            desc = clean_text(description).upper()
            ident_key = normalize_key(ident)
            if (
                not ident
                or not desc
                or not ident_key
                or ident_key == normalize_key(code)
            ):
                return False
            if ident_key in existing_keys:
                duplicates_skipped += 1
                return False
            existing_keys.add(ident_key)
            row = {column: "" for column in REVIEW_COLUMNS}
            row.update(
                {
                    "INCLUDE": True,
                    "READY": False,
                    "MACHINERY": name,
                    "PART NO": ident,
                    "DESCRIPTION": desc,
                    "CODE": ident,
                    "ITEM NO": clean_text(item_no).upper(),
                    "UNIT": unit,
                    "QNT": quantity_to_number(quantity),
                    "SOURCE PAGE": page_number,
                    "SECTION START PAGE": page_number,
                    "TABLE TITLE": name,
                    "SECTION CODE": code,
                    "SECTION MAKER": clean_text(candidate.get("MAKER", "")).upper(),
                    "SECTION MODEL": clean_text(candidate.get("MODEL", "")).upper(),
                    "CONFIDENCE": confidence,
                    "DETECTED MACHINERY": name,
                    "ASSIGNMENT SOURCE": source_kind,
                    "WARNING": f"Source: {source_kind.lower()}",
                }
            )
            added_rows.append(row)
            return True

        page_direct_rows = direct_by_page.get(page_number, [])
        if not page_direct_rows:
            # Headerless engineering variant lists (for example a repeated
            # 9007170 family with DN/A valve sizes) are still orderable source
            # rows even though no Article-No. heading is printed.
            page_direct_rows = _headerless_engineering_variant_rows(
                page_number,
                markdown,
                {
                    "section_code": code,
                    "section_name_raw": name,
                    "section_name_english": name,
                    "maker": clean_text(candidate.get("MAKER", "")),
                    "model": clean_text(candidate.get("MODEL", "")),
                },
            )
        if page_direct_rows:
            for direct in page_direct_rows:
                description = clean_text(
                    direct.get(
                        "description_raw",
                        direct.get("description_english", direct.get("description", "")),
                    )
                )
                if append_detail(
                    direct.get("ident_no", direct.get("code", "")),
                    description,
                    item_no=direct.get("item_no", ""),
                    quantity=direct.get("quantity"),
                    source_kind=(
                        "Headerless drawing variant table"
                        if direct.get("source_layout")
                        == "headerless two-column engineering variant table"
                        else "Embedded drawing Article-No. table"
                    ),
                    confidence=max(
                        0.90,
                        clamp_confidence(direct.get("confidence", 0.90)),
                    ),
                ):
                    article_added += 1
            # A genuine Article-No. table is the orderable source for this page;
            # legend callouts are not duplicated as separate rows.
            continue

        if parent_has_detail(code, name):
            continue
        for legend_code, legend_name in _engineering_legend_entries(markdown):
            if append_detail(
                legend_code,
                legend_name,
                source_kind="Equipment drawing legend",
                confidence=0.88,
            ):
                legend_added += 1

    if added_rows:
        review = pd.concat(
            [review, pd.DataFrame(added_rows, columns=REVIEW_COLUMNS)],
            ignore_index=True,
        )
    return (
        review[REVIEW_COLUMNS].reset_index(drop=True),
        article_added,
        legend_added,
        duplicates_skipped,
    )


def remove_component_assembly_spare_rows(
    review_frame: pd.DataFrame | None,
    component_candidates: pd.DataFrame | None,
) -> tuple[pd.DataFrame, int]:
    """Remove rows that repeat a drawing parent as one of its own spare parts.

    A title-block Document No./Title defines hierarchy only. Genuine children must
    come from Article-No., orderable table, explicit spare-number, or legend evidence.
    Therefore a row whose spare CODE/PART NO equals its own component drawing code is
    removed, including assembly fallbacks stored by builds 4.18.4-4.18.11.
    """
    review = (
        review_frame.copy()
        if review_frame is not None
        else empty_review_dataframe()
    )
    for column in REVIEW_COLUMNS:
        if column not in review.columns:
            review[column] = False if column in {"INCLUDE", "READY"} else ""
    if review.empty or component_candidates is None or component_candidates.empty:
        return review[REVIEW_COLUMNS].reset_index(drop=True), 0

    component_names: dict[str, set[str]] = {}
    for _, candidate in component_candidates.iterrows():
        code = _valid_drawing_heading_code(candidate.get("CODE", ""))
        name = _clean_equipment_drawing_title(candidate.get("NAME", "")).upper()
        code_key = normalize_key(code)
        if not code_key or not _is_equipment_component_title(name):
            continue
        component_names.setdefault(code_key, set()).add(normalize_key(name))

    known_component_names = {
        name_key
        for names in component_names.values()
        for name_key in names
        if name_key
    }
    remove_indexes: list[Any] = []
    for index, row in review.iterrows():
        row_identifiers = {
            normalize_key(row.get("CODE", "")),
            normalize_key(row.get("PART NO", "")),
        }
        row_identifiers.discard("")
        # A code belongs to exactly one hierarchy level. Once it has been proven
        # as a component drawing Document No., it cannot also remain a spare code,
        # even when a weak AI assignment attached that row to the wrong parent.
        # The previous same-parent condition allowed precisely that misassignment
        # to escape cleanup and appear in the Spares step.
        if row_identifiers.intersection(component_names):
            remove_indexes.append(index)
            continue

        # Builds 4.18.4-4.18.11 also emitted explicit title-block assembly rows.
        # Remove those by their source provenance when their description is the
        # component title; this covers an old OCR-suffixed identifier without
        # risking deletion of a genuine source-table row with a similar name.
        assignment_source = clean_text(row.get("ASSIGNMENT SOURCE", "")).lower()
        warning = clean_text(row.get("WARNING", "")).lower()
        title_block_fallback = bool(
            "equipment drawing title block" in assignment_source
            or "title block assembly" in warning
        )
        description_key = normalize_key(row.get("DESCRIPTION", ""))
        if title_block_fallback and description_key in known_component_names:
            remove_indexes.append(index)

    if remove_indexes:
        review = review.drop(index=remove_indexes)
    return review[REVIEW_COLUMNS].reset_index(drop=True), len(remove_indexes)


def ensure_component_assembly_spare_rows(
    review_frame: pd.DataFrame | None,
    component_candidates: pd.DataFrame | None,
    default_unit: str = "PCS",
) -> tuple[pd.DataFrame, int, int]:
    """Deprecated compatibility wrapper; assembly fallbacks are no longer created."""
    cleaned, removed = remove_component_assembly_spare_rows(
        review_frame,
        component_candidates,
    )
    return cleaned, 0, removed

def build_submachinery_candidates(
    review_frame: pd.DataFrame,
    main_row: dict[str, Any],
    source_document_name: str = "",
) -> pd.DataFrame:
    """Create source-code-driven Benefit sub-machinery rows from extracted parts."""
    if review_frame is None or review_frame.empty:
        return empty_submachinery_review_dataframe()

    observations: list[dict[str, Any]] = []
    for _, row in review_frame.iterrows():
        name = _clean_catalog_section_name(
            _clean_machinery_name(
                row.get("DETECTED MACHINERY", row.get("TABLE TITLE", ""))
            ),
            row.get("SECTION CODE", ""),
        )
        raw_code = clean_text(row.get("SECTION CODE", "")).upper()
        table_title = clean_text(row.get("TABLE TITLE", ""))
        explicit_source = any(
            marker in table_title.lower()
            for marker in (
                "explicit spare-number", "recommended spare parts",
            )
        )
        structured_short_code = bool(
            clean_text(row.get("ITEM NO", ""))
            and re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", raw_code)
        )
        valid_code = _valid_automatic_section_code(raw_code) if raw_code else ""
        if _is_generic_machinery_name(name):
            continue
        # Code-less semantic parents are accepted only from explicitly labelled
        # spare lists. Invalid OCR codes from prose/warnings cannot manufacture a
        # new SUB-### proposal. Numbered catalogue parents retain their established
        # item-position evidence.
        if not valid_code and not (explicit_source or structured_short_code):
            continue
        code = valid_code
        source = quantity_to_number(row.get("SOURCE PAGE"))
        section = quantity_to_number(row.get("SECTION START PAGE"))
        observations.append(
            {
                "name": name,
                "code": code,
                "maker": clean_text(row.get("SECTION MAKER", "")).upper(),
                "model": clean_text(row.get("SECTION MODEL", "")).upper(),
                "source_page": int(source) if source is not None else None,
                "section_page": int(section) if section is not None else None,
                "confidence": clamp_confidence(row.get("CONFIDENCE", 0.70)),
                "part_identifier": next(
                    (
                        value
                        for value in (
                            clean_text(row.get("CODE", "")),
                            clean_text(row.get("PART NO", "")),
                        )
                        if value and _engineering_article_parent_stem(value)
                    ),
                    clean_text(row.get("CODE", ""))
                    or clean_text(row.get("PART NO", "")),
                ),
                "part_description": clean_text(row.get("DESCRIPTION", "")).upper(),
                "item_no": clean_text(row.get("ITEM NO", "")),
            }
        )
    if not observations:
        return empty_submachinery_review_dataframe()

    groups: dict[str, list[dict[str, Any]]] = {}
    for observation in observations:
        group_key = normalize_key(observation["code"]) or normalize_key(observation["name"])
        if not group_key:
            continue
        groups.setdefault(group_key, []).append(observation)

    preliminary: list[dict[str, Any]] = []
    for group_observations in groups.values():
        code_values = [obs["code"] for obs in group_observations if obs["code"]]
        name_values = [obs["name"] for obs in group_observations if obs["name"]]
        maker_values = [obs["maker"] for obs in group_observations if obs["maker"]]
        model_values = [obs["model"] for obs in group_observations if obs["model"]]
        code = max(set(code_values), key=code_values.count) if code_values else ""
        name = max(set(name_values), key=lambda value: (name_values.count(value), -len(value))) if name_values else ""
        name = _submachinery_name_from_source_consensus(
            group_observations, code, name
        )
        if re.fullmatch(r"\d{1,2}(?:\.\d{1,2})?", clean_text(code)) and not any(
            clean_text(observation.get("item_no", ""))
            for observation in group_observations
        ):
            continue
        if code and not _valid_automatic_section_code(code):
            code = ""
        maker = max(set(maker_values), key=maker_values.count) if maker_values else clean_text(main_row.get("MAKER", "")).upper()
        model = max(set(model_values), key=model_values.count) if model_values else clean_text(main_row.get("MODEL", "")).upper()
        pages = [obs["source_page"] for obs in group_observations if obs["source_page"] is not None]
        section_pages = [obs["section_page"] for obs in group_observations if obs["section_page"] is not None]
        first_page = min(section_pages or pages) if (section_pages or pages) else None
        preliminary.append(
            {
                "code": code,
                "name": name.upper(),
                "maker": maker,
                "model": model,
                "first_page": first_page,
                "last_page": max(pages) if pages else first_page,
                "parts": len(group_observations),
                "confidence": _submachinery_hierarchy_confidence(
                    group_observations, code, name
                ),
                "detection_keys": "|".join(sorted({normalize_key(obs["name"]) for obs in group_observations if obs["name"]} | {normalize_key(code)})),
                "variants": " | ".join(sorted(set(name_values), key=str.upper)),
            }
        )

    records: list[dict[str, Any]] = []
    used_codes = {
        normalize_key(item["code"])
        for item in preliminary
        if clean_text(item["code"])
    }
    for item in preliminary:
        benefit_name = _clean_catalog_section_name(item["name"], item["code"])
        benefit_code = clean_text(item["code"]).upper()
        if not benefit_code:
            benefit_code = _generated_submachinery_code(
                len(records) + 1,
                used_codes,
            )
        records.append(
            {
                "INCLUDE": True,
                "CODE": benefit_code,
                "NAME": benefit_name,
                "MAKER": item["maker"],
                "MODEL": item["model"],
                "TYPE": "",
                "INSTR.BOOK": f"PDF PAGE.{item['first_page']}" if item["first_page"] is not None else "",
                "SPECIFICATIONS": f"PDF FILE: {clean_text(source_document_name)}" if clean_text(source_document_name) else "",
                "MCH_TP(M/S/U)": "SubMachinery",
                "FIRST PAGE": item["first_page"],
                "LAST PAGE": item["last_page"],
                "PARTS FOUND": item["parts"],
                "CONFIDENCE": item["confidence"],
                "VARIANTS": item["variants"],
                "DETECTION KEYS": item["detection_keys"],
                "ORIGIN": "Auto-matched by source code",
            }
        )

    frame = pd.DataFrame(records, columns=SUBMACHINERY_REVIEW_COLUMNS)
    frame["INCLUDE"] = frame["INCLUDE"].astype(bool)
    frame["FIRST PAGE"] = pd.to_numeric(frame["FIRST PAGE"], errors="coerce").astype("Int64")
    frame["LAST PAGE"] = pd.to_numeric(frame["LAST PAGE"], errors="coerce").astype("Int64")
    frame["PARTS FOUND"] = pd.to_numeric(frame["PARTS FOUND"], errors="coerce").fillna(0).astype(int)
    frame["CONFIDENCE"] = pd.to_numeric(frame["CONFIDENCE"], errors="coerce").fillna(0.70)
    return frame.sort_values(["FIRST PAGE", "NAME"], na_position="last").reset_index(drop=True)


def refresh_submachinery_derived_fields(
    existing: pd.DataFrame | None,
    review_frame: pd.DataFrame | None,
    main_row: dict[str, Any],
    source_document_name: str = "",
    component_candidates: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Refresh Step-3 evidence from both spare rows and component drawings.

    A genuine equipment drawing may have zero spare rows, so Step 3 audit evidence
    cannot be reconstructed solely from ``review_frame``. Export later removes
    zero-part candidates from the import hierarchy. Manual edits remain
    authoritative because ``merge_submachinery_candidates`` preserves manually
    overridden values.
    """
    spare_detected = (
        build_submachinery_candidates(
            review_frame,
            main_row,
            source_document_name=source_document_name,
        )
        if review_frame is not None and not review_frame.empty
        else empty_submachinery_review_dataframe()
    )
    # Spare-table headings establish semantic parents, while a corroborating
    # equipment drawing supplies the authoritative printed code/name.
    detected = merge_submachinery_candidates(spare_detected, component_candidates)
    # Automatic proposals are derived state: rebuild them from current evidence
    # instead of accumulating stale OCR candidates across Streamlit reruns. Only
    # rows explicitly edited/created by the user survive independently.
    manual_existing = empty_submachinery_review_dataframe()
    if existing is not None and not existing.empty:
        working_existing = existing.copy()
        for column in SUBMACHINERY_REVIEW_COLUMNS:
            if column not in working_existing.columns:
                working_existing[column] = False if column == "INCLUDE" else ""
        origin = working_existing["ORIGIN"].fillna("").astype(str)
        included = working_existing["INCLUDE"].fillna(False).astype(bool)
        parts_found = pd.to_numeric(
            working_existing["PARTS FOUND"], errors="coerce"
        ).fillna(0)
        page_marker_artifact = working_existing["CODE"].map(
            lambda value: bool(
                re.search(
                    r"(?:PAGE|PAPER|SHEET)\d+(?:OF\d+)?$",
                    normalize_key(value),
                    flags=re.I,
                )
            )
        )
        historical_false_override = (
            origin.eq("Manual override")
            & ~included
            & parts_found.eq(0)
            & page_marker_artifact
        )
        manual_existing = working_existing.loc[
            ~origin.str.startswith("Auto") & ~historical_false_override,
            SUBMACHINERY_REVIEW_COLUMNS,
        ].copy()

    result = (
        detected.copy()
        if detected is not None
        else empty_submachinery_review_dataframe()
    )
    if not manual_existing.empty:
        result = merge_submachinery_candidates(result, manual_existing)
    if result.empty:
        return empty_submachinery_review_dataframe()

    # INCLUDE is a review decision, not derived OCR evidence. Preserve it for a
    # candidate that still exists even when an older saved session did not mark
    # the row as a manual override. Stale automatic candidates are still omitted
    # because only rows present in the freshly rebuilt result are considered.
    if existing is not None and not existing.empty:
        decision_source = existing.copy()
        for column in SUBMACHINERY_REVIEW_COLUMNS:
            if column not in decision_source.columns:
                decision_source[column] = False if column == "INCLUDE" else ""
        for result_index, result_row in result.iterrows():
            result_code = normalize_key(result_row.get("CODE", ""))
            result_keys = _split_detection_keys(result_row.get("DETECTION KEYS", ""))
            matching_index: Any = None
            for old_index, old_row in decision_source.iterrows():
                old_code = normalize_key(old_row.get("CODE", ""))
                old_keys = _split_detection_keys(old_row.get("DETECTION KEYS", ""))
                if (result_code and old_code == result_code) or bool(result_keys & old_keys):
                    matching_index = old_index
                    break
            if matching_index is not None:
                result.at[result_index, "INCLUDE"] = bool(
                    decision_source.at[matching_index, "INCLUDE"]
                )

    result["MCH_TP(M/S/U)"] = "SubMachinery"
    result["CONFIDENCE"] = pd.to_numeric(result["CONFIDENCE"], errors="coerce").fillna(0.70)
    result["PARTS FOUND"] = pd.to_numeric(result["PARTS FOUND"], errors="coerce").fillna(0).astype(int)
    result["FIRST PAGE"] = pd.to_numeric(result["FIRST PAGE"], errors="coerce").astype("Int64")
    result["LAST PAGE"] = pd.to_numeric(result["LAST PAGE"], errors="coerce").astype("Int64")
    return result[SUBMACHINERY_REVIEW_COLUMNS].reset_index(drop=True)

def merge_submachinery_candidates(
    existing: pd.DataFrame | None,
    detected: pd.DataFrame | None,
) -> pd.DataFrame:
    """Merge refreshed detections while preserving manually created/edited rows."""
    if existing is None or existing.empty:
        return detected.copy() if detected is not None else empty_submachinery_review_dataframe()
    if detected is None or detected.empty:
        return existing.copy()

    result = existing.copy()
    for column in SUBMACHINERY_REVIEW_COLUMNS:
        if column not in result.columns:
            result[column] = False if column == "INCLUDE" else ""
    result = result[SUBMACHINERY_REVIEW_COLUMNS]

    for _, new_row in detected.iterrows():
        new_code = normalize_key(new_row.get("CODE", ""))
        new_keys = _split_detection_keys(new_row.get("DETECTION KEYS", ""))
        match_index: Any = None
        for index, old_row in result.iterrows():
            old_code = normalize_key(old_row.get("CODE", ""))
            old_keys = _split_detection_keys(old_row.get("DETECTION KEYS", ""))
            if (new_code and old_code == new_code) or bool(new_keys & old_keys):
                match_index = index
                break
        if match_index is None:
            result = pd.concat([result, pd.DataFrame([new_row], columns=SUBMACHINERY_REVIEW_COLUMNS)], ignore_index=True)
            continue

        old_origin = clean_text(result.at[match_index, "ORIGIN"])
        # Auto rows are fully refreshed. Manual rows retain all user-controlled values.
        refresh_columns = (
            "FIRST PAGE", "LAST PAGE", "PARTS FOUND", "CONFIDENCE", "VARIANTS", "DETECTION KEYS"
        )
        if old_origin.startswith("Auto"):
            refresh_columns = (
                "INCLUDE", "CODE", "NAME", "MAKER", "MODEL", "TYPE", "INSTR.BOOK", "SPECIFICATIONS",
                "FIRST PAGE", "LAST PAGE", "PARTS FOUND", "CONFIDENCE", "VARIANTS", "DETECTION KEYS", "ORIGIN",
            )
        for column in refresh_columns:
            result.at[match_index, column] = new_row.get(column, result.at[match_index, column])

    result["MCH_TP(M/S/U)"] = "SubMachinery"
    return result[SUBMACHINERY_REVIEW_COLUMNS].reset_index(drop=True)


def add_manual_submachinery_candidate(
    frame: pd.DataFrame | None,
    main_row: dict[str, Any],
) -> pd.DataFrame:
    existing = frame.copy() if frame is not None else empty_submachinery_review_dataframe()
    used_codes = {
        normalize_key(value)
        for value in existing.get("CODE", pd.Series(dtype=str)).tolist()
        if clean_text(value)
    }
    row = {
        "INCLUDE": True,
        "CODE": _generated_submachinery_code(len(existing) + 1, used_codes),
        "NAME": "",
        "MAKER": clean_text(main_row.get("MAKER", "")),
        "MODEL": clean_text(main_row.get("MODEL", "")),
        "TYPE": "",
        "INSTR.BOOK": clean_text(main_row.get("INSTR.BOOK", "")),
        "SPECIFICATIONS": "",
        "MCH_TP(M/S/U)": "SubMachinery",
        "FIRST PAGE": None,
        "LAST PAGE": None,
        "PARTS FOUND": 0,
        "CONFIDENCE": 1.0,
        "VARIANTS": "",
        "DETECTION KEYS": "",
        "ORIGIN": "Manual",
    }
    new_frame = pd.DataFrame([row], columns=SUBMACHINERY_REVIEW_COLUMNS)
    if existing.empty:
        return new_frame
    return pd.concat([existing, new_frame], ignore_index=True)


def included_submachinery_rows(frame: pd.DataFrame | None) -> pd.DataFrame:
    if frame is None or frame.empty:
        return empty_additional_machinery_dataframe()
    working = frame.copy()
    if "INCLUDE" in working.columns:
        working = working[working["INCLUDE"].astype(bool)]
    for column in MACHINERY_COLUMNS:
        if column not in working.columns:
            working[column] = ""
    working["MCH_TP(M/S/U)"] = "SubMachinery"
    return working[MACHINERY_COLUMNS].reset_index(drop=True)


def submachinery_identity_keys(row: Any) -> set[str]:
    """Return stable, type-prefixed keys for one Step-3 proposal.

    CODE is preferred, while full detection aliases allow an exclusion to survive
    a corrected OCR code/name. Prefixes prevent a code from accidentally matching
    an unrelated textual alias.
    """
    getter = getattr(row, "get", lambda *_: "")
    code_key = normalize_key(getter("CODE", ""))
    name_key = normalize_key(getter("NAME", ""))
    detection_keys = _split_detection_keys(getter("DETECTION KEYS", ""))
    keys = {f"C:{code_key}"} if code_key else set()
    keys.update(f"D:{key}" for key in detection_keys if key)
    if not keys and name_key:
        keys.add(f"N:{name_key}")
    return keys


def apply_submachinery_exclusion_registry(
    frame: pd.DataFrame | None,
    excluded_keys: Sequence[Any] | set[Any] | None,
) -> pd.DataFrame:
    """Reapply explicit user exclusions after any automatic candidate rebuild."""
    result = (
        frame.copy()
        if frame is not None
        else empty_submachinery_review_dataframe()
    )
    if result.empty:
        return result
    registry = {
        clean_text(value).strip().upper()
        for value in (excluded_keys or [])
        if clean_text(value).strip()
    }
    if not registry:
        return result
    for index, row in result.iterrows():
        if submachinery_identity_keys(row) & registry:
            result.at[index, "INCLUDE"] = False
    result["INCLUDE"] = result["INCLUDE"].fillna(False).astype(bool)
    return result


def apply_submachinery_assignments(
    review_frame: pd.DataFrame,
    submachinery_frame: pd.DataFrame | None,
    main_machinery: str,
    overwrite_auto_assignments: bool = True,
) -> pd.DataFrame:
    """Assign spare rows by exact source section code, then by canonical name."""
    if review_frame is None or review_frame.empty:
        return empty_review_dataframe()
    result = review_frame.copy()
    for column in REVIEW_COLUMNS:
        if column not in result.columns:
            result[column] = False if column in {"INCLUDE", "READY"} else ""

    code_map: dict[str, str] = {}
    name_map: dict[str, str] = {}
    approved_names: set[str] = set()
    if submachinery_frame is not None and not submachinery_frame.empty:
        for _, candidate in submachinery_frame.iterrows():
            if not bool(candidate.get("INCLUDE", False)):
                continue
            target_name = clean_text(candidate.get("NAME", ""))
            if not target_name:
                continue
            approved_names.add(normalize_key(target_name))
            code_key = normalize_key(candidate.get("CODE", ""))
            if code_key:
                code_map[code_key] = target_name
            for key in _split_detection_keys(candidate.get("DETECTION KEYS", "")) | {normalize_key(target_name)}:
                if key:
                    name_map[key] = target_name

    main_key = normalize_key(main_machinery)
    for index, row in result.iterrows():
        section_code_key = normalize_key(row.get("SECTION CODE", ""))
        detected_key = normalize_key(row.get("DETECTED MACHINERY", row.get("TABLE TITLE", "")))
        target = code_map.get(section_code_key) or name_map.get(detected_key)
        current = clean_text(row.get("MACHINERY", ""))
        current_source = clean_text(row.get("ASSIGNMENT SOURCE", ""))
        current_is_auto = (
            not current
            or normalize_key(current) == main_key
            or normalize_key(current) not in approved_names | ({main_key} if main_key else set())
            or current_source.startswith("Auto")
            or current_source.startswith("Main")
        )
        if target and (overwrite_auto_assignments or current_is_auto):
            result.at[index, "MACHINERY"] = target
            result.at[index, "ASSIGNMENT SOURCE"] = "Auto-matched by section code"
        elif not current:
            result.at[index, "MACHINERY"] = clean_text(main_machinery)
            result.at[index, "ASSIGNMENT SOURCE"] = "Main machinery default"
        elif not current_source:
            result.at[index, "ASSIGNMENT SOURCE"] = "Manual assignment"
    return result[REVIEW_COLUMNS]


# ---------------------------------------------------------------------------
# Review dataframe construction
# ---------------------------------------------------------------------------


def rows_to_review_dataframe(
    rows: Sequence[dict[str, Any]],
    default_machinery: str,
    default_unit: str = "PCS",
) -> pd.DataFrame:
    normalized: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    for raw in rows:
        identifier = clean_text(raw.get("ident_no", raw.get("code", raw.get("CODE", raw.get("part_no", raw.get("PART NO", ""))))))
        description = clean_text(raw.get("description_english", raw.get("description", raw.get("DESCRIPTION", "")))).upper()
        item_no = clean_text(raw.get("item_no", raw.get("ITEM NO", "")))
        detected_machinery = _clean_machinery_name(raw.get("section_name_english", raw.get("detected_machinery", raw.get("DETECTED MACHINERY", ""))))
        section_code = clean_text(raw.get("section_code", raw.get("SECTION CODE", ""))).upper()
        section_maker = _credible_section_maker(raw.get("section_maker", raw.get("SECTION MAKER", "")))
        section_model = _credible_section_model(raw.get("section_model", raw.get("SECTION MODEL", "")))
        table_title = clean_text(raw.get("table_title", raw.get("TABLE TITLE", detected_machinery)))
        source_page = quantity_to_number(raw.get("source_page", raw.get("SOURCE PAGE")))
        source_page_int = int(source_page) if source_page is not None else None
        section_start = quantity_to_number(raw.get("section_start_page", raw.get("SECTION START PAGE")))
        section_start_int = int(section_start) if section_start is not None else source_page_int
        quantity = quantity_to_number(raw.get("quantity", raw.get("QNT")))
        unit = normalize_unit(raw.get("unit", raw.get("UNIT", "")), default_unit)
        confidence = clamp_confidence(raw.get("confidence", raw.get("CONFIDENCE", 0.70)))
        duplicate_key = (str(source_page_int or ""), normalize_key(identifier), normalize_key(item_no), normalize_key(description))
        if duplicate_key in seen:
            continue
        seen.add(duplicate_key)
        useful = bool(description and (identifier or item_no))
        source_warnings: list[str] = []
        if bool(raw.get("description_source_mismatch", False)):
            source_warnings.append(
                "Source: AI wording differed - exact printed English was kept"
            )
        if bool(raw.get("language_review", False)):
            source_warnings.append("Source: English wording was not fully confirmed")
        section_assignment_source = clean_text(raw.get("section_assignment_source", ""))
        if bool(raw.get("section_review", False)):
            if section_assignment_source == "unconfirmed":
                source_warnings.append("Source: section could not be confirmed")
            else:
                source_warnings.append(
                    "Source: AI/page section conflict - printed page header was kept"
                )
        normalized.append(
            {
                "INCLUDE": useful,
                "READY": False,
                "MACHINERY": clean_text(default_machinery),
                "PART NO": identifier,
                "DESCRIPTION": description,
                "CODE": identifier,
                "ITEM NO": item_no,
                "UNIT": unit,
                "QNT": quantity,
                "SOURCE PAGE": source_page_int,
                "SECTION START PAGE": section_start_int,
                "TABLE TITLE": table_title,
                "SECTION CODE": section_code,
                "SECTION MAKER": section_maker,
                "SECTION MODEL": section_model,
                "CONFIDENCE": confidence,
                "DETECTED MACHINERY": detected_machinery,
                "ASSIGNMENT SOURCE": "Main machinery default",
                "WARNING": "; ".join(source_warnings),
            }
        )
    if not normalized:
        return empty_review_dataframe()
    frame = pd.DataFrame(normalized, columns=REVIEW_COLUMNS)
    frame["INCLUDE"] = frame["INCLUDE"].astype(bool)
    frame["READY"] = frame["READY"].astype(bool)
    frame["QNT"] = pd.to_numeric(frame["QNT"], errors="coerce")
    frame["SOURCE PAGE"] = pd.to_numeric(frame["SOURCE PAGE"], errors="coerce").astype("Int64")
    frame["SECTION START PAGE"] = pd.to_numeric(frame["SECTION START PAGE"], errors="coerce").astype("Int64")
    frame["CONFIDENCE"] = pd.to_numeric(frame["CONFIDENCE"], errors="coerce").fillna(0.70)
    return frame


def merge_review_dataframes(existing: pd.DataFrame, new: pd.DataFrame) -> pd.DataFrame:
    if existing is None or existing.empty:
        return new.copy()
    if new is None or new.empty:
        return existing.copy()

    combined = pd.concat([existing, new], ignore_index=True)
    def merge_key(row: pd.Series) -> str:
        source_page = str(row.get("SOURCE PAGE", ""))
        section_code = normalize_key(row.get("SECTION CODE", ""))
        item_no = normalize_key(row.get("ITEM NO", ""))
        # A manual PART NO override must not create a duplicate when the same
        # page range is appended later. Source page + section + drawing position
        # is the stable row identity whenever a position exists.
        if item_no:
            return "|".join([source_page, section_code, "ITEM", item_no])
        return "|".join(
            [
                source_page,
                section_code,
                "IDENT",
                normalize_key(row.get("CODE", row.get("PART NO", ""))),
                normalize_key(row.get("DESCRIPTION", "")),
            ]
        )

    keys = combined.apply(merge_key, axis=1)
    merged = combined.loc[~keys.duplicated()].copy()

    # Cross-range append de-duplication for explicitly labelled spare-number lists.
    # A maintenance schedule can mention a spare first and a later recommended-spares
    # section can repeat the same identifier. Prefer the later/more complete row for
    # the same parent, absorb parent-less duplicates when a parent-backed row exists,
    # but preserve the same identifier under genuinely different parent components.
    explicit_mask = merged["TABLE TITLE"].fillna("").astype(str).str.lower().apply(
        lambda value: (
            "explicit spare-number" in value
            or "recommended spare parts" in value
        )
    )
    explicit_indexes = list(merged.index[explicit_mask])
    drop_indexes: set[Any] = set()
    groups: dict[str, list[Any]] = {}
    for index in explicit_indexes:
        identifier = normalize_key(
            merged.at[index, "CODE"] or merged.at[index, "PART NO"]
        )
        if identifier:
            groups.setdefault(identifier, []).append(index)
    for indexes in groups.values():
        if len(indexes) < 2:
            continue
        by_parent: dict[str, list[Any]] = {}
        blank_parent: list[Any] = []
        for index in indexes:
            parent = normalize_key(merged.at[index, "DETECTED MACHINERY"])
            if parent:
                by_parent.setdefault(parent, []).append(index)
            else:
                blank_parent.append(index)
        for parent_indexes in by_parent.values():
            # Existing rows are concatenated before new append rows, so keep the last
            # occurrence to allow a later recommended-spares source to improve an
            # earlier maintenance-schedule reference without creating a duplicate.
            drop_indexes.update(parent_indexes[:-1])
        if by_parent:
            drop_indexes.update(blank_parent)
        elif len(blank_parent) > 1:
            drop_indexes.update(blank_parent[:-1])

    if drop_indexes:
        merged = merged.drop(index=list(drop_indexes))
    return merged.reset_index(drop=True)


def machinery_rows_from_main_and_additional(
    main_row: dict[str, Any],
    additional: pd.DataFrame | None,
) -> pd.DataFrame:
    records = [{column: clean_text(main_row.get(column, "")) for column in MACHINERY_COLUMNS}]
    if additional is not None and not additional.empty:
        for _, row in additional.iterrows():
            if "INCLUDE" in additional.columns and not bool(row.get("INCLUDE", False)):
                continue
            record = {column: clean_text(row.get(column, "")) for column in MACHINERY_COLUMNS}
            if any(record.values()):
                record["NAME"] = clean_text(record.get("NAME", "")).upper()
                record["MAKER"] = clean_text(record.get("MAKER", "")).upper()
                record["MODEL"] = clean_text(record.get("MODEL", "")).upper()
                record["MCH_TP(M/S/U)"] = "SubMachinery"
                records.append(record)
    return pd.DataFrame(records, columns=MACHINERY_COLUMNS)


def linked_machinery_rows_for_export(
    machinery_frame: pd.DataFrame,
    review_frame: pd.DataFrame,
    *,
    ready_only: bool = False,
) -> pd.DataFrame:
    """Keep the main machinery and only sub-machineries used by spare rows.

    Drawing/equipment candidates remain useful review and audit evidence, but the
    import workbook is a strict parent-child hierarchy: an exported sub-machinery
    must own at least one included spare-part row.
    """
    if machinery_frame is None or machinery_frame.empty:
        return pd.DataFrame(columns=MACHINERY_COLUMNS)
    machinery_frame = machinery_frame.copy()
    for column in MACHINERY_COLUMNS:
        if column not in machinery_frame.columns:
            machinery_frame[column] = ""

    selected = (
        review_frame.copy()
        if review_frame is not None
        else pd.DataFrame(columns=REVIEW_COLUMNS)
    )
    if not selected.empty and "INCLUDE" in selected.columns:
        selected = selected[selected["INCLUDE"].astype(bool)]
    if ready_only and not selected.empty and "READY" in selected.columns:
        selected = selected[selected["READY"].astype(bool)]

    linked_names = {
        normalize_key(value)
        for value in selected.get("MACHINERY", pd.Series(dtype=str)).tolist()
        if clean_text(value)
    }
    machinery_types = machinery_frame.get(
        "MCH_TP(M/S/U)", pd.Series("", index=machinery_frame.index)
    ).map(clean_text)
    machinery_names = machinery_frame.get(
        "NAME", pd.Series("", index=machinery_frame.index)
    ).map(normalize_key)
    keep = machinery_types.eq("Main Machinery") | (
        machinery_types.eq("SubMachinery") & machinery_names.isin(linked_names)
    )
    return machinery_frame.loc[keep, MACHINERY_COLUMNS].reset_index(drop=True)


def validate_spare_machinery_hierarchy(
    machinery_frame: pd.DataFrame,
    review_frame: pd.DataFrame,
) -> list[str]:
    """Validate the strict Main -> Sub-machinery -> Spare relationship."""
    errors: list[str] = []
    selected = (
        review_frame.copy()
        if review_frame is not None
        else pd.DataFrame(columns=REVIEW_COLUMNS)
    )
    if not selected.empty and "INCLUDE" in selected.columns:
        selected = selected[selected["INCLUDE"].astype(bool)]
    if selected.empty:
        return [
            "At least one included spare-part row is required; machinery without "
            "linked spare parts remains audit-only."
        ]

    frame = (
        machinery_frame.copy()
        if machinery_frame is not None
        else pd.DataFrame(columns=MACHINERY_COLUMNS)
    )
    for column in MACHINERY_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    machinery_types = frame.get(
        "MCH_TP(M/S/U)", pd.Series("", index=frame.index)
    ).map(clean_text)
    main_names = {
        normalize_key(value)
        for value in frame.loc[machinery_types.eq("Main Machinery"), "NAME"].tolist()
        if clean_text(value)
    }
    sub_names = {
        normalize_key(value)
        for value in frame.loc[machinery_types.eq("SubMachinery"), "NAME"].tolist()
        if clean_text(value)
    }
    parent_labels: dict[str, str] = {}
    for value in selected.get("MACHINERY", pd.Series(dtype=str)).tolist():
        label = clean_text(value)
        if label:
            parent_labels.setdefault(normalize_key(label), label)
    parent_names = set(parent_labels)

    direct_main = sorted(parent_names & main_names)
    if direct_main:
        errors.append(
            "Every included spare part must belong to a sub-machinery, not directly "
            "to the main machinery: "
            + ", ".join(parent_labels[key] for key in direct_main)
            + "."
        )

    missing_parents = sorted(parent_names - sub_names - main_names)
    if missing_parents:
        errors.append(
            "Included spare rows reference missing sub-machineries: "
            + ", ".join(parent_labels[key] for key in missing_parents)
            + "."
        )

    unlinked_subs = sorted(sub_names - parent_names)
    if unlinked_subs:
        display_by_key = {
            normalize_key(row.get("NAME", "")): clean_text(row.get("NAME", ""))
            for _, row in frame.loc[machinery_types.eq("SubMachinery")].iterrows()
        }
        errors.append(
            "Sub-machineries without included spare parts cannot be exported: "
            + ", ".join(display_by_key[key] for key in unlinked_subs)
            + "."
        )

    if len(sub_names) > len(selected):
        errors.append(
            f"Sub-machinery count ({len(sub_names)}) cannot exceed included spare-part "
            f"count ({len(selected)})."
        )
    return errors


def validate_machinery_dataframe(frame: pd.DataFrame) -> list[str]:
    errors: list[str] = []
    if frame is None or frame.empty:
        return ["At least one main machinery row is required."]
    if len(frame) > MAX_MACHINERY_ROWS:
        errors.append(
            f"The template supports at most {MAX_MACHINERY_ROWS} machinery rows."
        )

    required = ["CODE", "NAME", "MAKER", "MODEL", "MCH_TP(M/S/U)"]
    for index, row in frame.reset_index(drop=True).iterrows():
        excel_row = index + 5
        for column in required:
            if not clean_text(row.get(column, "")):
                errors.append(f"Machinery sheet row {excel_row}: {column} is required.")
        machinery_type = clean_text(row.get("MCH_TP(M/S/U)", ""))
        if machinery_type and machinery_type not in MACHINERY_TYPES:
            errors.append(
                f"Machinery sheet row {excel_row}: invalid MCH_TP value '{machinery_type}'."
            )

    main_count = sum(
        clean_text(value) == "Main Machinery"
        for value in frame["MCH_TP(M/S/U)"].tolist()
    )
    if main_count != 1:
        errors.append("The workbook must contain exactly one Main Machinery row.")

    names = [normalize_key(value) for value in frame["NAME"].tolist() if clean_text(value)]
    duplicate_names = {name for name in names if names.count(name) > 1}
    if duplicate_names:
        errors.append("Machinery NAME values must be unique.")

    codes = [normalize_key(value) for value in frame["CODE"].tolist() if clean_text(value)]
    duplicate_codes = {code for code in codes if codes.count(code) > 1}
    if duplicate_codes:
        errors.append("Machinery CODE values must be unique.")
    return errors


def recalculate_review_status(
    frame: pd.DataFrame,
    valid_machinery_names: Sequence[str],
    allow_duplicates: bool = False,
) -> pd.DataFrame:
    if frame is None or frame.empty:
        return empty_review_dataframe()
    result = frame.copy()
    for column in REVIEW_COLUMNS:
        if column not in result.columns:
            result[column] = False if column in {"INCLUDE", "READY"} else ""

    # Benefit automation rules: descriptions are uppercase and Ident/Code is copied
    # to both CODE and PART NO unless a reviewer intentionally changed PART NO.
    result["DESCRIPTION"] = result["DESCRIPTION"].map(lambda value: clean_text(value).upper())
    for index, row in result.iterrows():
        code = clean_text(row.get("CODE", ""))
        part_no = clean_text(row.get("PART NO", ""))
        if code and not part_no:
            result.at[index, "PART NO"] = code
        elif part_no and not code:
            result.at[index, "CODE"] = part_no

    valid_names = {normalize_key(name) for name in valid_machinery_names if clean_text(name)}
    included_indexes = [index for index, value in result["INCLUDE"].items() if bool(value)]
    duplicate_counter: dict[tuple[str, str, str, str], int] = {}
    for index in included_indexes:
        row = result.loc[index]
        key = (
            normalize_key(row.get("MACHINERY", "")),
            normalize_key(row.get("PART NO", "")),
            normalize_key(row.get("ITEM NO", "")),
            normalize_key(row.get("DESCRIPTION", "")),
        )
        duplicate_counter[key] = duplicate_counter.get(key, 0) + 1

    warnings: list[str] = []
    ready_values: list[bool] = []
    for index, row in result.iterrows():
        if not bool(row.get("INCLUDE", False)):
            warnings.append("Excluded from export")
            ready_values.append(False)
            continue
        existing_warning = clean_text(row.get("WARNING", ""))
        row_messages: list[str] = [
            message.strip()
            for message in existing_warning.split(";")
            if message.strip().startswith("Source:")
        ]
        blocking = False
        machinery = clean_text(row.get("MACHINERY", ""))
        part_no = clean_text(row.get("PART NO", ""))
        description = clean_text(row.get("DESCRIPTION", ""))
        item_no = clean_text(row.get("ITEM NO", ""))
        unit = clean_text(row.get("UNIT", "")).upper()
        if not machinery:
            row_messages.append("Missing sub-machinery")
            blocking = True
        elif normalize_key(machinery) not in valid_names:
            row_messages.append("Sub-machinery is not present on sheet 1")
            blocking = True
        if not description:
            row_messages.append("Missing English description")
            blocking = True
        if not part_no and not item_no:
            row_messages.append("Part No or Item No is required")
            blocking = True
        if unit not in UNIT_OPTIONS:
            row_messages.append("Unit must be blank, PCS, or SET")
            blocking = True
        raw_quantity = row.get("QNT")
        if clean_text(raw_quantity) and quantity_to_number(raw_quantity) is None:
            row_messages.append("Quantity is not numeric")
            blocking = True
        duplicate_key = (
            normalize_key(machinery), normalize_key(part_no), normalize_key(item_no), normalize_key(description)
        )
        if duplicate_counter.get(duplicate_key, 0) > 1:
            row_messages.append("Possible duplicate")
            if not allow_duplicates:
                blocking = True
        confidence = clamp_confidence(row.get("CONFIDENCE", 0.70))
        # OCR confidence is advisory only. A user-corrected row must become READY
        # when all required Benefit fields and machinery links are valid.
        if confidence < 0.65:
            row_messages.append("Very low OCR confidence - manually verified fields recommended")
        elif confidence < 0.80:
            row_messages.append("Low OCR confidence - verify when practical")
        warnings.append("; ".join(row_messages))
        ready_values.append(not blocking)

    result["WARNING"] = warnings
    result["READY"] = ready_values
    result["QNT"] = pd.to_numeric(result["QNT"], errors="coerce")
    result["SOURCE PAGE"] = pd.to_numeric(result["SOURCE PAGE"], errors="coerce").astype("Int64")
    result["SECTION START PAGE"] = pd.to_numeric(result["SECTION START PAGE"], errors="coerce").astype("Int64")
    result["CONFIDENCE"] = pd.to_numeric(result["CONFIDENCE"], errors="coerce").fillna(0.70)
    return result[REVIEW_COLUMNS]


# ---------------------------------------------------------------------------
# Benefit template and audit workbook generation
# ---------------------------------------------------------------------------


def _clear_values(worksheet: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def build_benefit_workbook(
    template_bytes: bytes,
    machinery_frame: pd.DataFrame,
    review_frame: pd.DataFrame,
    clear_existing: bool = True,
) -> bytes:
    selected = review_frame[
        review_frame["INCLUDE"].astype(bool) & review_frame["READY"].astype(bool)
    ].copy()
    if len(selected) > MAX_SPARE_ROWS:
        raise ValueError(f"Too many spare-parts rows; maximum is {MAX_SPARE_ROWS}.")

    export_machinery_frame = linked_machinery_rows_for_export(
        machinery_frame,
        selected,
    )
    hierarchy_errors = validate_spare_machinery_hierarchy(
        export_machinery_frame,
        selected,
    )
    if hierarchy_errors:
        raise ValueError(" ".join(hierarchy_errors))
    if len(export_machinery_frame) > MAX_MACHINERY_ROWS:
        raise ValueError(f"Too many machinery rows; maximum is {MAX_MACHINERY_ROWS}.")

    workbook = load_workbook(io.BytesIO(template_bytes), data_only=False, keep_links=True)
    missing_sheets = [
        name for name in (MACHINERY_SHEET, SPARE_PARTS_SHEET) if name not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "The selected workbook is not the expected import template. Missing sheets: "
            + ", ".join(missing_sheets)
        )

    machinery_sheet = workbook[MACHINERY_SHEET]
    spare_sheet = workbook[SPARE_PARTS_SHEET]

    if clear_existing:
        _clear_values(machinery_sheet, 5, 609, 1, 8)
        _clear_values(spare_sheet, 4, 1441, 1, 7)

    for offset, (_, row) in enumerate(export_machinery_frame.iterrows(), start=5):
        for column_index, column_name in enumerate(MACHINERY_COLUMNS, start=1):
            cell = machinery_sheet.cell(row=offset, column=column_index)
            cell.value = excel_safe_text(row.get(column_name, ""))
            cell.number_format = "@"

    for offset, (_, row) in enumerate(selected.iterrows(), start=4):
        values = [
            row.get("MACHINERY", ""),
            row.get("PART NO", ""),
            row.get("DESCRIPTION", ""),
            row.get("CODE", ""),
            row.get("ITEM NO", ""),
            normalize_unit(row.get("UNIT", ""), ""),
            quantity_to_number(row.get("QNT")),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = spare_sheet.cell(row=offset, column=column_index)
            if column_index == 7:
                cell.value = value
                cell.number_format = "0.###"
            else:
                cell.value = excel_safe_text(value)
                cell.number_format = "@"

    # The source template may arrive with sheet protection enabled. Review/import
    # workbooks must remain freely editable after export, including the Instructions
    # sheet and zero-spare exports that contain machinery rows only.
    for worksheet in workbook.worksheets:
        worksheet.protection.sheet = False

    workbook.active = workbook.sheetnames.index(SPARE_PARTS_SHEET)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_content_width(frame: pd.DataFrame, column: str, minimum: int = 10, maximum: int = 60) -> int:
    if column not in frame.columns:
        return minimum
    values = [clean_text(column)] + [clean_text(value) for value in frame[column].head(1000).tolist()]
    longest = max((len(value) for value in values), default=minimum)
    return max(minimum, min(maximum, int(longest * 1.05) + 2))


def build_audit_workbook(
    extracted_pages: Sequence[tuple[int, str]],
    machinery_frame: pd.DataFrame,
    review_frame: pd.DataFrame,
    page_classification: pd.DataFrame | None = None,
    extraction_log: Sequence[str] | None = None,
    vessels: Sequence[str] | None = None,
    submachinery_review: pd.DataFrame | None = None,
    job_metadata: dict[str, Any] | None = None,
) -> bytes:
    output = io.BytesIO()
    pages_frame = pd.DataFrame(extracted_pages, columns=["SOURCE PAGE", "OCR MARKDOWN"])
    classification_frame = (
        page_classification.copy()
        if page_classification is not None and not page_classification.empty
        else pd.DataFrame(columns=PAGE_CLASSIFICATION_COLUMNS)
    )
    log_frame = pd.DataFrame({"MESSAGE": list(extraction_log or [])})
    sub_frame = (
        submachinery_review.copy()
        if submachinery_review is not None and not submachinery_review.empty
        else empty_submachinery_review_dataframe()
    )

    summary_records: list[dict[str, str]] = [
        {
            "FIELD": "Generated at (UTC)",
            "VALUE": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        },
        {
            "FIELD": "Vessels",
            "VALUE": ", ".join(clean_text(value) for value in (vessels or []) if clean_text(value)),
        },
    ]
    for key, value in (job_metadata or {}).items():
        summary_records.append({"FIELD": clean_text(key), "VALUE": clean_text(value)})
    summary_frame = pd.DataFrame(summary_records, columns=["FIELD", "VALUE"])

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        summary_frame.to_excel(writer, index=False, sheet_name="Job Summary")
        pages_frame.to_excel(writer, index=False, sheet_name="OCR Pages")
        classification_frame.to_excel(writer, index=False, sheet_name="Page Classification")
        machinery_frame.to_excel(writer, index=False, sheet_name="Machinery Review")
        sub_frame.to_excel(writer, index=False, sheet_name="Sub-machinery Review")
        review_frame.to_excel(writer, index=False, sheet_name="Spare Parts Review")
        log_frame.to_excel(writer, index=False, sheet_name="Extraction Log")

        workbook = writer.book
        header_format = workbook.add_format(
            {"bold": True, "bg_color": "#D9EAF7", "border": 1}
        )
        wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})

        sheet_frames = (
            ("Job Summary", summary_frame),
            ("OCR Pages", pages_frame),
            ("Page Classification", classification_frame),
            ("Machinery Review", machinery_frame),
            ("Sub-machinery Review", sub_frame),
            ("Spare Parts Review", review_frame),
            ("Extraction Log", log_frame),
        )
        for sheet_name, frame in sheet_frames:
            sheet = writer.sheets[sheet_name]
            for column_index, column_name in enumerate(frame.columns):
                sheet.write(0, column_index, column_name, header_format)
                maximum = 100 if column_name in {"OCR MARKDOWN", "MESSAGE"} else 60
                width = _xlsx_content_width(frame, column_name, minimum=10, maximum=maximum)
                sheet.set_column(column_index, column_index, width, wrap_format if width >= 35 else None)
            sheet.freeze_panes(1, 0)
            if len(frame.columns):
                sheet.autofilter(0, 0, max(1, len(frame)), len(frame.columns) - 1)

        writer.sheets["Job Summary"].set_column(0, 0, 26)
        writer.sheets["Job Summary"].set_column(1, 1, 90, wrap_format)
        writer.sheets["OCR Pages"].set_column(0, 0, 12)
        writer.sheets["OCR Pages"].set_column(1, 1, 100, wrap_format)
        writer.sheets["Page Classification"].set_column(0, 3, 18)
        writer.sheets["Page Classification"].set_column(4, 4, 70, wrap_format)
        writer.sheets["Machinery Review"].set_column(0, len(MACHINERY_COLUMNS) - 1, 22)
        writer.sheets["Sub-machinery Review"].set_column(
            0, max(0, len(SUBMACHINERY_REVIEW_COLUMNS) - 1), 20
        )
        writer.sheets["Sub-machinery Review"].set_column(
            SUBMACHINERY_REVIEW_COLUMNS.index("VARIANTS"),
            SUBMACHINERY_REVIEW_COLUMNS.index("VARIANTS"),
            55,
            wrap_format,
        )
        writer.sheets["Spare Parts Review"].set_column(0, len(REVIEW_COLUMNS) - 1, 20)
        writer.sheets["Spare Parts Review"].set_column(
            REVIEW_COLUMNS.index("DESCRIPTION"),
            REVIEW_COLUMNS.index("DESCRIPTION"),
            50,
            wrap_format,
        )
        writer.sheets["Spare Parts Review"].set_column(
            REVIEW_COLUMNS.index("TABLE TITLE"),
            REVIEW_COLUMNS.index("TABLE TITLE"),
            40,
            wrap_format,
        )
        writer.sheets["Spare Parts Review"].set_column(
            REVIEW_COLUMNS.index("WARNING"),
            REVIEW_COLUMNS.index("WARNING"),
            45,
            wrap_format,
        )
        writer.sheets["Extraction Log"].set_column(0, 0, 100, wrap_format)

    return output.getvalue()


# Backward-compatible name used by app.py.
def build_workbook(
    template_bytes: bytes,
    machinery_frame: pd.DataFrame,
    review_frame: pd.DataFrame,
    clear_existing: bool = True,
) -> bytes:
    return build_benefit_workbook(
        template_bytes=template_bytes,
        machinery_frame=machinery_frame,
        review_frame=review_frame,
        clear_existing=clear_existing,
    )


def safe_filename(value: str, fallback: str = "spare_parts") -> str:
    stem = Path(clean_text(value)).stem
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return stem or fallback
